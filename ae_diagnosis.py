#!/usr/bin/env python3
"""
ae_diagnosis.py — root-cause diagnosis + recommender engine.

Strings together every report we have for a site and proposes a diagnosis:

  evidence sources
  ----------------
  - alert history        (live, classified by ae_alert_dashboard)
  - rule tool results    ae_ruleresults.xlsx  (comm / performance pass-fail)
  - AI site summaries    ae_ai_summaries.xlsx (natural-language health text)
  - live power + capacity ae_sites.xlsx       (is the site producing right now?)
  - hardware inventory   ae_hardware.xlsx     (inverter / tracker counts)

The rule base mirrors FAILURE_ROOT_CAUSES.md. Each rule looks for a
co-occurrence pattern (e.g. ground-fault keyword + same inverter offline
afterwards) and emits a diagnosis with confidence, evidence trail,
recommended action, and capacity at risk (kW) so the recommender can rank.

Used by ae_alert_dashboard.py; can also run standalone:
    python ae_diagnosis.py            # prints portfolio diagnoses to console
"""

import re
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

HERE = Path(__file__).parent

# ── keyword banks ─────────────────────────────────────────────────────────

KW = {
    "ground_fault": r"ground fault|gfdi|riso|isolation fault|insulation|leakage|earth fault",
    "grid":         r"islanding|grid power outage|recloser|grid under|grid over|"
                    r"underfrequency|overfrequency|under frequency|over frequency|"
                    r"phase lock loop|low voltage \d|lv\d|grid ?freq|breaker status|anti-island",
    "dc_overvolt":  r"bus hardware overvoltage|dc overvoltage|overvoltage fault",
    "thermal":      r"over ?temperature|thermal|overheat|temperature greater",
    "stop":         r"\bstop\b|stopped|shut ?down|standby",
    "tracker_soc":  r"state of charge|soc",
    "tracker_nocom":r"no communication",
    "tcu_fault":    r"tcu fault|system monitor|out of range",
    "meter_phase":  r"power phase error|phase error",
    "ups":          r"battery (over temperature|voltage low|low)|ups",
    "transformer":  r"transformer|liquid temp",
    "perf":         r"performance index|below limit|irradiance",
    "fan":          r"fan (fault|failure)|cooling",
    "arc":          r"arc fault|afci",
    "fuse":         r"fuse",
}

def _has(rx_key, *texts):
    rx = re.compile(KW[rx_key], re.I)
    return any(rx.search(t or "") for t in texts)


CONF_LABEL = {3: "High", 2: "Medium", 1: "Low"}

# negative phrases in AI summaries that corroborate an issue
AI_NEG = re.compile(r"offline|not communicat|fault|below|under ?perform|stopp|"
                    r"no production|issue|fail", re.I)


# ── evidence assembly ─────────────────────────────────────────────────────

def load_xlsx_evidence():
    """Read cached report files -> per-site dicts (best effort)."""
    import openpyxl
    rules, ai, power, caps = {}, {}, {}, {}

    f = HERE / "ae_ruleresults.xlsx"
    if f.exists():
        ws = openpyxl.load_workbook(f, read_only=True)["Site Pass-Fail"]
        it = ws.values; hdr = {h: i for i, h in enumerate(next(it))}
        for r in it:
            rules[r[hdr["Site Name"]]] = {
                "overall":   r[hdr["Overall"]],
                "comm_fail": int(r[hdr["Comm Fail"]] or 0),
                "perf_fail": int(r[hdr["Perf Fail"]] or 0),
                "weather":   r[hdr["Weather Condition"]],
                "run_at":    r[hdr["Run At"]],
            }

    f = HERE / "ae_ai_summaries.xlsx"
    if f.exists():
        ws = openpyxl.load_workbook(f, read_only=True)["AI Summaries"]
        it = ws.values; hdr = {h: i for i, h in enumerate(next(it))}
        for r in it:
            ai[r[hdr["Site Name"]]] = str(r[hdr["AI Summary (Plain Text)"]] or "")

    f = HERE / "ae_sites.xlsx"
    if f.exists():
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb["Live Power"]
        it = ws.values; hdr = {h: i for i, h in enumerate(next(it))}
        for r in it:
            power[r[hdr["Site Name"]]] = {
                "now_kw":  float(r[hdr["Current Power (kW)"]] or 0),
                "exp_kw":  float(r[hdr["Expected (kW)"]] or 0),
                "last_up": r[hdr["Last Upload"]],
            }
        ws = wb["Sites Overview"]
        it = ws.values; hdr = {h: i for i, h in enumerate(next(it))}
        for r in it:
            try:
                cap = float(r[hdr["Capacity AC (kW)"]] or 0)
            except (ValueError, TypeError):
                cap = 0
            if not cap:
                try:  # fall back: daily estimate / 5 sun-hours
                    cap = float(r[hdr["Daily Est (kWh)"]] or 0) / 5.0
                except (ValueError, TypeError):
                    cap = 0
            caps[r[hdr["Site Name"]]] = round(cap, 1)

    return rules, ai, power, caps


# ── core diagnosis rules ──────────────────────────────────────────────────

def diagnose_site(site, site_alerts, hw, rules, ai_text, pwr, cap_kw):
    """Return list of diagnosis dicts for one site."""
    out = []
    inv_total = max(len(hw.get(site, {}).get("inverters", [])), 1)

    inv_alerts = [a for a in site_alerts if a["category"] in
                  ("INVERTER_FAULT", "INVERTER_COMM")]
    trk_alerts = [a for a in site_alerts if a["category"].startswith("TRACKER")]
    open_alerts = [a for a in site_alerts if not a["is_resolved"]]

    def texts(alerts):
        return [f"{a['event_type']} {a['description']}" for a in alerts]

    def ev_alert(a):
        t = a["start"].strftime("%m-%d %H:%M") if a["start"] else "?"
        state = "OPEN" if not a["is_resolved"] else f"{a['duration_h']:.1f}h, resolved"
        return (f"[alert {t}] {a['hardware_name'] or 'site'}: "
                f"{a['event_type']} — {(a['description'] or '')[:90]} ({state})")

    def ai_corroboration():
        if ai_text and AI_NEG.search(ai_text):
            m = AI_NEG.search(ai_text)
            s = max(0, m.start() - 60); e = min(len(ai_text), m.end() + 80)
            return f"[AI summary] “…{ai_text[s:e].strip()}…”"
        return None

    def rule_corroboration(kind):
        r = rules.get(site)
        if not r:
            return None
        if kind == "comm" and r["comm_fail"] > 0:
            return f"[rule tool] {r['comm_fail']} communication check(s) FAILING"
        if kind == "perf" and r["perf_fail"] > 0:
            return (f"[rule tool] {r['perf_fail']} performance check(s) FAILING "
                    f"(weather: {r['weather']})")
        return None

    def add(key, title, conf, evidence, action, priority):
        # priority = editorial severity-class rank per diagnosis type (not a measured
        # quantity, never displayed) — used only to order the triage list. Ordering is
        # by real signals: is it an open alert, confidence from corroborating evidence,
        # severity class, and how many alerts corroborate it.
        evidence = [e for e in evidence if e]
        out.append({
            "site": site, "key": key, "title": title,
            "confidence": conf, "conf_label": CONF_LABEL[conf],
            "evidence": evidence, "action": action, "priority": priority,
            "open": any("OPEN" in e for e in evidence),
        })

    # 1 ── GROUND FAULT → inverter offline ────────────────────────────────
    gf = [a for a in inv_alerts if _has("ground_fault", a["event_type"], a["description"])]
    if gf:
        gf_hw = {a["hardware_name"] for a in gf}
        followers = [a for a in inv_alerts
                     if a["hardware_name"] in gf_hw and a is not gf[0] and
                     (_has("stop", a["event_type"], a["description"]) or
                      a["category"] == "INVERTER_COMM")]
        conf = 3 if followers else 2
        add("GROUND_FAULT",
            "Ground fault tripping inverter offline",
            conf,
            [ev_alert(a) for a in gf[:3]] + [ev_alert(a) for a in followers[:2]]
            + [rule_corroboration("comm"), ai_corroboration()],
            "Do NOT blind-reset. Dispatch electrician: insulation-resistance (megger) "
            "test the DC strings, inspect combiner boxes/wiring for water ingress or "
            "rodent damage, then clear GFDI and re-energize.",
            100)

    # 2 ── GRID EVENT: simultaneous multi-inverter trip or explicit grid alarms ─
    grid = [a for a in site_alerts if _has("grid", a["event_type"], a["description"])]
    simultaneous = []
    faults_sorted = sorted([a for a in inv_alerts if a["start"]], key=lambda a: a["start"])
    for i, a in enumerate(faults_sorted):
        cluster = [b for b in faults_sorted
                   if abs((b["start"] - a["start"]).total_seconds()) <= 1800]
        if len({b["hardware_name"] for b in cluster}) >= 3:
            simultaneous = cluster
            break
    if grid or simultaneous:
        conf = 3 if (grid and simultaneous) else 2
        n_inv = len({a['hardware_name'] for a in simultaneous}) if simultaneous else 0
        add("GRID_EVENT",
            "Grid disturbance / utility event tripped inverters (not equipment failure)",
            conf,
            [ev_alert(a) for a in grid[:3]] +
            ([f"[pattern] {n_inv} inverters faulted within 30 min of each other"]
             if simultaneous else []) + [ai_corroboration()],
            "Confirm utility outage/voltage event for the window; check recloser/breaker "
            "closed; verify inverters auto-restarted (production back to expected). "
            "Truck roll only if any unit failed to reconnect.",
            95)

    # 3 ── GATEWAY/DATALOGGER DOWN: site-wide comm loss ────────────────────
    comm = [a for a in site_alerts if a["category"] in
            ("INVERTER_COMM", "TRACKER_COMM", "COMMS_LOW")]
    gw = [a for a in site_alerts if _has_any(a, r"gateway heartbeat|site controller|datalogger")]
    comm_hw = {a["hardware_name"] for a in comm if a["hardware_name"]}
    site_dev_total = max(inv_total + len(hw.get(site, {}).get("trackers", [])), 1)
    if gw or len(comm_hw) >= max(3, int(0.8 * site_dev_total)):
        conf = 3 if gw and len(comm_hw) >= 3 else 2
        add("GATEWAY_DOWN",
            "Datalogger / gateway failure — site reporting dark (production may continue)",
            conf,
            [ev_alert(a) for a in gw[:2]] +
            [f"[pattern] {len(comm_hw)} of ~{site_dev_total} devices in comm loss"]
            + [rule_corroboration("comm"), ai_corroboration()],
            "Remote-reboot the datalogger / cell modem. If telemetry doesn't return in "
            "~4h, send tech to power-cycle; check cellular signal and modem data contract "
            "(contract expiry causes permanent disconnect).",
            55)

    # 4 ── RECURRING SINGLE-INVERTER HARDWARE FAULT ────────────────────────
    by_hw = defaultdict(list)
    for a in inv_alerts:
        if a["category"] == "INVERTER_FAULT" and a["hardware_name"]:
            by_hw[a["hardware_name"]].append(a)
    for hw_name, lst in by_hw.items():
        if _has("ground_fault", *texts(lst)):
            continue  # covered by rule 1
        long_open = [a for a in lst if not a["is_resolved"] and a["duration_h"] > 24]
        if len(lst) >= 3 or long_open:
            conf = 3 if long_open else 2
            add("INVERTER_HW",
                f"Recurring hardware fault on {hw_name} (board/IGBT/capacitor suspect)",
                conf,
                [f"[pattern] {len(lst)} fault alerts on this inverter in window"]
                + [ev_alert(a) for a in lst[:3]] + [ai_corroboration()],
                "Pull the inverter's internal event log, check warranty status, and "
                "schedule component or unit replacement. Same-model faults at other "
                "sites would indicate a batch defect.",
                85)

    # 5 ── DC OVERVOLTAGE ─────────────────────────────────────────────────
    dcov = [a for a in inv_alerts if _has("dc_overvolt", a["event_type"], a["description"])]
    if dcov:
        add("DC_OVERVOLT",
            "DC bus overvoltage — string sizing / cold-temp Voc or measurement issue",
            2,
            [ev_alert(a) for a in dcov[:3]],
            "Check string Voc against inverter max input at low ambient temps; verify "
            "no recent re-string; inspect DC surge protection. Recurs in cold mornings → "
            "string config issue.",
            70)

    # 6 ── THERMAL ────────────────────────────────────────────────────────
    th = [a for a in site_alerts if _has("thermal", a["event_type"], a["description"])
          and a["category"] != "EQUIPMENT"]
    if th:
        add("THERMAL",
            "Thermal derate/trip — cooling restriction suspected",
            2,
            [ev_alert(a) for a in th[:3]],
            "Clean inverter air filters/heat-sink fins, verify fans spin up, clear "
            "vegetation/obstructions around enclosures. Persistent → fan replacement.",
            60)

    # 7 ── TRACKER: battery / comm / TCU faults ───────────────────────────
    if trk_alerts:
        soc = [a for a in trk_alerts if _has("tracker_soc", a["description"])]
        nocom = [a for a in trk_alerts if _has("tracker_nocom", a["description"])]
        tcu = [a for a in trk_alerts if _has("tcu_fault", a["description"])]
        if soc:
            add("TRACKER_BATTERY",
                "Tracker TCU battery degraded (SOC below critical threshold)",
                3 if len(soc) > 2 else 2,
                [ev_alert(a) for a in soc[:3]],
                "Replace TCU batteries on the flagged rows; verify charge controller / "
                "panel on the TCU. Stuck trackers off-angle bleed production daily.",
                65)
        if nocom:
            n = len(set(re.findall(r"(?:Alarm|S)\s*(\d+)", " ".join(texts(nocom)))))
            add("TRACKER_MESH",
                f"Tracker comm mesh outage (~{max(n,len(nocom))} units not reporting)",
                2,
                [ev_alert(a) for a in nocom[:3]],
                "Check tracker network controller / repeater power and antennas. If units "
                "are stowed flat meanwhile, expect measurable production loss on clear days.",
                55)
        if tcu:
            add("TCU_FAULT",
                "TCU hardware fault (system monitor out-of-range)",
                2,
                [ev_alert(a) for a in tcu[:3]],
                "Inspect flagged TCUs: check 24V supply, motor current draw, and limit "
                "switches; clear fault and observe tracking through a full day.",
                60)

    # 8 ── METER MISREAD ──────────────────────────────────────────────────
    mtr = [a for a in site_alerts if a["category"] == "METER"]
    if mtr:
        add("METER",
            "Production meter / CT measurement issue — revenue data unreliable",
            2,
            [ev_alert(a) for a in mtr[:3]] + [rule_corroboration("perf")],
            "Compare meter kW vs sum of inverter kW. Phase reading ~0 on one leg → "
            "loose CT or wiring. Fix before settlement period closes; backfill from "
            "inverter data if needed.",
            65)

    # 9 ── UNDERPERFORMANCE (no faults to explain it) ─────────────────────
    perf = [a for a in site_alerts if a["category"] == "PERFORMANCE"]
    r = rules.get(site, {})
    if (perf or r.get("perf_fail", 0) > 0) and not inv_alerts:
        add("UNDERPERF",
            "Underperformance without equipment faults — soiling/shading/baseline suspect",
            1 if (r.get("weather") in ("Overcast", "Rain")) else 2,
            [ev_alert(a) for a in perf[:2]] + [rule_corroboration("perf"),
             ai_corroboration()],
            "Cross-check irradiance vs output. Clear-day deficit → schedule module wash "
            "/ vegetation survey; deficit only vs model → recalibrate expected baseline.",
            40)

    # 10 ── BALANCE OF PLANT (UPS / transformer) ──────────────────────────
    ups = [a for a in site_alerts if _has("ups", a["event_type"], a["description"])]
    if ups:
        add("UPS",
            "UPS battery end-of-life — datalogger will drop on next outage",
            2,
            [ev_alert(a) for a in ups[:2]],
            "Replace UPS battery module. Cheap fix that prevents site-dark events "
            "during grid blips.",
            35)
    tx = [a for a in site_alerts if _has("transformer", a["event_type"], a["description"])]
    if tx:
        add("TRANSFORMER",
            "Transformer over-temperature fault",
            3 if any(not a["is_resolved"] for a in tx) else 2,
            [ev_alert(a) for a in tx[:2]],
            "Verify cooling (radiators/fans), oil level and load profile. Sustained "
            ">90°C liquid temp degrades insulation — consider derate until inspected.",
            90)

    # live-power sanity: site producing nothing right now during day?
    p = pwr.get(site)
    if p and out:
        for d in out:
            if p["now_kw"] <= 0.1 and cap_kw > 0:
                d["evidence"].append(
                    f"[live power] site output now: {p['now_kw']:.1f} kW "
                    f"(capacity {cap_kw:.0f} kW)")
    return out


def _has_any(a, pattern):
    rx = re.compile(pattern, re.I)
    return bool(rx.search(a["event_type"] or "") or rx.search(a["description"] or ""))


# ── portfolio-level recommender ───────────────────────────────────────────

def diagnose_portfolio(alerts, hw):
    """alerts: enriched alert dicts (from ae_alert_dashboard). Returns ranked list."""
    rules, ai, power, caps = load_xlsx_evidence()
    by_site = defaultdict(list)
    for a in alerts:
        by_site[a["site_name"]].append(a)

    all_diag = []
    for site, site_alerts in by_site.items():
        all_diag.extend(diagnose_site(
            site, site_alerts, hw, rules, ai.get(site, ""), power,
            caps.get(site, 0)))

    # cross-site pattern: same fault type at many sites simultaneously
    # (supply-chain / portfolio-wide grid events)
    gf_sites = {d["site"] for d in all_diag if d["key"] == "GRID_EVENT"}
    if len(gf_sites) >= 3:
        all_diag.append({
            "site": "PORTFOLIO", "key": "REGIONAL_GRID",
            "title": f"Regional grid disturbance — {len(gf_sites)} sites tripped together",
            "confidence": 2, "conf_label": "Medium",
            "evidence": [f"[pattern] grid-event diagnosis at: {', '.join(sorted(gf_sites)[:6])}"],
            "action": "Treat as utility-side event; verify all sites auto-reconnected "
                      "rather than dispatching to each individually.",
            "priority": 95, "open": False,
        })

    all_diag.sort(key=lambda d: (-d["open"], -d["confidence"], -d["priority"],
                                 -len(d["evidence"])))
    return all_diag


if __name__ == "__main__":
    print("Run via ae_alert_dashboard.py (needs live classified alerts), "
          "or import diagnose_portfolio().")

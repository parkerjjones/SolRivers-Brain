#!/usr/bin/env python3
"""
Fetches AlsoEnergy AI-generated site summaries for every site in the portfolio.

Each summary is a Markdown natural language description of:
  - Current hardware communication status
  - Production vs expected output
  - Weather context
  - Any active alerts or anomalies

The API streams via Server-Sent Events (SSE). Summaries are cached for ~5 minutes.

Writes ae_ai_summaries.xlsx with one row per site.

USAGE
-----
    python ae_ai_summaries.py
    python ae_ai_summaries.py --output ae_ai_summaries.xlsx
"""

import argparse
import re
import shlex
import sys
import time

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CURL_FILE = "alsoenergy_curl.txt"
API_BASE  = "https://apps.alsoenergy.com/api"
PORTFOLIO = "C12941"
SLEEP     = 1.0

DARK  = "1F4E79"
MID   = "2E75B6"
WHITE = "FFFFFF"
ALT   = "EBF3FB"


def build_session():
    from ae_auth import get_session
    return get_session()


def fetch_ai_summary(session, site_id):
    """Reads the SSE stream and assembles the full AI summary text."""
    url = f"{API_BASE}/ai-site-summary?siteId={site_id}&lang=en-US"
    session.headers["accept"] = "text/event-stream"
    try:
        r = session.get(url, timeout=30, stream=True)
        if not r.ok:
            return None, None, r.status_code

        metadata = None
        chunks = []
        for raw_line in r.iter_lines(decode_unicode=True):
            if not raw_line:
                continue
            if raw_line.startswith("data:"):
                data = raw_line[5:].strip()
                if not data:
                    continue
                try:
                    import json
                    obj = json.loads(data)
                    if "metadata" in obj:
                        metadata = obj["metadata"]
                    elif "chunk" in obj:
                        chunks.append(obj["chunk"])
                except Exception:
                    chunks.append(data)

        full_text = "".join(chunks)
        # strip markdown links to get plain text for the cell
        plain = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", full_text)
        return full_text, plain, 200
    except Exception as e:
        return None, str(e), -1
    finally:
        session.headers["accept"] = "application/json"


def hdr(cell, text, bg=DARK):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_xlsx(path, rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1 — Full summaries
    ws = wb.create_sheet("AI Summaries")
    cols = ["Site Key", "Site Name", "AI Summary (Plain Text)", "Retrieved At"]
    for ci, c in enumerate(cols, 1):
        hdr(ws.cell(1, ci), c, MID)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 100
    ws.column_dimensions["D"].width = 20

    for ri, row in enumerate(rows, start=2):
        alt = ri % 2
        for ci, val in enumerate([row["site_key"], row["site_name"],
                                   row["plain_text"], row["retrieved_at"]], start=1):
            cell = ws.cell(ri, ci)
            cell.value = val
            if alt == 0:
                cell.fill = PatternFill("solid", fgColor=ALT)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 3))
        ws.row_dimensions[ri].height = max(15, min(len(row["plain_text"] or "") // 8, 200))

    # Sheet 2 — Keyword extraction
    ws2 = wb.create_sheet("Keywords")
    kw_cols = ["Site Key", "Site Name", "Mentions Alert", "Mentions Weather",
               "Mentions Communication", "Mentions Performance", "Mentions Production"]
    for ci, c in enumerate(kw_cols, 1):
        hdr(ws2.cell(1, ci), c, MID)
    ws2.freeze_panes = "A2"

    for ri, row in enumerate(rows, start=2):
        alt = ri % 2
        txt = (row["plain_text"] or "").lower()
        ws2.cell(ri, 1).value = row["site_key"]
        ws2.cell(ri, 2).value = row["site_name"]
        ws2.cell(ri, 3).value = "Yes" if any(w in txt for w in ["alert", "fault", "error", "issue"]) else "No"
        ws2.cell(ri, 4).value = "Yes" if any(w in txt for w in ["rain", "cloud", "weather", "wind", "snow", "irradiance"]) else "No"
        ws2.cell(ri, 5).value = "Yes" if any(w in txt for w in ["communicat", "offline", "not responding", "gateway"]) else "No"
        ws2.cell(ri, 6).value = "Yes" if any(w in txt for w in ["performance", "underperform", "below", "loss"]) else "No"
        ws2.cell(ri, 7).value = "Yes" if any(w in txt for w in ["production", "kwh", "kw", "output", "generat"]) else "No"
        for ci in range(1, 8):
            if ri % 2 == 0:
                ws2.cell(ri, ci).fill = PatternFill("solid", fgColor=ALT)

    for col in ws2.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 30)

    wb.save(path)
    print(f"Saved -> {path}  ({len(rows)} summaries)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", default="ae_ai_summaries.xlsx")
    args = ap.parse_args()

    from datetime import datetime, timezone

    session = build_session()

    print("Fetching site list...")
    r = session.get(f"{API_BASE}/view/portfolio/{PORTFOLIO}", timeout=20)
    if r.status_code in (401, 403):
        sys.exit("Auth expired — re-copy cURL.")
    EXCLUDE_SITES = {"S55935"}  # not in portfolio
    sites = [s for s in r.json().get("sites", []) if s.get("key") not in EXCLUDE_SITES]
    print(f"  {len(sites)} sites")

    rows = []
    for i, site in enumerate(sites):
        sk, sn = site["key"], site.get("name", site["key"])
        print(f"  [{i+1:2}/{len(sites)}] {sk} {sn}")
        md_text, plain, status = fetch_ai_summary(session, sk)
        if status != 200:
            print(f"    WARN HTTP {status}", file=sys.stderr)
        rows.append({
            "site_key":     sk,
            "site_name":    sn,
            "md_text":      md_text or "",
            "plain_text":   plain or "",
            "retrieved_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        })
        time.sleep(SLEEP)

    write_xlsx(args.output, rows)

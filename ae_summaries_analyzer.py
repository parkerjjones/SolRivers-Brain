#!/usr/bin/env python3
"""
Deep extraction from AlsoEnergy AI site summaries.

Pulls fresh summaries for all 35 sites, then extracts every piece of
structured information:
  - Site health classification
  - Production status vs expected
  - Weather conditions
  - Hardware issues / device names
  - Numeric values (kW, kWh, %, temperatures)
  - Embedded chart URLs decoded into site/hardware/date parameters
  - Action items and alerts mentioned
  - Sentiment / urgency score

Writes ae_summaries_deep.xlsx with one highly detailed sheet per topic.

USAGE
-----
    python ae_summaries_analyzer.py                  # fetch live + analyze
    python ae_summaries_analyzer.py --cached         # reuse ae_ai_summaries.xlsx
"""

import argparse
import re
import shlex
import sys
import time
from datetime import datetime, timezone
from urllib.parse import urlparse, parse_qs

import pandas as pd
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CURL_FILE = "alsoenergy_curl.txt"
API_BASE  = "https://apps.alsoenergy.com/api"
PORTFOLIO = "C12941"
SLEEP     = 0.8

DARK  = "1F4E79"
MID   = "2E75B6"
WHITE = "FFFFFF"
ALT   = "EBF3FB"
GREEN = "C6EFCE"
RED   = "FFC7CE"
YELL  = "FFEB9C"
BLUE  = "DDEEFF"

# -----------------------------------------------------------------------
# PATTERNS
# -----------------------------------------------------------------------

NUM_PATTERNS = {
    "production_kwh":    r"([\d,]+(?:\.\d+)?)\s*kWh?(?:\s+produced|\s+today|\s+generation)?",
    "power_kw":          r"([\d,]+(?:\.\d+)?)\s*kW\b",
    "pct_expected":      r"(\d+(?:\.\d+)?)\s*%\s*(?:of\s+expected|expected)",
    "performance_index": r"[Pp]erformance\s+[Ii]ndex[:\s]+([\d.]+)",
    "temp_f":            r"([-\d.]+)\s*°?F\b",
    "temp_c":            r"([-\d.]+)\s*°?C\b",
    "pct_loss":          r"(\d+(?:\.\d+)?)\s*%\s*(?:low|below|loss|less|reduction)",
}

HEALTH_KEYWORDS = {
    "all_comm":       ["all hardware is communicating", "all devices communicating",
                       "all communicating", "communicating normally", "100% communication"],
    "partial_comm":   ["partially communicating", "some devices", "intermittent",
                       "communication issue", "not communicating", "not responding",
                       "offline", "communication fail"],
    "production_ok":  ["in line with expected", "tracking normally", "on track",
                       "performing as expected", "meeting expectations", "above expected"],
    "production_low": ["tracking low", "below expected", "underperforming",
                       "production loss", "reduced output", "lower than expected",
                       "performance issue"],
    "has_alert":      ["alert", "fault", "error", "issue detected", "anomaly",
                       "failed", "failure", "degraded", "alarm"],
    "weather_impact": ["due to weather", "cloudy", "overcast", "rain", "snow",
                       "irradiance", "partly cloudy", "cloud cover"],
}

DEVICE_PATTERN = re.compile(
    r"(?:inverter|meter|tracker|gateway|logger|datalogger|modem|"
    r"weather station|pyranometer|sensor|relay|camera|UPS|combiner|"
    r"string\s+inverter|central\s+inverter)\s*[\w\s\-\.]*",
    re.IGNORECASE
)

URL_PATTERN = re.compile(r'https?://apps\.alsoenergy\.com[^\s\)"\']+')

URGENCY_WORDS = {
    3: ["critical", "severe", "down", "failed", "offline", "not communicating",
        "no production", "not responding", "fault"],
    2: ["alert", "error", "below expected", "tracking low", "underperforming",
        "issue", "reduced", "partial"],
    1: ["warning", "slightly", "minor", "intermittent", "occasional"],
    0: ["communicating", "normally", "in line", "as expected", "on track"],
}

# -----------------------------------------------------------------------
# SESSION
# -----------------------------------------------------------------------

def build_session():
    with open(CURL_FILE, "r", encoding="utf-8") as f:
        raw = f.read().replace("\\\r\n", " ").replace("\\\n", " ")
    tokens = shlex.split(raw)
    headers, cookie = {}, None
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if t in ("-H", "--header"):
            k, _, v = tokens[i + 1].partition(":")
            headers[k.strip()] = v.strip()
            i += 2
        elif t in ("-b", "--cookie"):
            cookie = tokens[i + 1]
            i += 2
        else:
            i += 1
    s = requests.Session()
    s.headers.update({k: v for k, v in headers.items() if k.lower() != "cookie"})
    if cookie:
        for part in cookie.split("; "):
            if "=" in part:
                n, _, v = part.partition("=")
                s.cookies.set(n, v)
    return s


def fetch_summaries(session):
    r = session.get(f"{API_BASE}/view/portfolio/{PORTFOLIO}", timeout=20)
    if r.status_code in (401, 403):
        sys.exit("Auth expired — re-copy cURL.")
    sites = r.json().get("sites", [])
    out = []
    for i, site in enumerate(sites):
        sk, sn = site["key"], site.get("name", "")
        print(f"  [{i+1:2}/{len(sites)}] {sk} {sn}")
        session.headers["accept"] = "text/event-stream"
        try:
            r2 = session.get(
                f"{API_BASE}/ai-site-summary?siteId={sk}&lang=en-US",
                timeout=25, stream=True
            )
            chunks = []
            for line in r2.iter_lines(decode_unicode=True):
                if line.startswith("data:"):
                    import json
                    try:
                        obj = json.loads(line[5:].strip())
                        if "chunk" in obj:
                            chunks.append(obj["chunk"])
                    except Exception:
                        pass
            text = "".join(chunks)
        except Exception as e:
            text = ""
            print(f"    WARN: {e}", file=sys.stderr)
        finally:
            session.headers["accept"] = "application/json"
        out.append({"site_key": sk, "site_name": sn, "text": text})
        time.sleep(SLEEP)
    return out


def load_cached_summaries(path):
    df = pd.read_excel(path, sheet_name="AI Summaries")
    out = []
    for _, row in df.iterrows():
        out.append({
            "site_key":  str(row.get("Site Key", "")),
            "site_name": str(row.get("Site Name", "")),
            # prefer raw markdown (Full Summaries sheet) but fall back to plain text
            "text":      str(row.get("AI Summary (Plain Text)", "")),
        })
    # try to load raw markdown from Full Summaries sheet if it exists
    try:
        df2 = pd.read_excel(path, sheet_name="Full Summaries")
        md_map = {str(r.get("Site Key","")): str(r.get("Full Summary Text",""))
                  for _, r in df2.iterrows()}
        for rec in out:
            if rec["site_key"] in md_map and md_map[rec["site_key"]]:
                rec["text"] = md_map[rec["site_key"]]
    except Exception:
        pass
    return out

# -----------------------------------------------------------------------
# EXTRACTION
# -----------------------------------------------------------------------

def classify_health(text):
    t = text.lower()
    flags = {}
    for key, phrases in HEALTH_KEYWORDS.items():
        flags[key] = any(p in t for p in phrases)
    if flags["all_comm"] and not flags["partial_comm"]:
        comm = "All Communicating"
    elif flags["partial_comm"]:
        comm = "Partial / Issues"
    else:
        comm = "Unknown"
    if flags["production_ok"] and not flags["production_low"]:
        prod = "On Track"
    elif flags["production_low"]:
        prod = "Below Expected"
    else:
        prod = "Unknown"
    return comm, prod, flags["has_alert"], flags["weather_impact"]


def score_urgency(text):
    t = text.lower()
    # strip negated phrases so "no active alerts" doesn't score high
    t = re.sub(r"no\s+\w+\s+alerts?", " ", t)
    t = re.sub(r"all\s+hardware\s+is\s+communicating", " ", t)
    t = re.sub(r"all\s+communicating", " ", t)
    best = 0
    for score, words in URGENCY_WORDS.items():
        if any(w in t for w in words):
            best = max(best, score)
    return best


def extract_numbers(text):
    result = {}
    for field, pattern in NUM_PATTERNS.items():
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            val_str = m.group(1).replace(",", "")
            try:
                result[field] = float(val_str)
            except ValueError:
                pass
    return result


def extract_urls(text):
    urls = URL_PATTERN.findall(text)
    parsed = []
    for url in urls:
        try:
            p = urlparse(url)
            qs = parse_qs(p.query)
            path_parts = p.path.strip("/").split("/")
            # extract site/hardware key from path
            site_k = next((x for x in path_parts if x.startswith("S") and x[1:].isdigit()), "")
            hw_k   = next((x for x in path_parts if x.startswith("H") and x[1:].isdigit()), "")
            section = path_parts[-1] if path_parts else ""
            parsed.append({
                "url":        url[:200],
                "site_key":   site_k,
                "hw_key":     hw_k,
                "section":    section,
                "start":      qs.get("start", [""])[0],
                "end":        qs.get("end", [""])[0],
                "bin":        qs.get("bin", [""])[0],
                "columns":    qs.get("c", [""])[0],
                "hardware_h": qs.get("h", [""])[0],  # e.g. "1*2*5*11"
            })
        except Exception:
            pass
    return parsed


def extract_devices(text):
    matches = DEVICE_PATTERN.findall(text)
    return list({m.strip() for m in matches if len(m.strip()) > 3})[:10]


def extract_action_items(text):
    lines = text.replace("\\n", "\n").split("\n")
    items = []
    for line in lines:
        line = line.strip("- •\t ")
        if any(w in line.lower() for w in ["recommend", "should", "need", "require",
                                             "check", "inspect", "replace", "investigate"]):
            items.append(line[:200])
    return items[:5]

# -----------------------------------------------------------------------
# ANALYZE ALL
# -----------------------------------------------------------------------

def analyze(summaries):
    rows = []
    url_rows = []
    device_rows = []
    action_rows = []

    for rec in summaries:
        sk, sn, text = rec["site_key"], rec["site_name"], rec["text"]
        if not text:
            continue

        comm, prod, has_alert, weather_impact = classify_health(text)
        urgency = score_urgency(text)
        nums = extract_numbers(text)
        devices = extract_devices(text)
        urls = extract_urls(text)
        actions = extract_action_items(text)

        # plain text (no markdown)
        plain = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", text)

        rows.append({
            "site_key":          sk,
            "site_name":         sn,
            "comm_status":       comm,
            "prod_status":       prod,
            "has_active_alert":  "Yes" if has_alert else "No",
            "weather_impact":    "Yes" if weather_impact else "No",
            "urgency_score":     urgency,
            "urgency_label":     ["Clean", "Low", "Medium", "High"][urgency],
            "production_kwh":    nums.get("production_kwh", ""),
            "power_kw":          nums.get("power_kw", ""),
            "pct_expected":      nums.get("pct_expected", ""),
            "perf_index":        nums.get("performance_index", ""),
            "temp_f":            nums.get("temp_f", ""),
            "pct_loss":          nums.get("pct_loss", ""),
            "device_count":      len(devices),
            "devices_mentioned": "; ".join(devices[:5]),
            "action_count":      len(actions),
            "embedded_url_count": len(urls),
            "summary_length":    len(text),
            "plain_text":        plain[:500],
        })

        for u in urls:
            url_rows.append({"site_key": sk, "site_name": sn, **u})

        for d in devices:
            device_rows.append({"site_key": sk, "site_name": sn, "device": d})

        for a in actions:
            action_rows.append({"site_key": sk, "site_name": sn, "action": a})

    return (pd.DataFrame(rows),
            pd.DataFrame(url_rows),
            pd.DataFrame(device_rows),
            pd.DataFrame(action_rows))

# -----------------------------------------------------------------------
# EXCEL
# -----------------------------------------------------------------------

def hdr(cell, text, bg=DARK):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_df_ws(wb, title, df, color_col=None, color_map=None):
    ws = wb.create_sheet(title)
    if df.empty:
        ws["A1"].value = "(No data)"
        return ws
    for ci, col in enumerate(df.columns, 1):
        hdr(ws.cell(1, ci), col, MID)
    ws.freeze_panes = "A2"
    for ri, row in enumerate(df.itertuples(index=False), start=2):
        alt = ri % 2
        bg = ALT if alt == 0 else None
        if color_col and color_map:
            val = getattr(row, color_col, None)
            bg = color_map.get(str(val), bg)
        for ci, v in enumerate(row, start=1):
            cell = ws.cell(ri, ci)
            cell.value = v if not (isinstance(v, float) and pd.isna(v)) else ""
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == len(df.columns)))
    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 60)
    return ws


def write_xlsx(path, main_df, url_df, device_df, action_df, summaries):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    urgency_colors = {
        "High": RED, "Medium": YELL, "Low": BLUE, "Clean": GREEN
    }

    # Sheet 1 — Main Analysis
    write_df_ws(wb, "Site Analysis", main_df, "urgency_label", urgency_colors)

    # Sheet 2 — Urgency Ranking
    ranked = main_df.sort_values("urgency_score", ascending=False)[
        ["site_key", "site_name", "urgency_label", "comm_status", "prod_status",
         "has_active_alert", "weather_impact", "plain_text"]
    ]
    write_df_ws(wb, "Urgency Ranking", ranked, "urgency_label", urgency_colors)

    # Sheet 3 — Numeric Extractions
    num_df = main_df[["site_key", "site_name", "production_kwh", "power_kw",
                       "pct_expected", "perf_index", "temp_f", "pct_loss"]].copy()
    write_df_ws(wb, "Numeric Values", num_df)

    # Sheet 4 — Embedded Chart URLs (decoded)
    write_df_ws(wb, "Chart URLs", url_df)

    # Sheet 5 — Hardware Mentions
    write_df_ws(wb, "Device Mentions", device_df)

    # Sheet 6 — Action Items
    write_df_ws(wb, "Action Items", action_df)

    # Sheet 7 — Full Summaries (raw text)
    ws7 = wb.create_sheet("Full Summaries")
    hdr(ws7.cell(1, 1), "Site Key", MID)
    hdr(ws7.cell(1, 2), "Site Name", MID)
    hdr(ws7.cell(1, 3), "Full Summary Text", MID)
    ws7.freeze_panes = "A2"
    ws7.column_dimensions["A"].width = 12
    ws7.column_dimensions["B"].width = 28
    ws7.column_dimensions["C"].width = 120
    for ri, rec in enumerate(summaries, start=2):
        plain = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", rec["text"])
        ws7.cell(ri, 1).value = rec["site_key"]
        ws7.cell(ri, 2).value = rec["site_name"]
        ws7.cell(ri, 3).value = plain
        ws7.cell(ri, 3).alignment = Alignment(wrap_text=True, vertical="top")
        ws7.row_dimensions[ri].height = min(max(15, len(plain) // 10), 250)
        if ri % 2 == 0:
            for ci in range(1, 4):
                ws7.cell(ri, ci).fill = PatternFill("solid", fgColor=ALT)

    # Sheet 8 — Communication + Production Cross-tab
    ct = main_df.groupby(["comm_status", "prod_status"]).size().reset_index(name="count")
    write_df_ws(wb, "Status Cross-Tab", ct)

    wb.save(path)
    print(f"Saved -> {path}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")

# -----------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--cached", action="store_true",
                    help="reuse ae_ai_summaries.xlsx instead of fetching live")
    ap.add_argument("--input",  default="ae_ai_summaries.xlsx")
    ap.add_argument("--output", default="ae_summaries_deep.xlsx")
    args = ap.parse_args()

    if args.cached:
        print(f"Loading cached summaries from {args.input}...")
        summaries = load_cached_summaries(args.input)
    else:
        print("Fetching live AI summaries...")
        session = build_session()
        summaries = fetch_summaries(session)

    print(f"\nAnalyzing {len(summaries)} summaries...")
    main_df, url_df, device_df, action_df = analyze(summaries)

    print("\n--- Urgency Distribution ---")
    if not main_df.empty:
        print(main_df["urgency_label"].value_counts().to_string())
        print("\n--- Comm Status ---")
        print(main_df["comm_status"].value_counts().to_string())
        print("\n--- Production Status ---")
        print(main_df["prod_status"].value_counts().to_string())
        print(f"\n--- Embedded URLs found: {len(url_df)} ---")

    write_xlsx(args.output, main_df, url_df, device_df, action_df, summaries)

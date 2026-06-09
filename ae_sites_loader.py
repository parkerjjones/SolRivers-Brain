#!/usr/bin/env python3
"""
Pulls full site details for every site in the portfolio using:
  GET /api/view/site/{siteKey}
  GET /api/view/portfolio/{portfolioKey}  (live power readings)

Writes ae_sites.xlsx with:
  - Sites Overview     : all 35 sites with capacity, location, contracts
  - Live Power         : current power output per site (at time of run)
  - Contract Status    : monitoring and cell modem contract dates
  - Capacity Summary   : AC/DC capacity breakdown

USAGE
-----
    python ae_sites_loader.py
    python ae_sites_loader.py --output ae_sites.xlsx
"""

import argparse
import shlex
import sys
import time
from datetime import datetime

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

CURL_FILE = "alsoenergy_curl.txt"
API_BASE  = "https://apps.alsoenergy.com/api"
PORTFOLIO = "C12941"
SLEEP     = 0.4

CONTRACT_STATUS = {0: "Active", 1: "Warning", 2: "Expired", 3: "Unknown"}
SITE_STATUS     = {0: "Offline", 1: "Online", 8: "Partial"}

DARK  = "1F4E79"
MID   = "2E75B6"
WHITE = "FFFFFF"
ALT   = "EBF3FB"
GREEN = "C6EFCE"
RED   = "FFC7CE"


def build_session():
    with open(CURL_FILE, "r", encoding="utf-8") as f:
        raw = f.read().replace("\\\r\n", " ").replace("\\\n", " ")
    tokens = shlex.split(raw)
    headers, cookie = {}, None
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if t in ("-H", "--header"):
            k, _, v = tokens[i + 1].partition(":")
            headers[k.strip()] = v.strip()
            i += 2
        elif t in ("-b", "--cookie"):
            cookie = tokens[i + 1]
            i += 2
        else:
            i += 1
    s = requests.Session()
    s.headers.update({k: v for k, v in headers.items() if k.lower() != "cookie"})
    if cookie:
        for part in cookie.split("; "):
            if "=" in part:
                n, _, v = part.partition("=")
                s.cookies.set(n, v)
    return s


def hdr(cell, text, bg=DARK):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def val(cell, v, row_idx=0, fmt=None):
    cell.value = v
    if row_idx % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=ALT)
    cell.alignment = Alignment(vertical="top")
    if fmt:
        cell.number_format = fmt


def autosize(ws, max_w=50):
    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, max_w)


def ts(v):
    if not v:
        return None
    try:
        return datetime.fromisoformat(v.replace("Z", ""))
    except Exception:
        return v


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", default="ae_sites.xlsx")
    args = ap.parse_args()

    session = build_session()

    # ── Fetch portfolio (live power + site list) ──────────────────────
    print("Fetching portfolio live data...")
    r = session.get(f"{API_BASE}/view/portfolio/{PORTFOLIO}", timeout=20)
    if r.status_code in (401, 403):
        sys.exit("Auth expired — re-copy a fresh cURL.")
    r.raise_for_status()
    portfolio = r.json()
    site_list = portfolio.get("sites", [])
    print(f"  {len(site_list)} sites in portfolio")

    # Build live-power lookup
    live = {s["key"]: s for s in site_list}

    # ── Fetch full site detail for each site ──────────────────────────
    details = {}
    for i, site in enumerate(site_list):
        sk = site["key"]
        print(f"  [{i+1:2}/{len(site_list)}] {sk} {site.get('name','')}")
        r2 = session.get(f"{API_BASE}/view/site/{sk}", timeout=20)
        if r2.ok:
            details[sk] = r2.json()
        else:
            print(f"    WARN HTTP {r2.status_code}", file=sys.stderr)
        time.sleep(SLEEP)

    # ── Write Excel ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1 — Sites Overview
    ws = wb.create_sheet("Sites Overview")
    oc = ["Site Key", "Site Name", "Address", "City", "State", "Lat", "Lon",
          "Timezone", "Status", "Capacity AC (kW)", "Capacity DC (kW)",
          "Daily Est (kWh)", "Monthly Est (kWh)", "Commission Date",
          "Valid Data From", "Is Monitored"]
    for ci, c in enumerate(oc, 1):
        hdr(ws.cell(1, ci), c, MID)
    ws.freeze_panes = "A2"

    for ri, sk in enumerate(details, start=2):
        d = details[sk]
        addr = d.get("address") or {}
        alt = ri % 2
        val(ws.cell(ri, 1), sk, alt)
        val(ws.cell(ri, 2), d.get("name", ""), alt)
        val(ws.cell(ri, 3), addr.get("address1", ""), alt)
        val(ws.cell(ri, 4), addr.get("city", ""), alt)
        val(ws.cell(ri, 5), addr.get("stateProvince", ""), alt)
        val(ws.cell(ri, 6), d.get("latitude"), alt)
        val(ws.cell(ri, 7), d.get("longitude"), alt)
        val(ws.cell(ri, 8), "", alt)          # timezone not in /view/site, use scriptsite
        val(ws.cell(ri, 9), SITE_STATUS.get(d.get("status", -1), str(d.get("status"))), alt)
        val(ws.cell(ri, 10), d.get("capacityAc") or 0, alt, "0.00")
        val(ws.cell(ri, 11), d.get("capacityDc") or 0, alt, "0.00")
        val(ws.cell(ri, 12), round(d.get("dailyProductionEstimate") or 0, 1), alt, "0.0")
        val(ws.cell(ri, 13), round(d.get("monthlyProductionEstimate") or 0, 1), alt, "0.0")
        val(ws.cell(ri, 14), ts(d.get("actualCommissioningDate")), alt)
        val(ws.cell(ri, 15), ts(d.get("validDataDate")), alt)
        val(ws.cell(ri, 16), "Yes" if d.get("isMonitored") else "No", alt)

    autosize(ws)

    # Sheet 2 — Live Power
    ws2 = wb.create_sheet("Live Power")
    lc = ["Site Key", "Site Name", "Current Power (kW)", "15-min Avg (kW)", "Expected (kW)", "Last Upload"]
    for ci, c in enumerate(lc, 1):
        hdr(ws2.cell(1, ci), c, MID)
    ws2.freeze_panes = "A2"

    for ri, site in enumerate(site_list, start=2):
        alt = ri % 2
        val(ws2.cell(ri, 1), site["key"], alt)
        val(ws2.cell(ri, 2), site.get("name", ""), alt)
        val(ws2.cell(ri, 3), round(site.get("power") or 0, 2), alt, "0.00")
        val(ws2.cell(ri, 4), round(site.get("powerAvg15") or 0, 2), alt, "0.00")
        val(ws2.cell(ri, 5), round(site.get("powerAvg15Exp") or 0, 2), alt, "0.00")
        val(ws2.cell(ri, 6), ts(site.get("lastDataUTC")), alt)
        # Highlight underperforming (actual < 50% of expected)
        pwr = site.get("powerAvg15") or 0
        exp = site.get("powerAvg15Exp") or 0
        if exp > 0 and pwr < exp * 0.5:
            for ci in range(1, 7):
                ws2.cell(ri, ci).fill = PatternFill("solid", fgColor=RED)

    autosize(ws2)

    # Sheet 3 — Contract Status
    ws3 = wb.create_sheet("Contract Status")
    cc = ["Site Key", "Site Name", "Monitor Status", "Monitor Start",
          "Monitor End", "Monitor Warn", "Modem Status", "Modem End"]
    for ci, c in enumerate(cc, 1):
        hdr(ws3.cell(1, ci), c, MID)
    ws3.freeze_panes = "A2"

    for ri, sk in enumerate(details, start=2):
        d = details[sk]
        alt = ri % 2
        mon_status = CONTRACT_STATUS.get(d.get("monitoringContractStatus", -1), "?")
        mod_status = CONTRACT_STATUS.get(d.get("cellModemContractStatus", -1), "?")
        val(ws3.cell(ri, 1), sk, alt)
        val(ws3.cell(ri, 2), d.get("name", ""), alt)
        val(ws3.cell(ri, 3), mon_status, alt)
        val(ws3.cell(ri, 4), ts(d.get("monitoringContractStartDate")), alt)
        val(ws3.cell(ri, 5), ts(d.get("monitoringContractEndDate")), alt)
        val(ws3.cell(ri, 6), ts(d.get("monitoringContractWarnDate")), alt)
        val(ws3.cell(ri, 7), mod_status, alt)
        val(ws3.cell(ri, 8), ts(d.get("cellModemContractEndDate")), alt)

    autosize(ws3)

    wb.save(args.output)
    print(f"\nSaved -> {args.output}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")

#!/usr/bin/env python3
import openpyxl, csv, io
from datetime import date, datetime
from collections import defaultdict

wb = openpyxl.load_workbook('ae_alerts.xlsx')
ws = wb.active
headers = [c.value for c in ws[1]]
col = {h: i for i, h in enumerate(headers)}

rows_609 = [row for row in ws.iter_rows(min_row=2, values_only=True)
            if (row[col['alert_start']] and hasattr(row[col['alert_start']], 'date')
                and row[col['alert_start']].date() == date(2026, 6, 9))]

groups = defaultdict(list)
for r in rows_609:
    key = (r[col['site_name']], r[col['asset_code']], r[col['event_type_name']])
    groups[key].append(r)

SITE_SHORT = {
    'C & B Graham Energy': 'Graham',
    'Longleaf Pine Solar, LLC': 'Longleaf',
    'Williams Solar, LLC': 'Williams',
    'Sunflower Solar': 'Sunflower',
    'Washington Solar': 'Washington',
    'Elk Solar': 'Elk',
    'Marble Solar': 'Marble',
    'Harding Solar': 'Harding',
    'Auburn Solar': 'Auburn',
    'Rio Grande Elementary': 'Rio Grande',
}
ASSET_EQ = {'INV':'Inverter','MTR':'Meter','TKR':'Tracker','GTW':'Gateway','DA':'Data Acq','UNK':'Site'}
EVENT_DESC = {
    'Device communication': 'Comm Failure',
    'Solectria XGI 1500 Series - Fault Alert': 'Inverter Fault',
    'Solectria XGI 1000 Series - Fault Alert': 'Inverter Fault',
    'Gamechange Tracker Alert': 'Tracker Alert',
    'Power Meter Check 2.0': 'Meter Check',
    'Rule Tool Alert': 'Rule Fail',
    'Phoenix Contact - Quint UPS Alerts': 'UPS Alert',
    'Phoenix Contact - Quint UPS Alerts (QUINT4 EIP)': 'UPS Alert',
    'Solar FlexRack Tracker Alert (revG)': 'Tracker Alert',
    '.Chint / Solectria / Canadian String Inv Faults': 'Inverter Fault',
    'Gateway heartbeat': 'Gateway Offline',
    'Transformer Monitor': 'Transformer Alert',
    'Sungrow 3150U/3425U/3600U inverter alerts': 'Inverter Fault',
}

now_ts = '6/9/2026 9:00:00'
email = 'parker@solrivercapital.com'

def fmt_date(dt):
    return f"{dt.month}/{dt.day}/{dt.year}" if dt else ''

def fmt_time(dt):
    h = dt.hour % 12 or 12
    ampm = 'AM' if dt.hour < 12 else 'PM'
    return f"{h}:{dt.minute:02d}:{dt.second:02d} {ampm}" if dt else ''

new_rows = []
for (site, asset, event), rows in sorted(groups.items()):
    short = SITE_SHORT.get(site, site)
    equipment = ASSET_EQ.get(asset, asset)
    desc = EVENT_DESC.get(event, event[:20])
    qty = len(rows)

    starts = [r[col['alert_start']] for r in rows if r[col['alert_start']]]
    ends   = [r[col['alert_end']]   for r in rows if r[col['alert_end']]]
    start_dt = min(starts) if starts else None
    end_dt   = max(ends)   if ends   else None

    # Device IDs
    names = [r[col['hardware_name']] for r in rows]
    nums = [(n or '').replace('INVERTER ', '').strip() for n in names]
    if qty <= 5:
        eq_id = ', '.join(nums)
    else:
        try:
            nums_sorted = sorted(nums, key=lambda x: int(x) if x.isdigit() else 9999)
        except Exception:
            nums_sorted = sorted(nums)
        eq_id = f"{nums_sorted[0]}-{nums_sorted[-1]} ({qty} devices)"

    # Category
    resolved_count = sum(1 for r in rows if r[col['is_resolved']])
    if 'Fault' in event or 'Faults' in event:
        category = 'Repair'
    elif 'communication' in event.lower() or 'heartbeat' in event.lower():
        category = 'Outage' if resolved_count < qty else 'Repair'
    else:
        category = 'Repair'

    prod_loss = 1 if asset in ('INV', 'GTW') or 'Fault' in event else 0

    # Comment
    descs = list(dict.fromkeys(r[col['description']] for r in rows if r[col['description']]))
    comment = descs[0] if descs else event
    if qty > 1:
        comment = f"{qty} {equipment.lower()}s affected. {comment}"

    row = [
        now_ts, email, short, equipment, '',
        desc, qty, eq_id,
        fmt_date(start_dt), comment[:300],
        category, fmt_date(end_dt) if end_dt else '',
        '', '', prod_loss,
        '','','','','','','','','','','',
        fmt_time(start_dt), fmt_time(end_dt) if end_dt else '',
        '', '',
    ]
    new_rows.append(row)
    print(f"{short:22s} | {equipment:10s} | {desc:20s} | n={qty:3d} | {fmt_date(start_dt)} | {'open' if not end_dt else fmt_date(end_dt):10s} | loss={prod_loss}")

print(f"\nTotal new rows: {len(new_rows)}")

# Write to CSV for inspection
import csv
with open('tracker_new_rows_609.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerows(new_rows)
print("Saved -> tracker_new_rows_609.csv")

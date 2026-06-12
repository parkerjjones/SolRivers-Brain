#!/usr/bin/env python3
"""
Comprehensive inverter issue tracker for the SolRiver Capital portfolio.

Fetches:
  - Hardware inventory  → identifies all inverters per site (fc=1)
  - Alert history       → filters to inverter alerts (assetCode=INV)
  - Rule results        → diagnostic check results per site

Writes ae_inverter_report.xlsx with 5 sheets:
  - Portfolio Summary     : site roll-up (inverter count, alert count, active issues)
  - Active Issues         : unresolved inverter alerts, severity-sorted
  - Alert History         : every inverter alert in the date window
  - Inverter Inventory    : every inverter with issue counts
  - Rule Results          : all rule check results (Fail/Warning highlighted)

USAGE
-----
    python ae_inverter_report.py
    python ae_inverter_report.py --from 2026-01-01 --to 2026-06-11
    python ae_inverter_report.py --output my_report.xlsx
"""

import argparse
import re
import sys
import time
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

API_BASE       = "https://apps.alsoenergy.com/api"
PORTFOLIO_KEYS = ["C12941", "C47197"]
SLEEP          = 0.4
TZ_OFFSET_MIN  = 360   # minutes west of UTC (US/Eastern)

DARK  = "1F4E79"
MID   = "2E75B6"
WHITE = "FFFFFF"
ALT   = "EBF3FB"
RED   = "FFC7CE"
YELL  = "FFEB9C"

SEVERITY_LABEL = {1: "Info", 2: "Warning", 3: "Critical", 4: "Emergency"}
RESULT_CODE    = {1: "N/A", 2: "Warning", 3: "Fail", 4: "Pass"}


# ── Excel helpers ─────────────────────────────────────────────────────────

def hdr(cell, text, bg=DARK):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def val(cell, v, row_idx=0, fmt=None, bg=None):
    cell.value = v
    fill = bg or (ALT if row_idx % 2 == 0 else None)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="top")
    if fmt:
        cell.number_format = fmt


def autosize(ws, max_w=55):
    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, max_w)


# ── Data fetching ─────────────────────────────────────────────────────────

def fetch_sites(session, portfolio_key):
    r = session.get(f"{API_BASE}/view/portfolio/{portfolio_key}", timeout=20)
    if r.status_code in (401, 403):
        sys.exit("Auth expired — re-run ae_auth.py")
    r.raise_for_status()
    return r.json().get("sites", [])


def fetch_hardware(session, site_key):
    url = f"{API_BASE}/scriptsite/{site_key}?lastChanged=1900-01-01T00:00:00.000Z"
    r = session.get(url, timeout=20)
    if not r.ok:
        print(f"  WARN {site_key}: HTTP {r.status_code}", file=sys.stderr)
        return None
    return r.json()


def fetch_alerts_window(session, portfolio_key, d_from, d_to):
    records = []
    cur = d_from
    while cur <= d_to:
        end = min(cur + timedelta(days=30), d_to)
        r = session.post(
            f"{API_BASE}/view/alerthistory",
            json={"key": portfolio_key, "from": cur.isoformat(),
                  "to": end.isoformat(), "offset": TZ_OFFSET_MIN},
            timeout=60,
        )
        if r.status_code == 204:
            cur = end + timedelta(days=1)
            continue
        if r.status_code in (401, 403):
            sys.exit("Auth expired during alert fetch")
        r.raise_for_status()
        data = r.json()
        chunk = data if isinstance(data, list) else (
            data.get("list") or data.get("alerts") or data.get("items") or []
        )
        records.extend(chunk)
        label = f"{cur}..{end}"
        print(f"    {portfolio_key} {label}: {len(chunk)} records")
        cur = end + timedelta(days=1)
        time.sleep(SLEEP)
    return records


def fetch_rule_results(session, portfolio_key):
    url = f"{API_BASE}/ruleresults/{portfolio_key}?lastChanged=1900-01-01T00:00:00.000Z&mergeHash="
    r = session.get(url, timeout=30)
    if not r.ok:
        print(f"  WARN rule results HTTP {r.status_code}", file=sys.stderr)
        return None
    return r.json()


# ── Normalization ─────────────────────────────────────────────────────────

def parse_ts(v):
    if not v:
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y %I:%M:%S %p"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def norm_alert(rec):
    ec = str(rec.get("eventCode", ""))
    m = re.match(r"^\s*(\d+)\s*-\s*(.+)$", ec)
    etype_name = m.group(2).strip() if m else ec
    a_start = parse_ts(rec.get("start"))
    a_end   = parse_ts(rec.get("end"))
    dur_h   = round((a_end - a_start).total_seconds() / 3600, 1) if (a_start and a_end) else None
    sev     = rec.get("severity") or 0
    return {
        "alert_id":       str(rec.get("alertId", "")),
        "site_key":       rec.get("siteKey"),
        "site_name":      rec.get("siteName"),
        "hardware_key":   rec.get("hardwareKey") or "",
        "hardware_name":  rec.get("hardwareName") or "",
        "asset_code":     rec.get("assetCode") or "",
        "event_type":     etype_name,
        "description":    rec.get("description") or "",
        "severity":       sev,
        "severity_label": SEVERITY_LABEL.get(sev, str(sev)),
        "impact":         rec.get("impact"),
        "capacity":       rec.get("capacity"),
        "is_resolved":    bool(rec.get("isResolved")),
        "is_ack":         bool(rec.get("isAcknowledged")),
        "resolved_by":    rec.get("resolvedByName") or "",
        "alert_start":    a_start,
        "alert_end":      a_end,
        "duration_h":     dur_h,
        "resolved_time":  parse_ts(rec.get("resolvedTime")),
    }


def is_inverter_alert(a):
    if (a.get("asset_code") or "").upper() == "INV":
        return True
    hw = (a.get("hardware_name") or "").lower()
    return any(kw in hw for kw in ["inv", "inverter", "pv inv"])


# ── Excel report ──────────────────────────────────────────────────────────

def write_report(path, inv_hw, inv_alerts, rule_data, d_from, d_to):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # pre-compute aggregates
    site_order, site_meta = [], {}
    for r in inv_hw:
        sk = r["site_key"]
        if sk not in site_meta:
            site_order.append(sk)
            site_meta[sk] = {"site_name": r["site_name"], "inv_count": 0}
        site_meta[sk]["inv_count"] += 1

    alert_by_site  = {}
    alert_by_hw    = {}
    active_by_site = {}
    active_by_hw   = {}
    for a in inv_alerts:
        sk  = a["site_key"]
        hwk = a["hardware_key"]
        alert_by_site[sk]  = alert_by_site.get(sk, 0) + 1
        alert_by_hw[hwk]   = alert_by_hw.get(hwk, 0) + 1
        if not a["is_resolved"]:
            active_by_site[sk] = active_by_site.get(sk, 0) + 1
            active_by_hw[hwk]  = active_by_hw.get(hwk, 0) + 1

    # ── Sheet 1: Portfolio Summary ────────────────────────────────────────
    ws1 = wb.create_sheet("Portfolio Summary")
    for ci, c in enumerate(
        ["Site Key", "Site Name", "Inverters", "Total Alerts",
         "Active Issues", "Issue Rate %", "Alerts / Inverter"], 1
    ):
        hdr(ws1.cell(1, ci), c, MID)
    ws1.freeze_panes = "A2"

    for ri, sk in enumerate(site_order, start=2):
        info  = site_meta[sk]
        n_inv = info["inv_count"]
        n_tot = alert_by_site.get(sk, 0)
        n_act = active_by_site.get(sk, 0)
        rate  = round(n_act / n_inv * 100, 1) if n_inv else 0
        per_i = round(n_tot / n_inv, 1) if n_inv else 0
        alt   = ri % 2
        val(ws1.cell(ri, 1), sk, alt)
        val(ws1.cell(ri, 2), info["site_name"], alt)
        val(ws1.cell(ri, 3), n_inv, alt)
        val(ws1.cell(ri, 4), n_tot, alt)
        val(ws1.cell(ri, 5), n_act, alt, bg=RED if n_act else None)
        val(ws1.cell(ri, 6), rate, alt, "0.0", bg=RED if rate > 20 else (YELL if rate > 0 else None))
        val(ws1.cell(ri, 7), per_i, alt, "0.0")
    autosize(ws1)

    # ── Sheet 2: Active Issues (unresolved, severity-sorted) ──────────────
    ws2 = wb.create_sheet("Active Issues")
    for ci, c in enumerate(
        ["Severity", "Site Name", "Site Key", "Inverter Name", "HW Key",
         "Event Type", "Description", "Started", "Duration (h)", "Impact", "Capacity kW"], 1
    ):
        hdr(ws2.cell(1, ci), c, MID)
    ws2.freeze_panes = "A2"

    active = sorted(
        [a for a in inv_alerts if not a["is_resolved"]],
        key=lambda x: (-(x["severity"] or 0), x["alert_start"] or datetime.min),
    )
    for ri, a in enumerate(active, start=2):
        alt    = ri % 2
        sev    = a["severity"] or 0
        row_bg = RED if sev >= 3 else (YELL if sev == 2 else None)
        val(ws2.cell(ri, 1), a["severity_label"], alt, bg=row_bg)
        val(ws2.cell(ri, 2), a["site_name"], alt)
        val(ws2.cell(ri, 3), a["site_key"], alt)
        val(ws2.cell(ri, 4), a["hardware_name"], alt)
        val(ws2.cell(ri, 5), a["hardware_key"], alt)
        val(ws2.cell(ri, 6), a["event_type"], alt)
        val(ws2.cell(ri, 7), a["description"], alt)
        val(ws2.cell(ri, 8), a["alert_start"], alt)
        val(ws2.cell(ri, 9), a["duration_h"], alt, "0.0")
        val(ws2.cell(ri, 10), a["impact"], alt)
        val(ws2.cell(ri, 11), a["capacity"], alt)
    autosize(ws2)

    # ── Sheet 3: Full Alert History ───────────────────────────────────────
    sheet_title = f"Alerts {d_from}..{d_to}"[:31]   # Excel tab name limit
    ws3 = wb.create_sheet(sheet_title)
    for ci, c in enumerate(
        ["Alert ID", "Site Name", "Site Key", "Inverter Name", "HW Key",
         "Asset Code", "Severity", "Event Type", "Description",
         "Resolved", "Acknowledged", "Started", "Ended", "Duration (h)",
         "Resolved Time", "Resolved By", "Impact", "Capacity kW"], 1
    ):
        hdr(ws3.cell(1, ci), c, MID)
    ws3.freeze_panes = "A2"

    sorted_all = sorted(
        inv_alerts,
        key=lambda x: (x["site_name"] or "", -(x["severity"] or 0),
                       x["alert_start"] or datetime.min),
    )
    for ri, a in enumerate(sorted_all, start=2):
        alt    = ri % 2
        sev    = a["severity"] or 0
        row_bg = RED if (sev >= 3 and not a["is_resolved"]) else None
        val(ws3.cell(ri, 1),  a["alert_id"],      alt)
        val(ws3.cell(ri, 2),  a["site_name"],      alt)
        val(ws3.cell(ri, 3),  a["site_key"],       alt)
        val(ws3.cell(ri, 4),  a["hardware_name"],  alt)
        val(ws3.cell(ri, 5),  a["hardware_key"],   alt)
        val(ws3.cell(ri, 6),  a["asset_code"],     alt)
        val(ws3.cell(ri, 7),  a["severity_label"], alt, bg=row_bg)
        val(ws3.cell(ri, 8),  a["event_type"],     alt)
        val(ws3.cell(ri, 9),  a["description"],    alt)
        val(ws3.cell(ri, 10), "Yes" if a["is_resolved"] else "No", alt)
        val(ws3.cell(ri, 11), "Yes" if a["is_ack"]      else "No", alt)
        val(ws3.cell(ri, 12), a["alert_start"],    alt)
        val(ws3.cell(ri, 13), a["alert_end"],      alt)
        val(ws3.cell(ri, 14), a["duration_h"],     alt, "0.0")
        val(ws3.cell(ri, 15), a["resolved_time"],  alt)
        val(ws3.cell(ri, 16), a["resolved_by"],    alt)
        val(ws3.cell(ri, 17), a["impact"],         alt)
        val(ws3.cell(ri, 18), a["capacity"],       alt)
    autosize(ws3)

    # ── Sheet 4: Inverter Inventory ───────────────────────────────────────
    ws4 = wb.create_sheet("Inverter Inventory")
    for ci, c in enumerate(
        ["Site Key", "Site Name", "HW Key", "Inverter Name",
         "HW Status", "Total Alerts", "Active Alerts", "Risk"], 1
    ):
        hdr(ws4.cell(1, ci), c, MID)
    ws4.freeze_panes = "A2"

    for ri, rec in enumerate(inv_hw, start=2):
        alt   = ri % 2
        n_tot = alert_by_hw.get(rec["hw_key"], 0)
        n_act = active_by_hw.get(rec["hw_key"], 0)
        risk  = "HIGH" if n_act else ("MED" if n_tot > 3 else "LOW")
        val(ws4.cell(ri, 1), rec["site_key"],        alt)
        val(ws4.cell(ri, 2), rec["site_name"],       alt)
        val(ws4.cell(ri, 3), rec["hw_key"],          alt)
        val(ws4.cell(ri, 4), rec["hw_name"],         alt)
        val(ws4.cell(ri, 5), rec["hardware_status"], alt)
        val(ws4.cell(ri, 6), n_tot, alt, bg=YELL if n_tot > 3 else None)
        val(ws4.cell(ri, 7), n_act, alt, bg=RED  if n_act      else None)
        val(ws4.cell(ri, 8), risk,  alt, bg=RED  if risk == "HIGH" else (YELL if risk == "MED" else None))
    autosize(ws4)

    # ── Sheet 5: Rule Results ─────────────────────────────────────────────
    ws5 = wb.create_sheet("Rule Results")
    for ci, c in enumerate(
        ["Site Key", "Site Name", "Rule Name", "Result", "Category", "Last Run"], 1
    ):
        hdr(ws5.cell(1, ci), c, MID)
    ws5.freeze_panes = "A2"

    if rule_data:
        ri_out = 2
        sites_rd = rule_data if isinstance(rule_data, list) else (
            rule_data.get("sites") or rule_data.get("results") or []
        )
        for site_rd in sites_rd:
            sk    = site_rd.get("key") or site_rd.get("siteKey", "")
            sn    = site_rd.get("name") or site_rd.get("siteName", "")
            rules = site_rd.get("results") or site_rd.get("rules") or []
            for rule in rules:
                r_int  = rule.get("result") or rule.get("resultCode") or 0
                r_str  = RESULT_CODE.get(r_int, str(r_int))
                row_bg = RED if r_int == 3 else (YELL if r_int == 2 else None)
                alt    = ri_out % 2
                val(ws5.cell(ri_out, 1), sk, alt)
                val(ws5.cell(ri_out, 2), sn, alt)
                val(ws5.cell(ri_out, 3), rule.get("ruleName") or rule.get("name", ""), alt)
                val(ws5.cell(ri_out, 4), r_str, alt, bg=row_bg)
                val(ws5.cell(ri_out, 5), rule.get("category") or rule.get("ruleCategory", ""), alt)
                val(ws5.cell(ri_out, 6), rule.get("runTime") or rule.get("lastRun", ""), alt)
                ri_out += 1
    else:
        ws5.cell(2, 1).value = "No rule results returned by API"
    autosize(ws5)

    wb.save(path)


# ── Entry point ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Inverter issue report for SolRiver portfolio")
    ap.add_argument("--from", dest="d_from", default="2026-01-01",
                    help="Alert history start (YYYY-MM-DD). Default: 2026-01-01")
    ap.add_argument("--to",   dest="d_to",   default=date.today().isoformat(),
                    help="Alert history end (YYYY-MM-DD). Default: today")
    ap.add_argument("--output", default="ae_inverter_report.xlsx")
    args = ap.parse_args()

    d_from = date.fromisoformat(args.d_from)
    d_to   = date.fromisoformat(args.d_to)

    from ae_auth import get_session
    session = get_session()

    # ── Step 1: Build inverter inventory ──────────────────────────────────
    print("Fetching hardware inventory...")
    all_sites = []
    for pk in PORTFOLIO_KEYS:
        sites = fetch_sites(session, pk)
        all_sites.extend(sites)
        print(f"  {pk}: {len(sites)} sites")

    inv_hw = []
    for i, site in enumerate(all_sites):
        sk = site["key"]
        sn = site.get("name", sk)
        print(f"  [{i+1:2}/{len(all_sites)}] {sk} {sn}")
        hw_data = fetch_hardware(session, sk)
        if not hw_data:
            continue
        for h in hw_data.get("hardware", []):
            if h.get("functionCode") == 1:
                inv_hw.append({
                    "site_key":       sk,
                    "site_name":      sn,
                    "hw_key":         h.get("key", ""),
                    "hw_name":        h.get("name", ""),
                    "hardware_status": h.get("hardwareStatus", ""),
                })
        time.sleep(SLEEP)

    print(f"\n  {len(inv_hw)} inverters across {len(all_sites)} sites")

    # ── Step 2: Alert history → inverter alerts only ──────────────────────
    print(f"\nFetching alerts {d_from} → {d_to}...")
    raw_alerts = []
    for pk in PORTFOLIO_KEYS:
        print(f"  Portfolio {pk}:")
        raw_alerts.extend(fetch_alerts_window(session, pk, d_from, d_to))

    all_norm   = [norm_alert(r) for r in raw_alerts if r.get("alertId")]
    inv_alerts = [a for a in all_norm if is_inverter_alert(a)]

    active_count = sum(1 for a in inv_alerts if not a["is_resolved"])
    print(f"\n  {len(raw_alerts)} total alerts → {len(inv_alerts)} inverter alerts")
    print(f"  Active (unresolved): {active_count}")

    # ── Step 3: Rule results ──────────────────────────────────────────────
    print("\nFetching rule results...")
    rule_data = None
    for pk in PORTFOLIO_KEYS:
        rd = fetch_rule_results(session, pk)
        if rd:
            rule_data = rd
            print(f"  Got rule results from {pk}")
            break

    # ── Step 4: Write report ──────────────────────────────────────────────
    print(f"\nWriting {args.output}...")
    write_report(args.output, inv_hw, inv_alerts, rule_data, d_from, d_to)
    print(f"\nDone → {args.output}")
    print("  Sheets: Portfolio Summary | Active Issues | Alert History | Inverter Inventory | Rule Results")

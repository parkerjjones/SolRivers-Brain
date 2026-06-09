#!/usr/bin/env python3
"""
AlsoEnergy / PowerTrack alert-history connector.

Pulls operational alert history per portfolio and loads it into the local
`solriver` Postgres/TimescaleDB instance, and optionally exports to Excel.

USAGE
-----
1. In the browser Network tab, right-click the `alerthistory` request,
   Copy > "Copy as cURL (bash)", and paste it into a file named
   `alsoenergy_curl.txt` next to this script. Do NOT commit that file.

2. Probe to verify the response shape:
       python ae_alert_loader.py --probe --key C12941

3. Run the full load:
       python ae_alert_loader.py --from 2025-01-01 --to 2026-06-08

4. Export to Excel without touching the DB:
       python ae_alert_loader.py --from 2025-01-01 --to 2026-06-08 --excel alerts.xlsx

When calls return 401/403, re-copy a fresh cURL and rerun. Idempotent (upsert).
"""

import argparse
import json
import re
import shlex
import sys
import time
from datetime import date, datetime, timedelta

import requests
try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    psycopg2 = None

# ----------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------

CURL_FILE = "alsoenergy_curl.txt"
API_URL = "https://apps.alsoenergy.com/api/view/alerthistory"  # hardcoded; cURL URL is ignored
DB_DSN = "host=localhost port=5432 dbname=solriver user=postgres password=postgres"

# Portfolio ("C...") keys from the PowerTrack tree.
PORTFOLIO_KEYS = [
    "C12941",
    # "C47197",
]

WINDOW_DAYS = 31
SLEEP_BETWEEN_CALLS = 1.5
TZ_OFFSET_MIN = 360  # minutes west of UTC; match offset in cURL body (360 = US/Eastern)

# Columns written to the Excel export (subset of normalize() output keys).
EXCEL_COLUMNS = [
    "alert_id", "site_name", "hardware_name", "event_type_name",
    "description", "severity", "is_resolved", "is_acknowledged",
    "alert_start", "alert_end", "resolved_time", "timezone",
    "asset_code", "impact", "capacity", "resolved_by",
]

# ----------------------------------------------------------------------
# cURL PARSING
# ----------------------------------------------------------------------

def parse_curl(path):
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()
    # handle both Windows (\r\n) and Unix (\n) line continuations
    raw = raw.replace("\\\r\n", " ").replace("\\\n", " ")
    tokens = shlex.split(raw)

    url, cookie = None, None
    headers = {}
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if t in ("-H", "--header"):
            k, _, v = tokens[i + 1].partition(":")
            headers[k.strip()] = v.strip()
            i += 2
        elif t in ("-b", "--cookie"):
            cookie = tokens[i + 1]
            i += 2
        elif t in ("--data-raw", "--data", "-d"):
            i += 2  # discard the captured body; we build our own
        elif t.startswith("http"):
            url = t
            i += 1
        else:
            i += 1

    if not url:
        sys.exit("Could not find a URL in the cURL file.")
    return url.split("?")[0], headers, cookie


def build_session(headers, cookie):
    s = requests.Session()
    # Filter out static Cookie header — the jar handles cookies to avoid duplication
    clean = {k: v for k, v in headers.items() if k.lower() != "cookie"}
    s.headers.update(clean)
    if cookie:
        for part in cookie.split("; "):
            if "=" in part:
                name, _, value = part.partition("=")
                s.cookies.set(name, value)
    return s

# ----------------------------------------------------------------------
# FETCH
# ----------------------------------------------------------------------

def fetch(session, base_url, key, d_from=None, d_to=None, since=None):
    if since:
        # Build URL manually to avoid requests percent-encoding the colons
        ts = since.strftime("%Y-%m-%dT%H:%M:%SZ") if isinstance(since, datetime) else since
        url = f"{base_url}?lastChanged={ts}"
    else:
        url = base_url

    payload = {"key": key, "offset": TZ_OFFSET_MIN}
    if d_from:
        payload["from"] = d_from.isoformat()
    if d_to:
        payload["to"] = d_to.isoformat()

    r = session.post(url, json=payload, timeout=60)
    if r.status_code == 204:
        return {}  # no alerts changed since lastChanged — not an error
    if r.status_code in (401, 403):
        sys.exit(f"Auth failed ({r.status_code}). Re-copy a fresh cURL into "
                 f"{CURL_FILE} and rerun.")
    if not r.ok:
        print(f"  HTTP {r.status_code} — response body: {r.text[:500]}", file=sys.stderr)
    r.raise_for_status()
    return r.json()


def extract_records(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for k in ("list", "data", "alerts", "items", "results", "rows"):
            if isinstance(payload.get(k), list):
                return payload[k]
        if payload:
            print(f"  [warn] unrecognised response shape — top-level keys: "
                  f"{list(payload.keys())}", file=sys.stderr)
    return []

# ----------------------------------------------------------------------
# NORMALIZE
# ----------------------------------------------------------------------

def parse_ts(v):
    """Parse ISO-8601 timestamps returned by the API (UTC 'Z' suffix)."""
    if not v:
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y %I:%M:%S %p"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    print(f"  [warn] could not parse timestamp: {v!r}", file=sys.stderr)
    return None


def split_event_code(s):
    # "614 - Rule Tool Alert" -> (614, "Rule Tool Alert")
    if not s:
        return None, None
    m = re.match(r"^\s*(\d+)\s*-\s*(.+)$", str(s))
    if m:
        return int(m.group(1)), m.group(2).strip()
    return None, str(s).strip()


def normalize(rec):
    event_type_code, event_type_name = split_event_code(rec.get("eventCode"))
    return {
        "alert_id":         str(rec.get("alertId", "")),
        "alert_key":        rec.get("key"),
        "site_key":         rec.get("siteKey"),
        "site_name":        rec.get("siteName"),
        "hardware_key":     rec.get("hardwareKey") or None,
        "hardware_name":    rec.get("hardwareName") or None,
        "asset_code":       rec.get("assetCode") or None,
        "category":         rec.get("category"),
        "event_type_code":  event_type_code,
        "event_type_name":  event_type_name,
        "name":             rec.get("name"),
        "description":      rec.get("description"),
        "severity":         rec.get("severity"),
        "impact":           rec.get("impact"),
        "impact_code":      rec.get("impactCode"),
        "capacity":         rec.get("capacity"),
        "is_acknowledged":  bool(rec.get("isAcknowledged")),
        "is_resolved":      bool(rec.get("isResolved")),
        "acknowledged_by":  rec.get("acknowledgedByName") or None,
        "resolved_by":      rec.get("resolvedByName") or None,
        "alert_start":      parse_ts(rec.get("start")),
        "alert_end":        parse_ts(rec.get("end")),
        "resolved_time":    parse_ts(rec.get("resolvedTime")),
        "trigger_time":     parse_ts(rec.get("triggerTime")),
        "last_changed":     parse_ts(rec.get("lastChanged")),
        "timezone":         rec.get("tz"),
    }

# ----------------------------------------------------------------------
# DATABASE
# ----------------------------------------------------------------------

DDL = """
CREATE TABLE IF NOT EXISTS ae_alerts (
    id              BIGSERIAL PRIMARY KEY,
    alert_id        TEXT UNIQUE NOT NULL,
    alert_key       TEXT,
    site_key        TEXT,
    site_name       TEXT,
    hardware_key    TEXT,
    hardware_name   TEXT,
    asset_code      TEXT,
    category        INT,
    event_type_code INT,
    event_type_name TEXT,
    name            TEXT,
    description     TEXT,
    severity        INT,
    impact          FLOAT,
    impact_code     INT,
    capacity        INT,
    is_acknowledged BOOLEAN NOT NULL DEFAULT FALSE,
    is_resolved     BOOLEAN NOT NULL DEFAULT FALSE,
    acknowledged_by TEXT,
    resolved_by     TEXT,
    alert_start     TIMESTAMP,
    alert_end       TIMESTAMP,
    resolved_time   TIMESTAMP,
    trigger_time    TIMESTAMP,
    last_changed    TIMESTAMP,
    timezone        TEXT
);
CREATE INDEX IF NOT EXISTS idx_ae_alerts_site  ON ae_alerts (site_key);
CREATE INDEX IF NOT EXISTS idx_ae_alerts_hw    ON ae_alerts (hardware_key);
CREATE INDEX IF NOT EXISTS idx_ae_alerts_start ON ae_alerts (alert_start);
CREATE INDEX IF NOT EXISTS idx_ae_alerts_event ON ae_alerts (event_type_code);
"""

UPSERT = """
INSERT INTO ae_alerts
 (alert_id, alert_key, site_key, site_name, hardware_key, hardware_name,
  asset_code, category, event_type_code, event_type_name, name, description,
  severity, impact, impact_code, capacity, is_acknowledged, is_resolved,
  acknowledged_by, resolved_by, alert_start, alert_end, resolved_time,
  trigger_time, last_changed, timezone)
VALUES
 (%(alert_id)s, %(alert_key)s, %(site_key)s, %(site_name)s, %(hardware_key)s,
  %(hardware_name)s, %(asset_code)s, %(category)s, %(event_type_code)s,
  %(event_type_name)s, %(name)s, %(description)s, %(severity)s, %(impact)s,
  %(impact_code)s, %(capacity)s, %(is_acknowledged)s, %(is_resolved)s,
  %(acknowledged_by)s, %(resolved_by)s, %(alert_start)s, %(alert_end)s,
  %(resolved_time)s, %(trigger_time)s, %(last_changed)s, %(timezone)s)
ON CONFLICT (alert_id) DO UPDATE SET
  is_resolved     = EXCLUDED.is_resolved,
  is_acknowledged = EXCLUDED.is_acknowledged,
  alert_end       = EXCLUDED.alert_end,
  resolved_time   = EXCLUDED.resolved_time,
  resolved_by     = EXCLUDED.resolved_by,
  severity        = EXCLUDED.severity,
  description     = EXCLUDED.description,
  last_changed    = EXCLUDED.last_changed
WHERE ae_alerts.last_changed IS NULL
   OR EXCLUDED.last_changed IS NULL
   OR EXCLUDED.last_changed > ae_alerts.last_changed;
"""


def ensure_schema(conn):
    with conn.cursor() as cur:
        cur.execute(DDL)
    conn.commit()


def upsert_rows(conn, rows):
    with conn.cursor() as cur:
        psycopg2.extras.execute_batch(cur, UPSERT, rows, page_size=500)
    conn.commit()

# ----------------------------------------------------------------------
# EXCEL EXPORT
# ----------------------------------------------------------------------

def export_excel(all_rows, path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("openpyxl is required for Excel export: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alerts"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    for col_idx, col in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(all_rows, start=2):
        for col_idx, col in enumerate(EXCEL_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col))

    # Auto-size columns (cap at 60 chars wide)
    for col_idx, col in enumerate(EXCEL_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"Saved {len(all_rows)} rows to {path}")

# ----------------------------------------------------------------------
# WINDOWS
# ----------------------------------------------------------------------

def windows(d_from, d_to, step):
    cur = d_from
    while cur <= d_to:
        end = min(cur + timedelta(days=step - 1), d_to)
        yield cur, end
        cur = end + timedelta(days=1)

# ----------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--probe", action="store_true",
                    help="dump the response shape for one portfolio/day and exit")
    ap.add_argument("--key", help="portfolio key (probe mode)")
    ap.add_argument("--date", help="YYYY-MM-DD (probe mode)")
    ap.add_argument("--from", dest="d_from", help="YYYY-MM-DD")
    ap.add_argument("--to", dest="d_to", help="YYYY-MM-DD")
    ap.add_argument("--since", metavar="YYYY-MM-DDTHH:MM:SSZ",
                    help="incremental mode: only fetch alerts changed after this timestamp")
    ap.add_argument("--excel", metavar="FILE.xlsx",
                    help="also write results to an Excel file")
    ap.add_argument("--excel-only", action="store_true",
                    help="write Excel only, skip DB load")
    args = ap.parse_args()

    from ae_auth import get_session
    session = get_session()

    if args.probe:
        key = args.key or PORTFOLIO_KEYS[0]
        d = date.fromisoformat(args.date) if args.date else date.today()
        payload = fetch(session, API_URL, key, d_from=d, d_to=d)
        recs = extract_records(payload)
        print("Top-level type:", type(payload).__name__)
        if isinstance(payload, dict):
            print("Top-level keys:", list(payload.keys()))
        print("Record count:", len(recs))
        if recs:
            print("\nFirst record keys:")
            print(json.dumps(list(recs[0].keys()), indent=2))
            print("\nFirst record sample:")
            print(json.dumps(recs[0], indent=2, default=str)[:2000])
        return

    if not (args.d_from and args.d_to) and not args.since:
        sys.exit("Provide --from and --to for a full load, --since for incremental, or --probe.")
    d_from = date.fromisoformat(args.d_from) if args.d_from else None
    d_to = date.fromisoformat(args.d_to) if args.d_to else None

    conn = None
    if not args.excel_only:
        if psycopg2 is None:
            sys.exit("psycopg2 is not installed. Run: pip install psycopg2-binary")
        conn = psycopg2.connect(DB_DSN)
        ensure_schema(conn)

    all_rows = []
    total = 0
    last_changed_cursor = None  # track for incremental --since hint

    for key in PORTFOLIO_KEYS:
        if args.since:
            # single incremental request per portfolio key
            payload = fetch(session, API_URL, key, since=args.since)
            window_iter = [(None, None)]
        else:
            window_iter = list(windows(d_from, d_to, WINDOW_DAYS))

        for w_start, w_end in window_iter:
            if args.since:
                pass  # payload already fetched above
            else:
                payload = fetch(session, API_URL, key, d_from=w_start, d_to=w_end)

            # capture the cursor for next incremental run
            if isinstance(payload, dict) and payload.get("lastChanged"):
                last_changed_cursor = payload["lastChanged"]

            recs = extract_records(payload)
            rows = [normalize(r) for r in recs]
            bad = sum(1 for r in rows if not r["alert_id"])
            if bad:
                print(f"  [warn] {bad} records missing alertId, skipping",
                      file=sys.stderr)
            rows = [r for r in rows if r["alert_id"]]
            if rows:
                if conn:
                    upsert_rows(conn, rows)
                all_rows.extend(rows)
                total += len(rows)
            window_label = f"{w_start}..{w_end}" if w_start else f"since {args.since}"
            print(f"{key} {window_label}: {len(rows)} rows (running total {total})")
            if not args.since:
                time.sleep(SLEEP_BETWEEN_CALLS)

    if conn:
        conn.close()
        print(f"\nUpserted {total} alert rows into ae_alerts.")

    if last_changed_cursor:
        print(f"\nNext incremental run:")
        print(f"  python ae_alert_loader.py --since {last_changed_cursor} --excel-only")

    if args.excel or args.excel_only:
        export_excel(all_rows, args.excel or "ae_alerts.xlsx")


if __name__ == "__main__":
    main()

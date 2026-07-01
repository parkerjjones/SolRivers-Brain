#!/usr/bin/env python3
"""
Reads and deeply analyzes AlsoEnergy AI site summaries.
Fetches fresh text for every site, writes ae_ai_deep.xlsx and ae_ai_deep.txt.

The AI generates natural language covering:
  - Hardware communication status per device
  - Production vs expected output with % figures
  - Weather context (temperature, cloud cover, rain)
  - Active faults, alerts, error codes
  - Links to relevant chart views
  - Trend commentary (improving / degrading)
"""

import re, shlex, time, json, sys
from datetime import datetime, timezone
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CURL_FILE = "alsoenergy_curl.txt"
API_BASE  = "https://apps.alsoenergy.com/api"
PORTFOLIO = "C12941"
SLEEP     = 0.8

DARK, MID, WHITE, ALT = "1F4E79", "2E75B6", "FFFFFF", "EBF3FB"
RED, GREEN, YELL = "FFC7CE", "C6EFCE", "FFEB9C"

def build_session():
    from ae_auth import get_session
    return get_session()


def fetch_full_summary(session, site_id):
    url = f"{API_BASE}/ai-site-summary?siteId={site_id}&lang=en-US"
    session.headers["accept"] = "text/event-stream"
    try:
        r = session.get(url, timeout=30, stream=True)
        if not r.ok:
            return "", False, ""
        chunks, cached, generated = [], False, ""
        for line in r.iter_lines(decode_unicode=True):
            if not line.startswith("data:"):
                continue
            try:
                obj = json.loads(line[5:].strip())
                if "metadata" in obj:
                    cached = obj["metadata"].get("cached", False)
                    generated = obj["metadata"].get("generatedAt", "")
                elif "chunk" in obj:
                    chunks.append(obj["chunk"])
            except Exception:
                pass
        return "".join(chunks), cached, generated
    finally:
        session.headers["accept"] = "application/json"


def clean(text):
    """Strip markdown links but keep label and URL as readable text."""
    # [label](url) → "label [url]"
    text = re.sub(r'\[([^\]]+)\]\((https?://[^\)]+)\)', r'\1', text)
    return text.replace("\\n", "\n").strip()


def extract_facts(text):
    """Pull structured facts from the AI text."""
    t = text.lower()
    facts = {}

    # Numbers
    m = re.search(r'([\d,]+(?:\.\d+)?)\s*kwh?', t)
    facts["energy_kwh"] = float(m.group(1).replace(",","")) if m else None

    m = re.search(r'([\d,]+(?:\.\d+)?)\s*kw\b', t)
    facts["power_kw"] = float(m.group(1).replace(",","")) if m else None

    m = re.search(r'(\d+(?:\.\d+)?)\s*%\s*(?:of\s+expected|expected)', t)
    facts["pct_of_expected"] = float(m.group(1)) if m else None

    m = re.search(r'([-\d.]+)\s*[°]?f\b', t)
    facts["temp_f"] = float(m.group(1)) if m else None

    m = re.search(r'([-\d.]+)\s*[°]?c\b', t)
    facts["temp_c"] = float(m.group(1)) if m else None

    # Status flags
    facts["all_communicating"]  = "all hardware is communicating" in t or "all communicating" in t
    facts["comm_issues"]        = any(w in t for w in ["not communicating","not responding","offline","intermittent"])
    facts["production_low"]     = any(w in t for w in ["tracking low","below expected","production loss","underperforming"])
    facts["production_ok"]      = any(w in t for w in ["in line with expected","on track","performing as expected"])
    facts["has_fault"]          = any(w in t for w in ["fault","faulted","tripped"])
    facts["has_alert"]          = "alert" in t and "no active alert" not in t
    facts["weather_sunny"]      = any(w in t for w in ["sunny","clear sky","clear skies","sunny day"])
    facts["weather_cloudy"]     = any(w in t for w in ["cloudy","overcast","cloud cover","partly cloudy"])
    facts["weather_rain"]       = any(w in t for w in ["rain","raining","rainy","precipitation"])
    facts["inverter_issue"]     = any(w in t for w in ["inverter","string inverter","central inverter"])
    facts["meter_issue"]        = "meter" in t and any(w in t for w in ["error","check","issue","fail"])

    # Extract device names mentioned
    device_re = re.compile(
        r'(?:inverter|meter|tracker|gateway|logger|modem|weather station|pyranometer)'
        r'[\w\s\-\.]*', re.IGNORECASE
    )
    facts["devices"] = list({m.strip() for m in device_re.findall(text) if len(m.strip()) > 3})[:6]

    return facts


def write_xlsx(path, records):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Sheet 1: Full AI Text ---
    ws1 = wb.create_sheet("Full AI Summaries")
    for ci, h in enumerate(["Site Key","Site Name","Open Alerts","Generated At","Cached",
                             "Full Summary (Plain Text)"], start=1):
        c = ws1.cell(1, ci, h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=MID)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.freeze_panes = "A2"
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 8
    ws1.column_dimensions["D"].width = 20
    ws1.column_dimensions["E"].width = 8
    ws1.column_dimensions["F"].width = 110

    for ri, rec in enumerate(records, start=2):
        alt = ri % 2
        bg = ALT if alt == 0 else None
        for ci, v in enumerate([rec["site_key"], rec["site_name"], rec["open_alerts"],
                                  rec["generated_at"], "Yes" if rec["cached"] else "No",
                                  rec["plain_text"]], start=1):
            cell = ws1.cell(ri, ci, v)
            if bg: cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci==6))
        ws1.row_dimensions[ri].height = min(max(30, len(rec["plain_text"])//8), 400)

    # --- Sheet 2: Extracted Facts ---
    ws2 = wb.create_sheet("Extracted Facts")
    fact_cols = ["site_key","site_name","open_alerts","all_communicating","comm_issues",
                 "production_low","production_ok","has_fault","has_alert",
                 "weather_sunny","weather_cloudy","weather_rain",
                 "inverter_issue","meter_issue",
                 "energy_kwh","power_kw","pct_of_expected","temp_f","devices_mentioned"]
    for ci, h in enumerate(fact_cols, start=1):
        c = ws2.cell(1, ci, h)
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=MID)
        c.alignment = Alignment(horizontal="center")
    ws2.freeze_panes = "A2"

    for ri, rec in enumerate(records, start=2):
        f = rec["facts"]
        alt = ri % 2
        # color by alert/comm status
        if f.get("comm_issues") or f.get("has_fault"):
            bg = RED
        elif f.get("production_low"):
            bg = YELL
        elif f.get("all_communicating") and f.get("production_ok"):
            bg = GREEN
        else:
            bg = ALT if alt == 0 else None

        row_vals = [rec["site_key"], rec["site_name"], rec["open_alerts"],
                    "YES" if f.get("all_communicating") else "no",
                    "YES" if f.get("comm_issues") else "no",
                    "YES" if f.get("production_low") else "no",
                    "YES" if f.get("production_ok") else "no",
                    "YES" if f.get("has_fault") else "no",
                    "YES" if f.get("has_alert") else "no",
                    "YES" if f.get("weather_sunny") else "",
                    "YES" if f.get("weather_cloudy") else "",
                    "YES" if f.get("weather_rain") else "",
                    "YES" if f.get("inverter_issue") else "",
                    "YES" if f.get("meter_issue") else "",
                    f.get("energy_kwh"), f.get("power_kw"), f.get("pct_of_expected"), f.get("temp_f"),
                    "; ".join(f.get("devices", [])[:4])]
        for ci, v in enumerate(row_vals, start=1):
            cell = ws2.cell(ri, ci, v)
            if bg: cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="top")

    for col in ws2.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 45)

    # --- Sheet 3: Site Rankings ---
    ws3 = wb.create_sheet("Site Rankings")
    rank_cols = ["Rank","Site Key","Site Name","Open Alerts",
                 "Comm OK","Prod OK","Has Fault","Has Alert","Weather"]
    for ci, h in enumerate(rank_cols, start=1):
        c = ws3.cell(1, ci, h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MID)
        c.alignment = Alignment(horizontal="center")
    ws3.freeze_panes = "A2"

    sorted_recs = sorted(records, key=lambda r: r["open_alerts"], reverse=True)
    for ri, rec in enumerate(sorted_recs, start=2):
        f = rec["facts"]
        weather = "Rainy" if f.get("weather_rain") else ("Cloudy" if f.get("weather_cloudy") else ("Sunny" if f.get("weather_sunny") else ""))
        bg = RED if rec["open_alerts"] >= 3 else (YELL if rec["open_alerts"] >= 1 else GREEN)
        for ci, v in enumerate([ri-1, rec["site_key"], rec["site_name"], rec["open_alerts"],
                                  "OK" if f.get("all_communicating") else ("ISSUE" if f.get("comm_issues") else "?"),
                                  "OK" if f.get("production_ok") else ("LOW" if f.get("production_low") else "?"),
                                  "YES" if f.get("has_fault") else "",
                                  "YES" if f.get("has_alert") else "",
                                  weather], start=1):
            cell = ws3.cell(ri, ci, v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="top")

    for col in ws3.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws3.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 40)

    wb.save(path)
    print(f"Saved -> {path}  ({len(records)} sites)")


def write_txt(path, records):
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"AlsoEnergy AI Site Summaries\n")
        f.write(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n")
        f.write("=" * 70 + "\n\n")
        for rec in sorted(records, key=lambda r: r["open_alerts"], reverse=True):
            f.write(f"{'=' * 70}\n")
            f.write(f"SITE: {rec['site_name']} ({rec['site_key']})  |  Open Alerts: {rec['open_alerts']}\n")
            f.write(f"Generated: {rec['generated_at']}  |  Cached: {rec['cached']}\n")
            f.write(f"{'-' * 70}\n")
            f.write(rec["plain_text"] + "\n\n")
    print(f"Saved -> {path}")


if __name__ == "__main__":
    import pandas as pd

    session = build_session()

    # Load real open-alert counts from master (missing → 0, meaning none recorded)
    alert_map = {}
    try:
        mdf = pd.read_excel("ae_master.xlsx", sheet_name="Portfolio Overview")
        for _, row in mdf.iterrows():
            sk = str(row.get("site_key", ""))
            oa = float(row.get("open_alerts", 0) or 0)
            alert_map[sk] = int(oa)
    except Exception:
        pass

    print("Fetching portfolio site list...")
    r = session.get(f"{API_BASE}/view/portfolio/{PORTFOLIO}", timeout=20)
    if r.status_code in (401, 403):
        sys.exit("Auth expired — paste fresh cURL")
    sites = r.json().get("sites", [])
    print(f"  {len(sites)} sites")

    records = []
    for i, site in enumerate(sites):
        sk, sn = site["key"], site.get("name", "")
        open_alerts = alert_map.get(sk, 0)
        print(f"  [{i+1:2}/{len(sites)}] {sk} {sn}  (open_alerts={open_alerts})")
        text, cached, generated_at = fetch_full_summary(session, sk)
        plain = clean(text)
        records.append({
            "site_key":     sk,
            "site_name":    sn,
            "open_alerts":  open_alerts,
            "cached":       cached,
            "generated_at": generated_at[:16] if generated_at else "",
            "raw_text":     text,
            "plain_text":   plain,
            "facts":        extract_facts(text),
        })
        time.sleep(SLEEP)

    write_xlsx("ae_ai_deep.xlsx", records)
    write_txt("ae_ai_deep.txt", records)

    print("\n--- Quick Facts ---")
    comm_ok    = sum(1 for r in records if r["facts"]["all_communicating"])
    comm_issue = sum(1 for r in records if r["facts"]["comm_issues"])
    prod_low   = sum(1 for r in records if r["facts"]["production_low"])
    has_fault  = sum(1 for r in records if r["facts"]["has_fault"])
    rainy      = sum(1 for r in records if r["facts"]["weather_rain"])
    print(f"  All communicating: {comm_ok}/35")
    print(f"  Comm issues:       {comm_issue}/35")
    print(f"  Production low:    {prod_low}/35")
    print(f"  Active faults:     {has_fault}/35")
    print(f"  Weather: rain:     {rainy}/35")

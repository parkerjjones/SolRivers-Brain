#!/usr/bin/env python3
"""
Fetches Rule Tool diagnostic results for the portfolio.

The Rule Tool runs automated checks across all sites and hardware —
communication tests, performance tests, data quality checks, etc.
Results include pass/fail per suite and per site.

Uses:
  GET /api/ruleresults/{portfolioKey}?lastChanged=1900-01-01T00:00:00.000Z&mergeHash=

Writes ae_ruleresults.xlsx with:
  - Site Pass/Fail     : high-level pass/fail per site
  - Rule Details       : individual rule results per site
  - Failing Sites      : sites with at least one failure
  - Weather Conditions : weather snapshot at time of last rule run

USAGE
-----
    python ae_ruleresults_loader.py
"""

import argparse
import json
import shlex
import sys
import time
from datetime import datetime

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CURL_FILE = "alsoenergy_curl.txt"
API_BASE  = "https://apps.alsoenergy.com/api"
PORTFOLIO = "C12941"

DARK  = "1F4E79"
MID   = "2E75B6"
WHITE = "FFFFFF"
ALT   = "EBF3FB"
GREEN = "C6EFCE"
RED   = "FFC7CE"
YELL  = "FFEB9C"

# result integer codes
RESULT_CODE = {1: "N/A", 2: "Warning", 3: "Fail", 4: "Pass"}
# weatherCondition integer codes (observed: 1=sunny,2=partly cloudy,3=overcast,5=rain)
WEATHER_CODE = {1: "Sunny", 2: "Partly Cloudy", 3: "Overcast", 4: "Fog", 5: "Rain", 6: "Snow"}


def build_session():
    with open(CURL_FILE, "r", encoding="utf-8") as f:
        raw = f.read().replace("\\\r\n", " ").replace("\\\n", " ")
    tokens = shlex.split(raw)
    headers, cookie = {}, None
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if t in ("-H", "--header"):
            k, _, v = tokens[i + 1].partition(":")
            headers[k.strip()] = v.strip()
            i += 2
        elif t in ("-b", "--cookie"):
            cookie = tokens[i + 1]
            i += 2
        else:
            i += 1
    s = requests.Session()
    s.headers.update({k: v for k, v in headers.items() if k.lower() != "cookie"})
    if cookie:
        for part in cookie.split("; "):
            if "=" in part:
                n, _, v = part.partition("=")
                s.cookies.set(n, v)
    return s


def hdr(cell, text, bg=DARK):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def val(cell, v, row_idx=0, color=None):
    cell.value = v
    bg = color or (ALT if row_idx % 2 == 0 else None)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def autosize(ws, max_w=60):
    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, max_w)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", default="ae_ruleresults.xlsx")
    args = ap.parse_args()

    session = build_session()

    print("Fetching rule results...")
    url = f"{API_BASE}/ruleresults/{PORTFOLIO}?lastChanged=1900-01-01T00:00:00.000Z&mergeHash="
    r = session.get(url, timeout=30)
    if r.status_code in (401, 403):
        sys.exit("Auth expired — re-copy cURL.")
    r.raise_for_status()
    data = r.json()

    print(f"  Last run:     {data.get('latestRun')}")
    print(f"  Last changed: {data.get('lastChanged')}")
    sites = data.get("sites", [])
    print(f"  Sites:        {len(sites)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Site Pass/Fail ───────────────────────────────────────
    ws = wb.create_sheet("Site Pass-Fail")
    site_cols = ["Site Key", "Site Name", "Overall", "Run At",
                 "Weather Condition", "Rules Count", "Passed", "Failed",
                 "Comm Pass", "Comm Fail", "Perf Pass", "Perf Fail"]
    for ci, c in enumerate(site_cols, 1):
        hdr(ws.cell(1, ci), c, MID)
    ws.freeze_panes = "A2"

    for ri, site in enumerate(sites, start=2):
        alt = ri % 2
        results = site.get("results", [])
        show_results = [x for x in results if x.get("show", True)]
        passed_r = [x for x in show_results if x.get("result") == 4]
        failed_r = [x for x in show_results if x.get("result") == 3]
        warn_r   = [x for x in show_results if x.get("result") == 2]
        comm_r   = [x for x in show_results if "comm" in x.get("id","").lower() or "network" in x.get("id","").lower()]
        perf_r   = [x for x in show_results if "perf" in x.get("id","").lower() or "inverter" in x.get("id","").lower()]

        passed_count = site.get("passed", len(passed_r))
        total_show = len(show_results)
        overall_ok = len(failed_r) == 0 and total_show > 0
        row_color = GREEN if overall_ok else (RED if failed_r else None)
        weather_label = WEATHER_CODE.get(site.get("weatherCondition"), str(site.get("weatherCondition", "")))

        val(ws.cell(ri, 1),  site.get("key", ""), alt, row_color)
        val(ws.cell(ri, 2),  site.get("name", ""), alt, row_color)
        val(ws.cell(ri, 3),  "PASS" if overall_ok else ("FAIL" if failed_r else "?"), alt, row_color)
        val(ws.cell(ri, 4),  site.get("runAt", ""), alt, row_color)
        val(ws.cell(ri, 5),  weather_label, alt, row_color)
        val(ws.cell(ri, 6),  total_show, alt, row_color)
        val(ws.cell(ri, 7),  len(passed_r), alt, row_color)
        val(ws.cell(ri, 8),  len(failed_r), alt, row_color)
        val(ws.cell(ri, 9),  sum(1 for x in comm_r if x.get("result") == 4), alt, row_color)
        val(ws.cell(ri, 10), sum(1 for x in comm_r if x.get("result") == 3), alt, row_color)
        val(ws.cell(ri, 11), sum(1 for x in perf_r if x.get("result") == 4), alt, row_color)
        val(ws.cell(ri, 12), sum(1 for x in perf_r if x.get("result") == 3), alt, row_color)

    autosize(ws)

    # ── Sheet 2: Individual Rule Results ─────────────────────────────
    ws2 = wb.create_sheet("Rule Details")
    rule_cols = ["Site Key", "Site Name", "Rule Name", "Type", "Passed",
                 "Message", "Score", "Last Changed"]
    for ci, c in enumerate(rule_cols, 1):
        hdr(ws2.cell(1, ci), c, MID)
    ws2.freeze_panes = "A2"

    row2 = 2
    for site in sites:
        sk, sn = site.get("key", ""), site.get("name", "")
        for res in site.get("results", []):
            if not res.get("show", True):
                continue
            alt = row2 % 2
            rc = res.get("result", 0)
            color = GREEN if rc == 4 else (RED if rc == 3 else (YELL if rc == 2 else None))
            val(ws2.cell(row2, 1), sk, alt, color)
            val(ws2.cell(row2, 2), sn, alt, color)
            val(ws2.cell(row2, 3), res.get("id", ""), alt, color)
            val(ws2.cell(row2, 4), RESULT_CODE.get(rc, str(rc)), alt, color)
            val(ws2.cell(row2, 5), RESULT_CODE.get(rc, str(rc)), alt, color)
            val(ws2.cell(row2, 6), res.get("detail", ""), alt, color)
            val(ws2.cell(row2, 7), res.get("success"), alt, color)
            val(ws2.cell(row2, 8), ", ".join(str(e) for e in res.get("errors", [])[:3]), alt, color)
            row2 += 1

    autosize(ws2)

    # ── Sheet 3: Failing Sites (quick action list) ────────────────────
    ws3 = wb.create_sheet("Failing Sites")
    fail_cols = ["Site Key", "Site Name", "Failing Rules", "Weather", "Run At"]
    for ci, c in enumerate(fail_cols, 1):
        hdr(ws3.cell(1, ci), c, MID)
    ws3.freeze_panes = "A2"

    row3 = 2
    for site in sites:
        failed = [r for r in site.get("results", []) if r.get("result") == 3 and r.get("show", True)]
        if not failed:
            continue
        alt = row3 % 2
        fail_names = "; ".join(r.get("id", "") for r in failed[:5])
        if len(failed) > 5:
            fail_names += f" (+{len(failed)-5} more)"
        val(ws3.cell(row3, 1), site.get("key", ""), alt, RED)
        val(ws3.cell(row3, 2), site.get("name", ""), alt, RED)
        val(ws3.cell(row3, 3), fail_names, alt, RED)
        val(ws3.cell(row3, 4), site.get("weatherCondition", ""), alt, RED)
        val(ws3.cell(row3, 5), site.get("runAt", ""), alt, RED)
        row3 += 1

    autosize(ws3)

    # ── Sheet 4: Raw JSON dump (for exploration) ──────────────────────
    ws4 = wb.create_sheet("Raw Sample")
    ws4["A1"].value = "First site raw JSON (for exploration):"
    ws4["A1"].font = Font(bold=True)
    if sites:
        raw_str = json.dumps(sites[0], indent=2, default=str)
        for li, line in enumerate(raw_str.split("\n"), start=2):
            ws4.cell(li, 1).value = line
            ws4.cell(li, 1).font = Font(name="Courier New", size=9)
    ws4.column_dimensions["A"].width = 80

    wb.save(args.output)

    # Console summary
    total = len(sites)
    passed = sum(1 for s in sites if s.get("passed"))
    failed = sum(1 for s in sites if not s.get("passed") and s.get("passed") is not None)
    print(f"\nResults: {passed}/{total} pass, {failed} fail")
    print(f"Saved -> {args.output}")
    failing_names = [s.get("name", s["key"]) for s in sites
                     if any(r.get("result") == 3 for r in s.get("results", []))]
    if failing_names:
        print("Failing sites:", ", ".join(failing_names[:8]))

#!/usr/bin/env python3
"""
AlsoEnergy API explorer + schema design spreadsheet generator.

Probes multiple AlsoEnergy endpoints, prints what comes back,
and writes a multi-sheet Excel file describing the proposed DB schema.

USAGE
-----
    python ae_schema_explorer.py           # probe all + write schema.xlsx
    python ae_schema_explorer.py --probe   # probe only, no spreadsheet
    python ae_schema_explorer.py --schema  # write spreadsheet only
"""

import json
import shlex
import sys
import time
from datetime import date, datetime

import requests

# -----------------------------------------------------------------------
# CONFIG
# -----------------------------------------------------------------------

CURL_FILE  = "alsoenergy_curl.txt"
API_BASE   = "https://apps.alsoenergy.com/api"
PORTFOLIO  = "C12941"
SAMPLE_SITE = "S40834"          # Green Elementary — used for drill-downs
SAMPLE_HW   = "H119159"         # Elkor MKII Production Meter
SLEEP       = 1.0

OUTPUT_FILE = "ae_schema.xlsx"

# -----------------------------------------------------------------------
# cURL PARSING  (same as ae_alert_loader.py)
# -----------------------------------------------------------------------

def parse_curl(path):
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()
    raw = raw.replace("\\\r\n", " ").replace("\\\n", " ")
    tokens = shlex.split(raw)
    url, cookie = None, None
    headers = {}
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if t in ("-H", "--header"):
            k, _, v = tokens[i + 1].partition(":")
            headers[k.strip()] = v.strip()
            i += 2
        elif t in ("-b", "--cookie"):
            cookie = tokens[i + 1]
            i += 2
        elif t in ("--data-raw", "--data", "-d"):
            i += 2
        elif t.startswith("http"):
            url = t
            i += 1
        else:
            i += 1
    return headers, cookie


def build_session(headers, cookie):
    s = requests.Session()
    clean = {k: v for k, v in headers.items() if k.lower() != "cookie"}
    s.headers.update(clean)
    if cookie:
        for part in cookie.split("; "):
            if "=" in part:
                name, _, value = part.partition("=")
                s.cookies.set(name, value)
    return s

# -----------------------------------------------------------------------
# PROBING
# -----------------------------------------------------------------------

def probe(session, url, method="GET", body=None, label=None):
    label = label or url
    try:
        if method == "POST":
            r = session.post(url, json=body, timeout=30)
        else:
            r = session.get(url, timeout=30)
        print(f"\n{'='*60}")
        print(f"  {method} {label}  -> HTTP {r.status_code}")
        if r.ok and r.text.strip():
            try:
                data = r.json()
                if isinstance(data, list):
                    print(f"  List of {len(data)} items")
                    if data:
                        print(f"  First item keys: {list(data[0].keys()) if isinstance(data[0], dict) else type(data[0]).__name__}")
                        print(f"  Sample:\n{json.dumps(data[0], indent=4, default=str)[:800]}")
                elif isinstance(data, dict):
                    print(f"  Dict keys: {list(data.keys())}")
                    for k, v in data.items():
                        if isinstance(v, list):
                            print(f"    '{k}': list of {len(v)}", end="")
                            if v and isinstance(v[0], dict):
                                print(f"  keys={list(v[0].keys())[:8]}")
                            else:
                                print()
                        else:
                            print(f"    '{k}': {str(v)[:120]}")
                return data
            except Exception:
                print(f"  (non-JSON) {r.text[:300]}")
        else:
            print(f"  (empty or error) {r.text[:200]}")
    except Exception as e:
        print(f"  ERROR: {e}")
    return None


def run_probes(session):
    results = {}
    today = date.today().isoformat()
    yesterday = "2026-06-08"

    # ---- Alert history (known working) --------------------------------
    results["alerts"] = probe(
        session,
        f"{API_BASE}/view/alerthistory",
        "POST",
        {"key": PORTFOLIO, "from": yesterday, "to": today, "offset": 360},
        "alerthistory (portfolio)"
    )
    time.sleep(SLEEP)

    # ---- Portfolio / tree hierarchy -----------------------------------
    for path in [
        f"/view/tree/{PORTFOLIO}",
        f"/view/portfolio/{PORTFOLIO}",
        f"/portfolio/{PORTFOLIO}",
        f"/view/customer/{PORTFOLIO}",
    ]:
        results[f"tree_{path}"] = probe(session, API_BASE + path, label=path)
        time.sleep(SLEEP)

    # ---- Sites list ---------------------------------------------------
    for path in [
        f"/view/sites?key={PORTFOLIO}",
        f"/view/sites/{PORTFOLIO}",
        f"/site/list?portfolio={PORTFOLIO}",
    ]:
        results[f"sites_{path}"] = probe(session, API_BASE + path, label=path)
        time.sleep(SLEEP)

    # ---- Site detail --------------------------------------------------
    for path in [
        f"/view/site/{SAMPLE_SITE}",
        f"/site/{SAMPLE_SITE}",
        f"/view/siteinfo/{SAMPLE_SITE}",
    ]:
        results[f"site_{path}"] = probe(session, API_BASE + path, label=path)
        time.sleep(SLEEP)

    # ---- Hardware / devices ------------------------------------------
    for path in [
        f"/view/devices?siteKey={SAMPLE_SITE}",
        f"/view/hardware?siteKey={SAMPLE_SITE}",
        f"/hardware/list/{SAMPLE_SITE}",
    ]:
        results[f"hw_{path}"] = probe(session, API_BASE + path, label=path)
        time.sleep(SLEEP)

    # ---- Production / energy -----------------------------------------
    prod_body = {"key": PORTFOLIO, "from": yesterday, "to": today, "interval": "day", "offset": 360}
    for path in [
        "/view/production",
        "/view/energy",
        "/view/performance",
        "/view/summary",
    ]:
        results[f"prod_{path}"] = probe(session, API_BASE + path, "POST", prod_body, path)
        time.sleep(SLEEP)

    # ---- Site-level production ---------------------------------------
    site_body = {"key": SAMPLE_SITE, "from": yesterday, "to": today, "interval": "hour", "offset": 360}
    for path in [
        "/view/production",
        "/view/siteproduction",
        "/view/energydata",
    ]:
        results[f"sitedata_{path}"] = probe(
            session, API_BASE + path, "POST", site_body, f"{path} (site)"
        )
        time.sleep(SLEEP)

    # ---- Work orders / service requests ------------------------------
    for path in [
        f"/view/workorders?key={PORTFOLIO}",
        f"/view/servicerequests?key={PORTFOLIO}",
        f"/view/notes?key={PORTFOLIO}",
    ]:
        results[f"wo_{path}"] = probe(session, API_BASE + path, label=path)
        time.sleep(SLEEP)

    return results

# -----------------------------------------------------------------------
# SCHEMA SPREADSHEET
# -----------------------------------------------------------------------

SCHEMA = {
    "ae_portfolios": {
        "description": "Top-level portfolio / customer accounts",
        "source": "Manually populated or from tree endpoint",
        "columns": [
            ("portfolio_key", "TEXT PK", "AlsoEnergy portfolio ID, e.g. C12941"),
            ("portfolio_name", "TEXT", "Display name of the portfolio"),
            ("created_at", "TIMESTAMP", "Row insert timestamp"),
        ],
    },
    "ae_sites": {
        "description": "Individual solar sites within a portfolio",
        "source": "Site list / tree endpoint",
        "columns": [
            ("site_key",       "TEXT PK",    "AlsoEnergy site ID, e.g. S40834"),
            ("portfolio_key",  "TEXT FK",    "Parent portfolio"),
            ("site_name",      "TEXT",       "Human-readable site name, e.g. Green Elementary"),
            ("capacity_kw",    "FLOAT",      "DC nameplate capacity in kW"),
            ("timezone",       "TEXT",       "IANA tz string, e.g. US/Eastern"),
            ("latitude",       "FLOAT",      "Site latitude"),
            ("longitude",      "FLOAT",      "Site longitude"),
            ("install_date",   "DATE",       "Commission / install date"),
            ("state",          "TEXT",       "US state abbreviation"),
            ("created_at",     "TIMESTAMP",  "Row insert timestamp"),
            ("updated_at",     "TIMESTAMP",  "Last refresh from API"),
        ],
    },
    "ae_hardware": {
        "description": "Devices / equipment at each site",
        "source": "Hardware list endpoint per site",
        "columns": [
            ("hardware_key",  "TEXT PK",   "AlsoEnergy device ID, e.g. H119159"),
            ("site_key",      "TEXT FK",   "Parent site"),
            ("hardware_name", "TEXT",      "Device name, e.g. Elkor MKII Production Meter"),
            ("asset_code",    "TEXT",      "Device class: INV / MTR / MET / UNK etc."),
            ("capacity_kw",   "FLOAT",     "Device rated capacity in kW (NULL for meters/sensors)"),
            ("model",         "TEXT",      "Model / type string from API"),
            ("icon_url",      "TEXT",      "Thumbnail URL from AlsoEnergy"),
            ("created_at",    "TIMESTAMP", "Row insert timestamp"),
        ],
    },
    "ae_alerts": {
        "description": "Operational alert history — one row per alert instance",
        "source": "/api/view/alerthistory POST",
        "columns": [
            ("id",              "BIGSERIAL PK",   "Internal surrogate key"),
            ("alert_id",        "TEXT UNIQUE",    "AlsoEnergy alertId string, e.g. 154085477"),
            ("alert_key",       "TEXT",           "Short record key, e.g. x752f286e"),
            ("site_key",        "TEXT FK",        "Site where alert fired"),
            ("site_name",       "TEXT",           "Denormalized site name for fast queries"),
            ("hardware_key",    "TEXT FK",        "Device that triggered alert (NULL for site-level)"),
            ("hardware_name",   "TEXT",           "Denormalized device name"),
            ("asset_code",      "TEXT",           "INV / MTR / MET / UNK etc."),
            ("category",        "INT",            "AE category integer (2=Comm, 3=Meter, 4=Rule, 5=Perf)"),
            ("event_type_code", "INT",            "Numeric prefix of eventCode, e.g. 551"),
            ("event_type_name", "TEXT",           "Text portion of eventCode, e.g. Power Meter Check 2.0"),
            ("name",            "TEXT",           "Full alert name / title"),
            ("description",     "TEXT",           "Detail message from alert engine"),
            ("severity",        "INT",            "0 = resolved/low, 5 = active/high"),
            ("impact",          "FLOAT",          "Estimated production impact (kW)"),
            ("impact_code",     "INT",            "AE impact code integer"),
            ("capacity",        "INT",            "Device capacity kW at time of alert"),
            ("is_acknowledged", "BOOLEAN",        "Has alert been acknowledged"),
            ("is_resolved",     "BOOLEAN",        "Has alert been resolved"),
            ("acknowledged_by", "TEXT",           "Name of acknowledging user"),
            ("resolved_by",     "TEXT",           "Name of resolving user or 'Unable to verify name'"),
            ("alert_start",     "TIMESTAMP",      "UTC time alert condition began"),
            ("alert_end",       "TIMESTAMP",      "UTC time alert condition ended (NULL = still open)"),
            ("resolved_time",   "TIMESTAMP",      "UTC time alert marked resolved"),
            ("trigger_time",    "TIMESTAMP",      "UTC time alert was triggered in system"),
            ("last_changed",    "TIMESTAMP",      "UTC time any field on this alert last changed"),
            ("timezone",        "TEXT",           "Local IANA timezone of the site"),
        ],
    },
    "ae_production_daily": {
        "description": "Daily energy production per site — TimescaleDB hypertable on ts",
        "source": "/api/view/production or /api/view/energy POST (to be confirmed)",
        "columns": [
            ("ts",              "TIMESTAMP PK",  "Start of the UTC day (hypertable partition key)"),
            ("site_key",        "TEXT PK",       "Site"),
            ("energy_kwh",      "FLOAT",         "kWh generated during the day"),
            ("peak_kw",         "FLOAT",         "Peak instantaneous power in kW"),
            ("irradiance_avg",  "FLOAT",         "Average POA irradiance W/m² (if weather station)"),
            ("temp_avg_c",      "FLOAT",         "Average module temperature °C"),
            ("performance_index","FLOAT",        "Performance index 0–1"),
            ("expected_kwh",    "FLOAT",         "Expected production from model"),
            ("loss_kwh",        "FLOAT",         "Estimated production loss"),
            ("created_at",      "TIMESTAMP",     "Row insert timestamp"),
        ],
    },
    "ae_production_hourly": {
        "description": "Hourly energy production per device — TimescaleDB hypertable",
        "source": "/api/view/production or /api/view/energydata POST with interval=hour",
        "columns": [
            ("ts",              "TIMESTAMP PK", "Start of UTC hour"),
            ("site_key",        "TEXT PK",      "Site"),
            ("hardware_key",    "TEXT PK",      "Device (inverter or meter)"),
            ("energy_kwh",      "FLOAT",        "kWh during the hour"),
            ("power_kw_avg",    "FLOAT",        "Average kW during the hour"),
            ("power_kw_max",    "FLOAT",        "Max instantaneous kW"),
            ("created_at",      "TIMESTAMP",    "Row insert timestamp"),
        ],
    },
    "ae_work_orders": {
        "description": "Service requests / work orders linked to alerts",
        "source": "/api/view/workorders or /api/view/servicerequests (to be confirmed)",
        "columns": [
            ("work_order_key",  "TEXT PK",    "AlsoEnergy work order key"),
            ("work_order_id",   "TEXT",       "Numeric ID string"),
            ("alert_id",        "TEXT FK",    "Related alert"),
            ("site_key",        "TEXT FK",    "Site"),
            ("title",           "TEXT",       "Work order title / description"),
            ("status",          "TEXT",       "Open / In Progress / Closed"),
            ("assigned_to",     "TEXT",       "Assignee name"),
            ("created_at",      "TIMESTAMP",  "Work order creation time"),
            ("closed_at",       "TIMESTAMP",  "Work order close time"),
        ],
    },
    "ae_notes": {
        "description": "Notes / comments attached to alerts",
        "source": "/api/view/notes (to be confirmed)",
        "columns": [
            ("note_key",    "TEXT PK",    "AlsoEnergy note key"),
            ("alert_id",    "TEXT FK",    "Related alert"),
            ("site_key",    "TEXT FK",    "Site"),
            ("author",      "TEXT",       "Note author"),
            ("body",        "TEXT",       "Note text"),
            ("created_at",  "TIMESTAMP",  "Note timestamp"),
        ],
    },
}

CATEGORY_LEGEND = [
    (2, "Communication", "Device not reporting"),
    (3, "Meter",         "Meter / power quality issue"),
    (4, "Rule Tool",     "Automated rule-based alert"),
    (5, "Performance",   "Performance index below threshold"),
]

ASSET_CODE_LEGEND = [
    ("INV", "Inverter"),
    ("MTR", "Meter (production)"),
    ("MET", "Weather station / met station"),
    ("UNK", "Unknown / site-level"),
]


def write_schema_xlsx(path, probe_results=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("pip install openpyxl")

    DARK_BLUE  = "1F4E79"
    MID_BLUE   = "2E75B6"
    LIGHT_BLUE = "D6E4F0"
    HEADER_FG  = "FFFFFF"
    ALT_ROW    = "EBF3FB"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    def hdr(cell, text, bg=DARK_BLUE, bold=True, size=11):
        cell.value = text
        cell.font = Font(bold=bold, color=HEADER_FG, size=size)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def val(cell, text, alt=False, wrap=False):
        cell.value = text
        if alt:
            cell.fill = PatternFill("solid", fgColor=ALT_ROW)
        cell.alignment = Alignment(wrap_text=wrap, vertical="top")

    def autosize(ws, min_w=12, max_w=60):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)

    # ── Sheet 1: Overview / ERD description ──────────────────────────
    ws = wb.create_sheet("Overview")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 40

    hdr(ws["A1"], "Table", MID_BLUE)
    hdr(ws["B1"], "Description", MID_BLUE)
    hdr(ws["C1"], "Source / Notes", MID_BLUE)
    ws.row_dimensions[1].height = 22

    for i, (tname, tinfo) in enumerate(SCHEMA.items(), start=2):
        alt = i % 2 == 0
        val(ws.cell(i, 1), tname, alt)
        val(ws.cell(i, 2), tinfo["description"], alt, wrap=True)
        val(ws.cell(i, 3), tinfo["source"], alt, wrap=True)
        ws.row_dimensions[i].height = 28

    ws.freeze_panes = "A2"

    # ── Sheet 2: Column definitions per table ─────────────────────────
    ws = wb.create_sheet("Column Definitions")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 55

    hdr(ws["A1"], "Table")
    hdr(ws["B1"], "Column")
    hdr(ws["C1"], "Data Type")
    hdr(ws["D1"], "Description")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    row = 2
    for tname, tinfo in SCHEMA.items():
        for col_name, col_type, col_desc in tinfo["columns"]:
            alt = row % 2 == 0
            val(ws.cell(row, 1), tname, alt)
            val(ws.cell(row, 2), col_name, alt)
            val(ws.cell(row, 3), col_type, alt)
            val(ws.cell(row, 4), col_desc, alt, wrap=True)
            ws.row_dimensions[row].height = 20
            row += 1

    # ── Sheet 3: DDL (CREATE TABLE statements) ─────────────────────────
    ws = wb.create_sheet("DDL")
    ws.column_dimensions["A"].width = 100

    hdr(ws["A1"], "Postgres / TimescaleDB DDL  (copy-paste into psql)", MID_BLUE)
    ws.row_dimensions[1].height = 22

    ddl_lines = []
    for tname, tinfo in SCHEMA.items():
        ddl_lines.append(f"-- {tinfo['description']}")
        ddl_lines.append(f"CREATE TABLE IF NOT EXISTS {tname} (")
        col_strs = []
        for col_name, col_type, _ in tinfo["columns"]:
            pg_type = (col_type
                       .replace(" PK", " PRIMARY KEY")
                       .replace(" FK", "")
                       .replace("BIGSERIAL", "BIGSERIAL")
                       .replace("UNIQUE", ""))
            constraint = ""
            if "PK" in col_type:
                constraint = " PRIMARY KEY"
            if "UNIQUE" in col_type:
                constraint += " UNIQUE"
            base = col_type.split()[0]
            col_strs.append(f"    {col_name:<24} {base}{constraint}")
        ddl_lines.append(",\n".join(col_strs))
        ddl_lines.append(");\n")
        # hypertable hint
        if tname in ("ae_production_daily", "ae_production_hourly"):
            ddl_lines.append(f"-- SELECT create_hypertable('{tname}', 'ts');")
            ddl_lines.append("")

    for i, line in enumerate(ddl_lines, start=2):
        ws.cell(i, 1).value = line
        ws.cell(i, 1).font = Font(name="Courier New", size=9)
        ws.row_dimensions[i].height = 14

    # ── Sheet 4: Reference codes ────────────────────────────────────────
    ws = wb.create_sheet("Reference Codes")

    hdr(ws["A1"], "Category (ae_alerts.category)", MID_BLUE)
    ws.merge_cells("A1:C1")
    hdr(ws["A2"], "Code", DARK_BLUE)
    hdr(ws["B2"], "Name", DARK_BLUE)
    hdr(ws["C2"], "Meaning", DARK_BLUE)
    for i, (code, name, meaning) in enumerate(CATEGORY_LEGEND, start=3):
        alt = i % 2 == 0
        val(ws.cell(i, 1), code, alt)
        val(ws.cell(i, 2), name, alt)
        val(ws.cell(i, 3), meaning, alt)

    ws["A8"].value = ""
    hdr(ws["A9"], "Asset Code (ae_alerts.asset_code / ae_hardware.asset_code)", MID_BLUE)
    ws.merge_cells("A9:C9")
    hdr(ws["A10"], "Code", DARK_BLUE)
    hdr(ws["B10"], "Device Type", DARK_BLUE)
    for i, (code, desc) in enumerate(ASSET_CODE_LEGEND, start=11):
        alt = i % 2 == 0
        val(ws.cell(i, 1), code, alt)
        val(ws.cell(i, 2), desc, alt)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 42

    # ── Sheet 5: Sample alert data (from probe if available) ────────────
    if probe_results and probe_results.get("alerts"):
        raw = probe_results["alerts"]
        records = raw.get("list", []) if isinstance(raw, dict) else (raw if isinstance(raw, list) else [])
        if records:
            ws = wb.create_sheet("Sample Alerts")
            keys = list(records[0].keys())
            for ci, k in enumerate(keys, start=1):
                hdr(ws.cell(1, ci), k, MID_BLUE, size=9)
            for ri, rec in enumerate(records[:50], start=2):
                for ci, k in enumerate(keys, start=1):
                    v = rec.get(k)
                    ws.cell(ri, ci).value = str(v) if v is not None else ""
                    ws.cell(ri, ci).font = Font(size=9)
            ws.freeze_panes = "A2"
            for col in ws.columns:
                ml = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 40)

    # ── Sheet 6: Discovered endpoints ──────────────────────────────────
    if probe_results:
        ws = wb.create_sheet("API Probe Results")
        hdr(ws["A1"], "Endpoint / Label", MID_BLUE)
        hdr(ws["B1"], "Result", MID_BLUE)
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 80
        row = 2
        for label, data in probe_results.items():
            ws.cell(row, 1).value = label
            if data is None:
                summary = "No response / error"
            elif isinstance(data, dict):
                summary = f"Dict keys: {list(data.keys())}  |  list sizes: " + \
                          ", ".join(f"{k}={len(v)}" for k, v in data.items() if isinstance(v, list))
            elif isinstance(data, list):
                summary = f"List of {len(data)} items"
            else:
                summary = str(data)[:200]
            ws.cell(row, 2).value = summary
            ws.cell(row, 2).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 28
            row += 1

    wb.save(path)
    print(f"\nSchema spreadsheet saved ->{path}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")


# -----------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--probe",  action="store_true", help="probe endpoints only, skip spreadsheet")
    ap.add_argument("--schema", action="store_true", help="write spreadsheet only, skip probing")
    args = ap.parse_args()

    probe_results = None

    if not args.schema:
        from ae_auth import get_session
        session = get_session()
        probe_results = run_probes(session)

    if not args.probe:
        write_schema_xlsx(OUTPUT_FILE, probe_results)

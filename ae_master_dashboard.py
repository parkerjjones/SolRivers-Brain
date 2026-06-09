#!/usr/bin/env python3
"""
Master portfolio health dashboard.

Joins all collected data sources into one definitive spreadsheet:
  - ae_alerts.xlsx
  - ae_ruleresults.xlsx
  - ae_summaries_deep.xlsx
  - ae_hardware.xlsx
  - ae_sites.xlsx  (optional)

Writes ae_master.xlsx with:
  - Portfolio Overview    : live power + urgency + alert count per site
  - Risk Matrix           : sites ranked by composite risk score
  - Alert Summary         : alerts grouped by site and event type
  - Hardware Failure Map  : which device types fail most per site
  - Rule vs Alert         : sites where rule failures correlate with alerts
  - Capacity Breakdown    : device count and type by site
  - Raw Data              : all sites flattened

USAGE
-----
    python ae_master_dashboard.py
"""

import os
import sys
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

DARK  = "1F4E79"
MID   = "2E75B6"
WHITE = "FFFFFF"
ALT   = "EBF3FB"
GREEN = "C6EFCE"
RED   = "FFC7CE"
YELL  = "FFEB9C"
ORANGE= "FFCC99"


def hdr(cell, text, bg=DARK, size=10):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def val(cell, v, row_idx=0, color=None):
    cell.value = v if not (isinstance(v, float) and pd.isna(v)) else ""
    bg = color or (ALT if row_idx % 2 == 0 else None)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="top")


def autosize(ws, max_w=55):
    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, max_w)


def write_df(ws, df, start_row=2, color_col=None, color_map=None):
    for ci, col in enumerate(df.columns, 1):
        hdr(ws.cell(start_row - 1, ci), col, MID)
    ws.freeze_panes = ws.cell(start_row, 1)
    for ri, row in enumerate(df.itertuples(index=False), start=start_row):
        alt = ri % 2
        bg = ALT if alt == 0 else None
        if color_col and color_map:
            cv = getattr(row, color_col.replace(" ", "_"), None)
            bg = color_map.get(str(cv), bg)
        for ci, v in enumerate(row, start=1):
            c = ws.cell(ri, ci)
            c.value = v if not (isinstance(v, float) and pd.isna(v)) else ""
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(vertical="top", wrap_text=(ci == len(df.columns)))
    autosize(ws)


# -----------------------------------------------------------------------
# LOAD DATA
# -----------------------------------------------------------------------

def load(path, sheet=0):
    if not os.path.exists(path):
        print(f"  MISSING: {path}", file=sys.stderr)
        return pd.DataFrame()
    return pd.read_excel(path, sheet_name=sheet)


print("Loading data sources...")
alerts_df   = load("ae_alerts.xlsx")
rules_df    = load("ae_ruleresults.xlsx", "Site Pass-Fail")
summary_df  = load("ae_summaries_deep.xlsx", "Site Analysis")
hw_df       = load("ae_hardware.xlsx", "Hardware Inventory")
hw_cols_df  = load("ae_hardware.xlsx", "Archive Columns")
sites_df    = load("ae_sites.xlsx", "Sites Overview") if os.path.exists("ae_sites.xlsx") else pd.DataFrame()
live_df     = load("ae_sites.xlsx", "Live Power") if os.path.exists("ae_sites.xlsx") else pd.DataFrame()

for name, df in [("alerts", alerts_df), ("rules", rules_df),
                  ("summaries", summary_df), ("hardware", hw_df)]:
    print(f"  {name}: {len(df)} rows")

# -----------------------------------------------------------------------
# SITE MASTER JOIN
# -----------------------------------------------------------------------

# Base: unique sites from rules or summaries
if not rules_df.empty:
    site_master = rules_df[["Site Key", "Site Name"]].drop_duplicates().copy()
    site_master.columns = ["site_key", "site_name"]
elif not summary_df.empty:
    site_master = summary_df[["site_key", "site_name"]].drop_duplicates().copy()
else:
    site_master = pd.DataFrame(columns=["site_key", "site_name"])

# Alert counts per site
if not alerts_df.empty:
    acols = {c.lower().replace(" ", "_"): c for c in alerts_df.columns}
    sk_col = next((v for k, v in acols.items() if "site_name" in k), None)
    res_col = next((v for k, v in acols.items() if "is_resolved" in k), None)
    if sk_col:
        alert_counts = alerts_df.groupby(sk_col).agg(
            total_alerts=(sk_col, "count"),
            open_alerts=(res_col, lambda x: (~x.astype(bool)).sum()) if res_col else (sk_col, "count"),
        ).reset_index().rename(columns={sk_col: "site_name"})
        site_master = site_master.merge(alert_counts, on="site_name", how="left")
    else:
        site_master["total_alerts"] = 0
        site_master["open_alerts"] = 0
else:
    site_master["total_alerts"] = 0
    site_master["open_alerts"] = 0

# Rule results
if not rules_df.empty:
    rcols = {c.lower().replace(" ", "_").replace("-", "_"): c for c in rules_df.columns}
    fail_col = next((v for k, v in rcols.items() if "fail" in k and "comm" not in k and "perf" not in k), None)
    pass_col = next((v for k, v in rcols.items() if k == "passed" or "pass" in k and "comm" not in k), None)
    rr = rules_df[["Site Key", "Site Name"]].copy()
    rr.columns = ["site_key", "site_name"]
    if fail_col:
        rr["rule_failures"] = rules_df[fail_col].fillna(0)
    else:
        rr["rule_failures"] = 0
    site_master = site_master.merge(
        rr[["site_name", "rule_failures"]].drop_duplicates("site_name"),
        on="site_name", how="left"
    )

# AI urgency
if not summary_df.empty:
    urg = summary_df[["site_key", "site_name", "urgency_score", "urgency_label",
                       "comm_status", "prod_status", "has_active_alert"]].copy()
    site_master = site_master.merge(
        urg.drop_duplicates("site_key"),
        on=["site_key", "site_name"], how="left"
    )

# Hardware counts
if not hw_df.empty:
    hcols = {c.lower().replace(" ", "_"): c for c in hw_df.columns}
    sk_col = next((v for k, v in hcols.items() if "site_key" in k or k == "site_key"), None)
    if sk_col:
        hw_counts = hw_df.groupby(sk_col).size().reset_index(name="device_count")
        hw_counts.columns = ["site_key", "device_count"]
        site_master = site_master.merge(hw_counts, on="site_key", how="left")

# Live power
if not live_df.empty:
    lcols = {c.lower().replace(" ", "_"): c for c in live_df.columns}
    sk_col = next((v for k, v in lcols.items() if "site_key" in k), None)
    pwr_col = next((v for k, v in lcols.items() if "current" in k and "kw" in k), None)
    if sk_col and pwr_col:
        lv = live_df[[sk_col, pwr_col]].copy()
        lv.columns = ["site_key", "current_kw"]
        site_master = site_master.merge(lv, on="site_key", how="left")

# Composite risk score (0-100)
def risk_score(row):
    score = 0
    score += min(row.get("open_alerts", 0) * 5, 35)
    score += min(row.get("rule_failures", 0) * 3, 25)
    score += {"High": 25, "Medium": 15, "Low": 5, "Clean": 0}.get(
        str(row.get("urgency_label", "")), 10)
    score += {"Partial / Issues": 15, "Unknown": 5, "All Communicating": 0}.get(
        str(row.get("comm_status", "")), 5)
    score += {"Below Expected": 10, "Unknown": 5, "On Track": 0}.get(
        str(row.get("prod_status", "")), 5)
    return min(score, 100)

site_master["risk_score"] = site_master.apply(risk_score, axis=1)
site_master["risk_tier"] = pd.cut(
    site_master["risk_score"],
    bins=[-1, 20, 40, 60, 101],
    labels=["Low", "Medium", "High", "Critical"]
)

site_master = site_master.sort_values("risk_score", ascending=False)
site_master["risk_tier"] = site_master["risk_tier"].astype(str)
site_master = site_master.fillna("")

# -----------------------------------------------------------------------
# WRITE XLSX
# -----------------------------------------------------------------------

print("Building ae_master.xlsx...")
wb = openpyxl.Workbook()
wb.remove(wb.active)
risk_colors = {"Critical": RED, "High": ORANGE, "Medium": YELL, "Low": GREEN}

# Sheet 1 — Portfolio Overview
ws1 = wb.create_sheet("Portfolio Overview")
overview_cols = ["site_key", "site_name", "risk_score", "risk_tier",
                 "open_alerts", "total_alerts", "rule_failures",
                 "urgency_label", "comm_status", "prod_status",
                 "has_active_alert", "device_count"]
ov = site_master[[c for c in overview_cols if c in site_master.columns]]
write_df(ws1, ov, color_col="risk_tier", color_map=risk_colors)

# Sheet 2 — Risk Matrix
ws2 = wb.create_sheet("Risk Matrix")
risk_cols = ["site_name", "risk_score", "risk_tier", "open_alerts",
             "rule_failures", "urgency_label", "comm_status", "prod_status"]
rm = site_master[[c for c in risk_cols if c in site_master.columns]]
write_df(ws2, rm, color_col="risk_tier", color_map=risk_colors)

# Sheet 3 — Alert Summary by Site & Type
if not alerts_df.empty:
    ws3 = wb.create_sheet("Alerts by Site")
    acols2 = {c.lower().replace(" ", "_"): c for c in alerts_df.columns}
    sn = next((v for k, v in acols2.items() if "site_name" in k), None)
    et = next((v for k, v in acols2.items() if "event_type_name" in k), None)
    sev = next((v for k, v in acols2.items() if "severity" in k), None)
    res = next((v for k, v in acols2.items() if "is_resolved" in k), None)
    if sn and et:
        agg = alerts_df.groupby([sn, et]).agg(
            count=(sn, "count"),
            open_count=(res, lambda x: (~x.astype(bool)).sum()) if res else (sn, "count"),
            avg_severity=(sev, "mean") if sev else (sn, "count"),
        ).reset_index().sort_values("count", ascending=False)
        write_df(ws3, agg)

# Sheet 4 — Hardware Failure Map
if not alerts_df.empty and not hw_df.empty:
    ws4 = wb.create_sheet("Hardware Failure Map")
    acols3 = {c.lower().replace(" ", "_"): c for c in alerts_df.columns}
    sn_a = next((v for k, v in acols3.items() if "site_name" in k), None)
    hn_a = next((v for k, v in acols3.items() if "hardware_name" in k), None)
    ac_a = next((v for k, v in acols3.items() if "asset_code" in k), None)
    if sn_a and hn_a:
        hw_fails = alerts_df[alerts_df[hn_a].notna() & (alerts_df[hn_a] != "")].groupby(
            [sn_a, hn_a, ac_a] if ac_a else [sn_a, hn_a]
        ).size().reset_index(name="alert_count").sort_values("alert_count", ascending=False)
        write_df(ws4, hw_fails.head(100))

# Sheet 5 — Rule Failures vs Alert Count
if not rules_df.empty and not alerts_df.empty:
    ws5 = wb.create_sheet("Rules vs Alerts")
    rule_detail = load("ae_ruleresults.xlsx", "Rule Details")
    if not rule_detail.empty:
        rcols2 = {c.lower().replace(" ", "_"): c for c in rule_detail.columns}
        sn_r = next((v for k, v in rcols2.items() if k == "site_name"), None)
        id_r = next((v for k, v in rcols2.items() if k == "rule_name" or "id" in k), None)
        pass_r = next((v for k, v in rcols2.items() if "passed" in k), None)
        if sn_r and id_r:
            fails = rule_detail[rule_detail[pass_r] == "FAIL"] if pass_r else rule_detail
            fail_agg = fails.groupby([sn_r, id_r]).size().reset_index(name="fail_count")
            # join with alert counts
            if sn_a:
                ac = alerts_df.groupby(sn_a).size().reset_index(name="total_alerts")
                ac.columns = [sn_r, "total_alerts"]
                fail_agg = fail_agg.merge(ac, on=sn_r, how="left")
            write_df(ws5, fail_agg.sort_values("fail_count", ascending=False).head(100))

# Sheet 6 — Device Type Breakdown
if not hw_df.empty:
    ws6 = wb.create_sheet("Device Type Breakdown")
    hcols2 = {c.lower().replace(" ", "_"): c for c in hw_df.columns}
    sk2 = next((v for k, v in hcols2.items() if k == "site_key"), None)
    sn2 = next((v for k, v in hcols2.items() if k == "site_name"), None)
    dt2 = next((v for k, v in hcols2.items() if "device_type" in k), None)
    if sk2 and dt2:
        dtype_pivot = hw_df.groupby([sn2 or sk2, dt2]).size().unstack(fill_value=0).reset_index()
        dtype_pivot.columns.name = None
        write_df(ws6, dtype_pivot)

# Sheet 7 — Archive Columns (what telemetry is available per device type)
if not hw_cols_df.empty:
    ws7 = wb.create_sheet("Available Telemetry")
    cc = {c.lower().replace(" ","_"): c for c in hw_cols_df.columns}
    dt_c = next((v for k,v in cc.items() if "device_type" in k), None)
    col_c = next((v for k,v in cc.items() if "column" in k), None)
    if dt_c and col_c:
        telem = hw_cols_df.groupby([dt_c, col_c]).size().reset_index(name="site_count")
        telem = telem.sort_values(["device_type" if dt_c=="device_type" else dt_c, "site_count"], ascending=[True, False])
        write_df(ws7, telem)

# Sheet 8 — Full Site Master
ws8 = wb.create_sheet("Full Site Master")
write_df(ws8, site_master, color_col="risk_tier", color_map=risk_colors)

wb.save("ae_master.xlsx")
print(f"Saved -> ae_master.xlsx")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")

print("\n--- Portfolio Health Summary ---")
if "risk_tier" in site_master.columns:
    print(site_master["risk_tier"].value_counts().to_string())
def safe_sum(df, col):
    if col not in df.columns: return "N/A"
    return pd.to_numeric(df[col], errors="coerce").sum()

print(f"\nTotal open alerts:   {safe_sum(site_master, 'open_alerts'):.0f}")
print(f"Total rule failures: {safe_sum(site_master, 'rule_failures'):.0f}")
if "risk_score" in site_master.columns:
    top5 = site_master[["site_name","risk_score","risk_tier"]].head(5)
    print("\nTop 5 highest-risk sites:")
    print(top5.to_string(index=False))

#!/usr/bin/env python3
"""
Automated Operational Alert Dashboard for the SolRiver portfolio.

Pulls the most recent alert history (default 7 days), filters to the alerts
that matter (inverter faults, production stops, grid events, tracker/TCU
faults), and renders dashboards/alerts.html with:

  - KPI strip            : alert volume, critical count, unresolved, sites hit
  - Critical alert table : faults / production stops, severity-ranked
  - Inverter heatmaps    : per affected site — inverter x day grid colored by
                           fault-hours (red) vs comm-loss-hours (orange)
  - Tracker / TCU section: per-site alarm-per-day charts + parsed TCU faults
  - Grid & meter events  : recloser / islanding / power-meter checks

Chart selection logic (alert category -> visualization + measurement unit):
  INVERTER_FAULT  -> inverter heatmap          (context: Power kW / AC Current A)
  INVERTER_COMM   -> inverter heatmap (orange) (context: comm hours)
  TRACKER         -> alarms-per-day bar chart  (context: tracker count / angle deg)
  GRID            -> event timeline            (context: AC Voltage V)
  METER           -> event table               (context: Power kW per phase)
  PERFORMANCE     -> event table               (context: Performance Index %)

Measurement -> unit map mirrors the PowerTrack Chart Builder:
  AC Current=A, AC Voltage=V, DC Current=A, DC Voltage=V, Power=kW,
  Energy=kWh, Irradiance=W/m^2, Temperature=degC, Availability=%,
  Performance Index=%, Energy Ratio=%, Capacity Factor=%, Phase Angle=deg

USAGE
-----
    python ae_alert_dashboard.py                 # last 7 days, live fetch
    python ae_alert_dashboard.py --days 14       # 14-day look-back
    python ae_alert_dashboard.py --offline       # use cached ae_alerts.xlsx
    python ae_alert_dashboard.py --out dashboards/alerts.html
"""

import argparse
import json
import re
import sys
import time
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

API_BASE       = "https://apps.alsoenergy.com/api"
PORTFOLIO_KEYS = ["C12941", "C47197"]
TZ_OFFSET_MIN  = 360
SLEEP          = 0.5
HERE           = Path(__file__).parent
CHARTJS_FILE   = HERE / "assets" / "chart.umd.min.js"
HW_XLSX        = HERE / "ae_hardware.xlsx"
ALERTS_XLSX    = HERE / "ae_alerts.xlsx"

# ── Measurement -> unit map (PowerTrack Chart Builder) ────────────────────
MEASUREMENT_UNITS = {
    "AC Current": "A", "AC Voltage": "V", "DC Current": "A", "DC Voltage": "V",
    "Power": "kW", "Energy": "kWh", "Energy (cumulative)": "kWh",
    "Estimated Power": "kW", "Estimated Energy": "kWh",
    "Expected Power": "kW", "Expected Energy": "kWh",
    "Irradiance": "W/m²", "Temperature": "°C", "Weather": "code",
    "Availability": "%", "Capacity Factor": "%", "Energy Ratio": "%",
    "Performance Index": "%", "Phase Angle": "°", "Faults": "count",
    "Control Setpoints": "kW", "Power On Time": "h", "String Current": "A",
}

SEVERITY_LABEL = {0: "Normal", 1: "Info", 2: "Warning", 3: "Critical",
                  4: "Emergency", 5: "High"}

# ── Alert classification ──────────────────────────────────────────────────
# (regex on event_type_name + description, asset_code) -> category
# First match wins. Category drives chart type, measurement and importance.

CATEGORY_META = {
    # category        weight  label                      measurement            chart
    "INVERTER_FAULT": (100, "Inverter Fault / Stop",     "Power",               "heatmap"),
    "GRID":           (95,  "Grid / Protection Event",   "AC Voltage",          "timeline"),
    "TRACKER_FAULT":  (80,  "Tracker TCU Fault",         "Phase Angle",         "tracker"),
    "METER":          (70,  "Production Meter Issue",    "Power",               "table"),
    "INVERTER_COMM":  (60,  "Inverter Comm Loss",        "Availability",        "heatmap"),
    "TRACKER_COMM":   (50,  "Tracker Comm Loss",         "Availability",        "tracker"),
    "PERFORMANCE":    (45,  "Performance / Irradiance",  "Performance Index",   "table"),
    "EQUIPMENT":      (40,  "Balance-of-Plant Equipment","Temperature",         "table"),
    "COMMS_LOW":      (10,  "Site Comms / Gateway",      "Availability",        "none"),
    "OTHER":          (5,   "Other",                     "Availability",        "none"),
}

def classify(event_type, description, asset):
    et   = (event_type or "").lower()
    desc = (description or "").lower()
    a    = (asset or "").upper()

    # Grid / protection first — these stop whole-site production
    if any(k in et or k in desc for k in
           ("recloser", "islanding", "grid power outage", "breaker status")):
        return "GRID"

    # Inverter faults & stops
    if a == "INV" or "inverter" in et or "inv fault" in et or "string inv" in et:
        if any(k in et for k in ("fault", "stopped", "inverter alerts")) \
           or any(k in desc for k in ("fault", "stop", "islanding", "shut")):
            return "INVERTER_FAULT"
        if any(k in et for k in ("communication", "heartbeat")):
            return "INVERTER_COMM"
        if "irradiance" in et or "performance" in et:
            return "PERFORMANCE"
        return "INVERTER_FAULT"   # unknown inverter alert: treat as important

    # Trackers
    if a == "TKR" or "tracker" in et or "tcu" in desc:
        if any(k in et for k in ("communication", "heartbeat")) and "tracker alert" not in et:
            return "TRACKER_COMM"
        if "no communication" in desc and "fault" not in desc:
            return "TRACKER_COMM"
        return "TRACKER_FAULT"

    # Meters
    if a == "MTR" or "power meter" in et or "meter" in et:
        return "METER"

    # Performance / rule tool
    if any(k in et for k in ("performance index", "irradiance check", "rule tool",
                             "data comparison")):
        return "PERFORMANCE"

    # Balance of plant
    if any(k in et for k in ("ups", "transformer", "sel ", "recloser", "controller health")):
        return "EQUIPMENT"

    # Low-value comms noise
    if any(k in et for k in ("communication", "heartbeat", "gateway")):
        return "COMMS_LOW"

    return "OTHER"


def importance(alert):
    """Score an alert: category weight + severity + duration + unresolved."""
    cat_w = CATEGORY_META[alert["category"]][0]
    sev   = alert.get("severity") or 0
    score = cat_w + sev * 4
    if not alert.get("is_resolved"):
        score += 20
    score += min(alert.get("duration_h") or 0, 48) * 0.5
    if alert.get("capacity"):
        score += min(float(alert["capacity"]) / 100.0, 10)
    return round(score, 1)


# ── Data fetching ─────────────────────────────────────────────────────────

def get_session():
    from ae_auth import get_session as _gs
    return _gs()


def fetch_alerts_live(session, d_from, d_to):
    rows = []
    for key in PORTFOLIO_KEYS:
        payload = {"key": key, "offset": TZ_OFFSET_MIN,
                   "from": d_from.isoformat(), "to": d_to.isoformat()}
        r = session.post(f"{API_BASE}/view/alerthistory", json=payload, timeout=60)
        if r.status_code == 204:
            continue
        r.raise_for_status()
        data = r.json()
        recs = data if isinstance(data, list) else next(
            (data[k] for k in ("list", "data", "alerts", "items", "rows")
             if isinstance(data.get(k), list)), [])
        for rec in recs:
            rows.append(normalize_live(rec))
        time.sleep(SLEEP)
    return rows


def parse_ts(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y %I:%M:%S %p"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def split_event_code(s):
    if not s:
        return None
    m = re.match(r"^\s*(\d+)\s*-\s*(.+)$", str(s))
    return m.group(2).strip() if m else str(s).strip()


def normalize_live(rec):
    return {
        "site_key":      rec.get("siteKey"),
        "site_name":     rec.get("siteName"),
        "hardware_key":  rec.get("hardwareKey") or None,
        "hardware_name": rec.get("hardwareName") or None,
        "asset_code":    rec.get("assetCode") or None,
        "event_type":    split_event_code(rec.get("eventCode")),
        "description":   rec.get("description") or "",
        "severity":      rec.get("severity") or 0,
        "is_resolved":   bool(rec.get("isResolved")),
        "capacity":      rec.get("capacity"),
        "start":         parse_ts(rec.get("start")),
        "end":           parse_ts(rec.get("end")),
    }


def fetch_alerts_offline(d_from, d_to):
    import openpyxl
    wb = openpyxl.load_workbook(ALERTS_XLSX, read_only=True)
    ws = wb.active
    rows_iter = ws.values
    hdr = next(rows_iter)
    idx = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows_iter:
        start = r[idx["alert_start"]]
        if isinstance(start, str):
            start = parse_ts(start)
        if not start or not (d_from <= start.date() <= d_to):
            continue
        out.append({
            "site_key":      None,
            "site_name":     r[idx["site_name"]],
            "hardware_key":  None,
            "hardware_name": r[idx["hardware_name"]],
            "asset_code":    r[idx["asset_code"]],
            "event_type":    r[idx["event_type_name"]],
            "description":   r[idx["description"]] or "",
            "severity":      r[idx["severity"]] or 0,
            "is_resolved":   bool(r[idx["is_resolved"]]),
            "capacity":      r[idx["capacity"]],
            "start":         start,
            "end":           r[idx["alert_end"]] if isinstance(r[idx["alert_end"]], datetime)
                             else parse_ts(r[idx["alert_end"]]),
        })
    return out


INV_CAPACITY_RE = [
    (re.compile(r"SG(\d+)HV",  re.I), lambda m: int(m.group(1))),
    (re.compile(r"SG(\d+)",    re.I), lambda m: int(m.group(1))),
    (re.compile(r"SCH(\d+)",   re.I), lambda m: int(m.group(1))),
    (re.compile(r"SCA(\d+)",   re.I), lambda m: int(m.group(1))),
    (re.compile(r"PVI\s*(\d+)",re.I), lambda m: int(m.group(1))),
    (re.compile(r"XGI\s*1500", re.I), lambda m: 200),
    (re.compile(r"XGI\s*1000", re.I), lambda m: 166),
    (re.compile(r"(\d+)\s*kTL", re.I), lambda m: int(m.group(1))),
    (re.compile(r"(\d+)\s*kW",  re.I), lambda m: int(m.group(1))),
]

def guess_inv_capacity(name):
    for rx, fn in INV_CAPACITY_RE:
        m = rx.search(name or "")
        if m:
            return fn(m)
    return None


def load_hardware():
    """site_name -> {'inverters': [names...], 'trackers': [names...],
                     'inv_keys': {name: hw_key}, 'inv_cap': {name: kW},
                     'weather_keys': [hw_keys...], 'site_key': str}"""
    import openpyxl
    hw = defaultdict(lambda: {"inverters": [], "trackers": [],
                               "inv_keys": {}, "inv_cap": {},
                               "weather_keys": [], "site_key": None})
    if not HW_XLSX.exists():
        return hw
    wb = openpyxl.load_workbook(HW_XLSX, read_only=True)
    ws = wb["Hardware Inventory"]
    rows_iter = ws.values
    hdr = next(rows_iter)
    idx = {h: i for i, h in enumerate(hdr)}
    for r in rows_iter:
        fc = r[idx["Function Code"]]
        site = r[idx["Site Name"]]
        name = r[idx["HW Name"]]
        hw_key = r[idx["HW Key"]]
        if fc == 1:
            hw[site]["inverters"].append(name)
            hw[site]["inv_keys"][name] = hw_key
            cap = guess_inv_capacity(name)
            if cap:
                hw[site]["inv_cap"][name] = cap
        elif fc == 24:
            hw[site]["trackers"].append(name)
        if fc == 5:
            name_low = (name or "").lower()
            if "poa" in name_low and "offline" not in name_low:
                hw[site]["weather_keys"].insert(0, hw_key)
            elif "offline" not in name_low:
                hw[site]["weather_keys"].append(hw_key)
        if hw[site]["site_key"] is None and "Site Key" in idx:
            hw[site]["site_key"] = r[idx["Site Key"]]
    return hw


SITE_EMAIL_ALIASES = {
    "elk": "Elk Solar", "elk solar": "Elk Solar",
    "whitetail": "Whitetail", "sunflower": "Sunflower Solar",
    "williams": "Williams Solar, LLC", "williams solar": "Williams Solar, LLC",
    "washington": "Washington Solar", "washington solar": "Washington Solar",
    "wallace": "Wallace Solar", "wallace solar": "Wallace Solar",
    "graham": "C & B Graham Energy", "c&b graham": "C & B Graham Energy",
    "c & b graham": "C & B Graham Energy", "c&b graham energy": "C & B Graham Energy",
    "green solar": "Green Solar", "green elementary": "Green Elementary",
    "gallia": "Gallia Academy", "gallia academy": "Gallia Academy",
    "richmond": "Richmond", "shorthorn": "Shorthorn",
    "bulloch": "Bulloch 1A", "bulloch 1a": "Bulloch 1A", "bulloch 1b": "Bulloch 1B",
    "bulloch 1a & 1b": "Bulloch 1A",
    "monroe": "Monroe Landfill", "eagle": "Eagle",
    "warbler": "Warbler", "mclean": "McLean",
    "harding": "Harding Solar", "whitehall": "Whitehall Solar",
    "gray fox": "Gray Fox Solar", "longleaf": "Longleaf Pine Solar, LLC",
    "marble": "Marble Solar", "sheridan": "Sheridan Solar",
    "auburn": "Auburn Solar", "rit": "RIT", "rrh": "RRH 1 & 2",
    "butler maple": "Butler Maple", "maple": "Butler Maple",
    "upson": "Upson",
}


def load_site_emails():
    """Load the AE Reports Site Emails xlsx. Returns {canonical_site: [email_dicts]}"""
    import glob as _glob
    files = sorted(_glob.glob(str(HERE / "AE_Reports_Site_Emails*.xlsx")), reverse=True)
    if not files:
        return {}
    import openpyxl
    wb = openpyxl.load_workbook(files[0], read_only=True)
    ws = wb.active
    rows_iter = ws.values
    hdr = next(rows_iter)
    idx = {h: i for i, h in enumerate(hdr)}
    by_site = defaultdict(list)
    for r in rows_iter:
        raw_sites = str(r[idx.get("Site(s)", 3)] or "")
        date_val = r[idx.get("Date", 0)]
        email = {
            "date": str(date_val)[:10] if date_val else "",
            "from": str(r[idx.get("From / Participants", 1)] or "")[:120],
            "subject": str(r[idx.get("Subject", 2)] or ""),
            "category": str(r[idx.get("Category", 4)] or ""),
            "contents": str(r[idx.get("Contents", 5)] or ""),
            "raw_sites": raw_sites,
        }
        matched = set()
        raw_lower = raw_sites.lower()
        for alias, canonical in SITE_EMAIL_ALIASES.items():
            if alias in raw_lower:
                matched.add(canonical)
        if "bulloch 1a & 1b" in raw_lower or "bulloch 1a and 1b" in raw_lower:
            matched.add("Bulloch 1A")
            matched.add("Bulloch 1B")
        if "duke nc" in raw_lower:
            for s in ["Elk Solar", "Williams Solar, LLC", "Shorthorn",
                       "Sunflower Solar", "Gray Fox Solar"]:
                matched.add(s)
        if "oregon" in raw_lower or "or site" in raw_lower:
            for s in ["Green Solar", "Wallace Solar", "Marble Solar",
                       "Auburn Solar", "Sheridan Solar"]:
                matched.add(s)
        if "ga site" in raw_lower or "ga-" in raw_lower:
            for s in ["Bulloch 1A", "Bulloch 1B", "Richmond", "Upson"]:
                matched.add(s)
        if not matched:
            matched.add("PORTFOLIO")
        for site in matched:
            by_site[site].append(email)
    print(f"  Loaded {sum(len(v) for v in by_site.values())} email refs across {len(by_site)} sites from {Path(files[0]).name}")
    return by_site


SUMMARIES_XLSX = HERE / "ae_ai_summaries.xlsx"

def _fix_encoding(s):
    """Fix UTF-8 bytes stored as individual Latin-1/cp1252 characters."""
    # Reassemble UTF-8 byte sequences from their Latin-1 character codes
    out = []
    i = 0
    while i < len(s):
        c = ord(s[i])
        # Three-byte UTF-8: 0xE0-0xEF + two continuation bytes
        if (0xE0 <= c <= 0xEF and i + 2 < len(s)
                and 0x80 <= ord(s[i+1]) <= 0xBF
                and 0x80 <= ord(s[i+2]) <= 0xBF):
            try:
                out.append(bytes([c, ord(s[i+1]), ord(s[i+2])]).decode("utf-8"))
                i += 3
                continue
            except UnicodeDecodeError:
                pass
        # Two-byte UTF-8: 0xC0-0xDF + one continuation byte
        if (0xC0 <= c <= 0xDF and i + 1 < len(s)
                and 0x80 <= ord(s[i+1]) <= 0xBF):
            try:
                out.append(bytes([c, ord(s[i+1])]).decode("utf-8"))
                i += 2
                continue
            except UnicodeDecodeError:
                pass
        out.append(s[i])
        i += 1
    return "".join(out)


def _fix_encoding_old(s):
    """Fallback replacement-based fix."""
    replacements = [
        ("Â°", "°"),   # Â° -> °
        ("Â·", "·"),   # Â· -> ·
        ("Â±", "±"),   # Â± -> ±
        ("Ã", "×"),   # Ã— -> ×
        ("â", "–"),  # â€" -> –
        ("â", "—"),  # â€" -> —
        ("â", "’"),  # â€™ -> '
        ("â", "“"),  # â€œ -> "
        ("â", "”"),  # â€ -> "
        ("â¦", "…"),  # â€¦ -> …
        ("â¤", "≤"),  # â‰¤ -> ≤
        ("â¥", "≥"),  # â‰¥ -> ≥
    ]
    for bad, good in replacements:
        s = s.replace(bad, good)
    return s


def load_ai_summaries():
    """site_name -> summary text"""
    import openpyxl
    sums = {}
    if not SUMMARIES_XLSX.exists():
        return sums
    wb = openpyxl.load_workbook(SUMMARIES_XLSX, read_only=True)
    ws = wb.active
    rows_iter = ws.values
    hdr = next(rows_iter)
    idx = {h: i for i, h in enumerate(hdr)}
    for r in rows_iter:
        site = r[idx.get("Site Name", 1)]
        txt = r[idx.get("AI Summary (Plain Text)", 2)]
        if site and txt:
            sums[site] = _fix_encoding(str(txt)[:800])
    return sums


SITES_XLSX = HERE / "ae_sites.xlsx"

def load_site_keys():
    """site_name -> site_key (e.g. 'S59656')"""
    import openpyxl
    keys = {}
    if not SITES_XLSX.exists():
        return keys
    wb = openpyxl.load_workbook(SITES_XLSX, read_only=True)
    ws = wb["Sites Overview"]
    rows_iter = ws.values
    hdr = next(rows_iter)
    idx = {h: i for i, h in enumerate(hdr)}
    for r in rows_iter:
        keys[r[idx["Site Name"]]] = r[idx["Site Key"]]
    return keys


IRRAD_MEASUREMENT_CATS = [
    {"value": 1, "enabled": True, "checked": True, "modelOptions": None},
    {"value": 4, "enabled": True, "checked": True, "modelOptions": None},
    {"value": 8, "enabled": True, "checked": True, "modelOptions": None},
]
IRRAD_INLINE = {
    "aggregationMode": 3, "autoSource": False, "lineType": 0,
    "weatherMode": 0, "modelIndex": 0,
    "useOnSiteWeatherStations": True, "primaryWeatherSource": 10,
    "showNetEnergy": True, "powerAverage": True,
    "includePOA": True, "includeGHI": False, "fillGaps": False,
    "showExternalTemperature": False, "showDeviceTemperature": False,
    "showAggregateLayers": True, "showSourceLayers": False,
    "xSeriesKey": None,
}


PROD_CACHE = HERE / "ae_production_cache.json"


def save_prod_cache(prod_by_site, prod_bins_by_site, irrad_by_site, d_from, d_to):
    cache = {
        "d_from": d_from.isoformat(), "d_to": d_to.isoformat(),
        "saved_at": datetime.now().isoformat(),
        "prod": {s: {k: v for k, v in data.items()} for s, data in prod_by_site.items()},
        "bins": prod_bins_by_site,
        "irrad": irrad_by_site,
    }
    PROD_CACHE.write_text(json.dumps(cache), encoding="utf-8")
    print(f"  Production cache saved ({len(prod_by_site)} sites)")


def load_prod_cache():
    if not PROD_CACHE.exists():
        return {}, {}, {}, None, None
    try:
        cache = json.loads(PROD_CACHE.read_text(encoding="utf-8"))
        d_from = date.fromisoformat(cache["d_from"])
        d_to = date.fromisoformat(cache["d_to"])
        print(f"  Production cache loaded ({len(cache['prod'])} sites, {cache['d_from']}..{cache['d_to']}, saved {cache['saved_at'][:16]})")
        return cache["prod"], cache["bins"], cache.get("irrad", {}), d_from, d_to
    except Exception as e:
        print(f"  [warn] cache load failed: {e}")
        return {}, {}, {}, None, None


CHART_MEASUREMENT_CATS = [
    {"value": 1, "enabled": True, "checked": True, "modelOptions": None},
]
CHART_INLINE = {
    "aggregationMode": 3, "autoSource": False, "lineType": 0,
    "weatherMode": 0, "modelIndex": 0,
    "useOnSiteWeatherStations": True, "primaryWeatherSource": 10,
    "showNetEnergy": True, "powerAverage": True,
    "includePOA": True, "includeGHI": False, "fillGaps": False,
    "showExternalTemperature": False, "showDeviceTemperature": False,
    "showAggregateLayers": True, "showSourceLayers": False,
    "xSeriesKey": None,
}


def fetch_inverter_production(session, site_key, inv_keys_dict, d_from, d_to):
    """Fetch 15-min AC Power (kW) for all inverters at a site.
    Returns {inv_name: [kW values per 15-min bin]} and bin count."""
    hw_keys = list(inv_keys_dict.values())
    if not hw_keys or not site_key:
        return {}, 0
    key_to_name = {v: k for k, v in inv_keys_dict.items()}
    payload = {
        "chartType": 1, "binSize": 15, "context": "site",
        "start": d_from.isoformat(), "end": d_to.isoformat(),
        "futureDays": 0, "hardwareSet": hw_keys, "sectionCode": 0,
        "query": {
            "name": "CustomChart", "title": "Custom Chart",
            "initialSpan": 1, "dataItems": [],
            "kpiChart": {
                "siteKeys": [site_key],
                "categories": {
                    "measurements": CHART_MEASUREMENT_CATS,
                    "calculations": [], "losses": [], "financials": [],
                    "events": [], "special": [],
                },
                "availabilityReferenceMode": 0, "availabilityPassMode": 0,
                "availabilityPowerThreshold": 0, "weatherModes": [],
                "weatherSources": [0],
                "inlineOptions": CHART_INLINE,
            },
        },
        "source": [site_key],
    }
    try:
        r = session.post(f"{API_BASE}/view/chart?lastChanged=1900-01-01T00:00:00.000Z",
                         json=payload, timeout=120)
        if not r.ok:
            return {}, 0
        data = r.json()
        result = {}
        n_bins = 0
        hw_names_map = data.get("hardwareNames", {})
        for s in data.get("series", []):
            name_full = s.get("name", "")
            bins = s.get("dataBinned", [])
            n_bins = max(n_bins, len(bins))
            inv_part = name_full.rsplit(",", 1)[0].strip()
            if inv_part in inv_keys_dict:
                result[inv_part] = bins
                continue
            for inv_name in inv_keys_dict:
                if inv_name in name_full or inv_name == inv_part:
                    result[inv_name] = bins
                    break
        if len(result) < len(inv_keys_dict) and hw_names_map:
            name_to_inv = {}
            for hk_low, hw_name in hw_names_map.items():
                hk_upper = hk_low[0].upper() + hk_low[1:]
                for inv_name, inv_key in inv_keys_dict.items():
                    if inv_key == hk_upper:
                        name_to_inv[hw_name] = inv_name
                        break
            for s in data.get("series", []):
                name_full = s.get("name", "")
                bins = s.get("dataBinned", [])
                inv_part = name_full.rsplit(",", 1)[0].strip()
                if inv_part in name_to_inv and name_to_inv[inv_part] not in result:
                    result[name_to_inv[inv_part]] = bins
                    n_bins = max(n_bins, len(bins))
        return result, n_bins
    except Exception as e:
        print(f"  [chart] {site_key}: {e}")
        return {}, 0


def fetch_irradiance(session, site_key, weather_keys, d_from, d_to):
    """Fetch 15-min POA irradiance for a site. Returns list of W/m2 values."""
    if not weather_keys or not site_key:
        return []
    hw_key = weather_keys[0]
    payload = {
        "chartType": 1, "binSize": 15, "context": "site",
        "start": d_from.isoformat(), "end": d_to.isoformat(),
        "futureDays": 0, "hardwareSet": [hw_key], "sectionCode": 0,
        "query": {
            "name": "CustomChart", "title": "Custom Chart",
            "initialSpan": 1, "dataItems": [],
            "kpiChart": {
                "siteKeys": [site_key],
                "categories": {
                    "measurements": IRRAD_MEASUREMENT_CATS,
                    "calculations": [], "losses": [], "financials": [],
                    "events": [], "special": [],
                },
                "availabilityReferenceMode": 0, "availabilityPassMode": 0,
                "availabilityPowerThreshold": 0, "weatherModes": [],
                "weatherSources": [0],
                "inlineOptions": IRRAD_INLINE,
            },
        },
        "source": [site_key],
    }
    try:
        r = session.post(f"{API_BASE}/view/chart?lastChanged=1900-01-01T00:00:00.000Z",
                         json=payload, timeout=30)
        if not r.ok:
            return []
        data = r.json()
        for s in data.get("series", []):
            name = str(s.get("name", "")).lower()
            if "poa" in name or "irradiance" in name or "w/m" in name:
                return s.get("dataBinned", [])
        if data.get("series"):
            return data["series"][0].get("dataBinned", [])
    except Exception as e:
        print(f"  [irrad] {site_key}: {e}")
    return []


# ── Enrichment ────────────────────────────────────────────────────────────

def enrich(alerts, d_to):
    now = datetime.utcnow()
    for a in alerts:
        a["category"] = classify(a["event_type"], a["description"], a["asset_code"])
        end = a["end"] or now
        a["duration_h"] = round(max((end - a["start"]).total_seconds() / 3600, 0), 1) \
                          if a["start"] else 0
        a["score"] = importance(a)
        meta = CATEGORY_META[a["category"]]
        a["cat_label"], a["measurement"] = meta[1], meta[2]
        a["unit"] = MEASUREMENT_UNITS.get(meta[2], "")
    return alerts


# ── Aggregations for charts ──────────────────────────────────────────────

def day_range(d_from, d_to):
    days = []
    cur = d_from
    while cur <= d_to:
        days.append(cur)
        cur += timedelta(days=1)
    return days


def overlap_hours(a_start, a_end, day):
    """Hours of [a_start, a_end] overlapping calendar `day`."""
    d0 = datetime(day.year, day.month, day.day)
    d1 = d0 + timedelta(days=1)
    s = max(a_start, d0)
    e = min(a_end or datetime.utcnow(), d1)
    return max((e - s).total_seconds() / 3600, 0)


def build_inverter_heatmaps(alerts, hw, days):
    """site -> {inverter -> {day -> {'fault': h, 'comm': h, 'titles': set}}}"""
    sites = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: {"fault": 0.0, "comm": 0.0, "titles": set()})))
    for a in alerts:
        if a["category"] not in ("INVERTER_FAULT", "INVERTER_COMM"):
            continue
        if not a["hardware_name"] or not a["start"]:
            continue
        kind = "fault" if a["category"] == "INVERTER_FAULT" else "comm"
        for d in days:
            h = overlap_hours(a["start"], a["end"], d)
            if h > 0.01:
                cell = sites[a["site_name"]][a["hardware_name"]][d]
                cell[kind] += h
                cell["titles"].add(f"{a['event_type']}: {a['description'][:80]}")
    # ensure every inverter of an affected site shows (healthy rows too)
    for site in list(sites.keys()):
        for inv in hw.get(site, {}).get("inverters", []):
            _ = sites[site][inv]
    return sites


TCU_PATTERNS = [
    (re.compile(r"TCU Fault\s*(\d+)", re.I), "TCU Fault {}"),
    (re.compile(r"Individual Tracker Alarm\s*(\d+):\s*([^\t\n]+)", re.I), "Tracker {}: {}"),
    (re.compile(r"Slave Alarms?\s*S?\s*(\d+):\s*([^\t\n]+)", re.I), "Slave {}: {}"),
]

def parse_tcu_details(description):
    out = []
    for rx, fmt in TCU_PATTERNS:
        for m in rx.finditer(description or ""):
            out.append(fmt.format(*m.groups()))
    return out


def build_tracker_data(alerts, days):
    """site -> {'per_day': {day: {'fault': n, 'comm': n}}, 'details': [..]}"""
    sites = defaultdict(lambda: {"per_day": defaultdict(lambda: {"fault": 0, "comm": 0}),
                                 "details": []})
    for a in alerts:
        if a["category"] not in ("TRACKER_FAULT", "TRACKER_COMM"):
            continue
        kind = "fault" if a["category"] == "TRACKER_FAULT" else "comm"
        d = a["start"].date() if a["start"] else None
        if d and d in [x for x in days]:
            sites[a["site_name"]]["per_day"][d][kind] += 1
        det = parse_tcu_details(a["description"])
        for item in det[:12]:
            sites[a["site_name"]]["details"].append({
                "when": a["start"].strftime("%m-%d %H:%M") if a["start"] else "",
                "hw": a["hardware_name"] or "", "item": item,
                "resolved": a["is_resolved"],
            })
    return sites


# ── HTML rendering ────────────────────────────────────────────────────────

def esc(s):
    return (str(s or "").replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def heat_color(fault_h, comm_h):
    """Red scale for fault hours, orange for comm-only."""
    if fault_h > 0.05:
        a = min(0.15 + fault_h / 12.0, 1.0)
        return f"rgba(198,40,40,{a:.2f})"
    if comm_h > 0.05:
        a = min(0.15 + comm_h / 16.0, 0.9)
        return f"rgba(239,140,30,{a:.2f})"
    return "#eef3ea"


def build_strip_data(site, hw, d_from, d_to, production=None, n_bins_api=0):
    """Build 15-min strip data for the heatmap.

    If production data is available (from chart API), compute CF%.
    Otherwise fall back to alert-derived binary availability.

    Returns dict for JS:
      invs    : inverter names
      bins    : number of 15-min bins
      binMin  : 15
      vals    : per inverter, per bin: CF% (0-100) or null (no data)
      kw      : per inverter, per bin: raw kW values (for tooltip)
      caps    : per inverter: rated capacity kW
      alerts  : per inverter, alert strings
    """
    start = datetime(d_from.year, d_from.month, d_from.day)
    end = datetime(d_to.year, d_to.month, d_to.day) + timedelta(days=1)
    bins_15 = int((end - start).total_seconds() // 900)

    hwd = hw.get(site, {})
    inv_names = sorted(hwd.get("inverters", []),
                       key=lambda x: (len(str(x)), str(x)))
    if not inv_names:
        return {"invs": [], "bins": 0, "binMin": 15,
                "start": start.strftime("%Y-%m-%d %H:%M"),
                "vals": [], "kw": [], "caps": [], "alerts": []}

    inv_caps = hwd.get("inv_cap", {})
    caps = [inv_caps.get(n, 0) for n in inv_names]

    if production and n_bins_api > 0:
        n_bins = n_bins_api
        vals = []
        kw_data = []
        for name in inv_names:
            kw_raw = production.get(name, [])
            cap = inv_caps.get(name, 0)
            if not cap and kw_raw:
                cap = max((v for v in kw_raw if v is not None), default=0) * 1.05
                if cap > 0:
                    caps[inv_names.index(name)] = round(cap, 1)
            row_cf = []
            row_kw = []
            for i in range(n_bins):
                v = kw_raw[i] if i < len(kw_raw) else None
                if v is None:
                    row_cf.append(None)
                    row_kw.append(None)
                else:
                    row_kw.append(round(v, 2))
                    row_cf.append(round(v / cap * 100, 2) if cap > 0 else 0)
            vals.append(row_cf)
            kw_data.append(row_kw)
    else:
        n_bins = bins_15
        vals = [[None] * n_bins for _ in inv_names]
        kw_data = [[None] * n_bins for _ in inv_names]

    return {"invs": inv_names, "bins": n_bins, "binMin": 15,
            "start": start.strftime("%Y-%m-%d %H:%M"),
            "vals": vals, "kw": kw_data, "caps": caps,
            "alerts": [[] for _ in inv_names]}


def apply_alerts_to_strips(strip, site_alerts, d_from):
    """Annotate strip data with alert info for tooltips."""
    start = datetime(d_from.year, d_from.month, d_from.day)
    idx = {name: i for i, name in enumerate(strip["invs"])}
    for a in site_alerts:
        if a["category"] not in ("INVERTER_FAULT", "INVERTER_COMM"):
            continue
        i = idx.get(a["hardware_name"])
        if i is None or not a["start"]:
            continue
        t = a["start"].strftime("%m-%d %H:%M")
        state = "OPEN" if not a["is_resolved"] else f"{a['duration_h']:.1f}h"
        strip["alerts"][i] = (strip["alerts"][i] +
            [f"{t} · {a['event_type']} ({state})"])[:8]
    return strip


def render_email_items(emails, limit=4):
    if not emails:
        return ""
    items = []
    for e in emails[:limit]:
        date = e["date"]
        cat = e["category"]
        subj = e["subject"][:80]
        body = e["contents"][:200]
        items.append(f"<li class='email-ev'><b>[{esc(cat)}]</b> {esc(date)} — {esc(subj)}<br>"
                     f"<span class='email-body'>{esc(body)}</span></li>")
    return "".join(items)


def render_diag_card(d, rank=None, emails=None):
    ev = "".join(f"<li>{esc(e)}</li>" for e in d["evidence"][:6])
    email_ev = render_email_items(emails or [])
    rk = f"<span class='diag-rank'>#{rank}</span>" if rank else ""
    return f"""
    <div class="diag-card">
      <div class="diag-head">{rk}<b>{esc(d['site'])}</b> — {esc(d['title'])}</div>
      <ul class="diag-ev">{ev}{email_ev}</ul>
      <div class="diag-action">→ {esc(d['action'])}</div>
    </div>"""


def render_site_emails(emails):
    if not emails:
        return ""
    rows = []
    for e in emails[:6]:
        rows.append(f"""<tr>
          <td>{esc(e['date'])}</td>
          <td><span class="email-cat">{esc(e['category'])}</span></td>
          <td><b>{esc(e['subject'][:60])}</b><br><span class="email-body">{esc(e['contents'][:250])}</span></td>
          <td class="muted">{esc(e['from'][:50])}</td>
        </tr>""")
    return f"""<div class="email-section">
      <div class="email-header">Site Emails ({len(emails)})</div>
      <table class="mini"><tr><th>Date</th><th>Category</th><th>Subject / Summary</th><th>From</th></tr>
      {''.join(rows)}</table></div>"""


def render_site_accordion(sid, site, strip, site_diags, fault_h, comm_h,
                          ai_summary="", irrad_data=None, site_emails=None):
    """Clickable <details> per site: badge + canvas strip heatmap + diagnoses + AI context + irradiance."""
    n_inv = len(strip["invs"])
    if fault_h > 1:
        badge = f"<span class='acc-badge acc-red'>{fault_h:.0f}h fault</span>"
    elif comm_h > 1:
        badge = f"<span class='acc-badge acc-orange'>{comm_h:.0f}h comm</span>"
    else:
        badge = "<span class='acc-badge acc-green'>healthy</span>"
    diag_html = "".join(render_diag_card(d) for d in site_diags[:3])
    emails_html = render_site_emails(site_emails) if site_emails else ""
    open_attr = " open" if fault_h > 1 or any(d["open"] for d in site_diags) or site_emails else ""
    summary_html = ""
    if ai_summary:
        summary_html = f"""<div class="ai-summary"><div class="ai-label">AI Site Summary</div>{esc(ai_summary)}</div>"""
    if irrad_data:
        strip["irrad"] = irrad_data
    return f"""
    <details class="site-acc" data-site="{esc(site.lower())}"{open_attr}>
      <summary><span class="acc-name">{esc(site)}</span>
        <span class="acc-meta">{n_inv} inverters</span>{badge}</summary>
      <div class="acc-body">
        {emails_html}
        {summary_html}
        {diag_html}
        <div class="hm-controls">
          <label>View:</label>
          <select id="sel{sid}" onchange="switchMode({sid})">
            <option value="cf">Capacity Factor (%)</option>
            <option value="kw">Production &ndash; AC (kW)</option>
          </select>
        </div>
        <div class="strip-wrap">
          <div class="strip-labels" id="lb{sid}"></div>
          <div style="flex:1;min-width:0"><canvas id="hm{sid}" class="strip-canvas"></canvas></div>
        </div>
        <div class="legend" style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">
          <span class="legend-title" id="lt{sid}">Capacity Factor (%)</span>
          <div class="scale-ruler">
            <div class="scalebar"></div>
            <div class="scale-ticks" id="st{sid}">
              <span>0%</span><span>20%</span><span>40%</span><span>60%</span><span>80%</span><span>100%</span>
            </div>
          </div>
          <span class="legend"><i style="background:#e8e8e8"></i>No data</span>
          <span class="legend" style="display:flex;align-items:center;gap:3px">
            <span class="night-swatch"></span>Night</span>
        </div>
      </div>
      <script type="application/json" id="d{sid}">{json.dumps(strip)}</script>
    </details>"""


def render_tracker_html(site, data, days, chart_id):
    labels = [d.strftime("%m/%d") for d in days]
    faults = [data["per_day"].get(d, {}).get("fault", 0) for d in days]
    comms  = [data["per_day"].get(d, {}).get("comm", 0) for d in days]
    det_rows = "".join(
        f"<tr><td>{esc(d['when'])}</td><td>{esc(d['hw'])}</td><td>{esc(d['item'])}</td>"
        f"<td>{'✔' if d['resolved'] else '<b class=red>open</b>'}</td></tr>"
        for d in data["details"][:25])
    details_tbl = (f"<table class='mini'><tr><th>When</th><th>Controller</th>"
                   f"<th>TCU / tracker detail</th><th>Resolved</th></tr>{det_rows}</table>"
                   if det_rows else "<div class='muted'>No per-TCU details parsed.</div>")
    return f"""
    <div class="card">
      <div class="card-title">Tracker / TCU — {esc(site)}
        <span class="unit-tag">alarms per day · context: tracker angle (°), SOC (%)</span></div>
      <div class="chart-wrap"><canvas id="{chart_id}"></canvas></div>
      {details_tbl}
    </div>
    <script>
    new Chart(document.getElementById('{chart_id}'), {{
      type: 'bar',
      data: {{ labels: {json.dumps(labels)},
        datasets: [
          {{label: 'TCU faults', data: {json.dumps(faults)}, backgroundColor: 'rgba(198,40,40,.75)'}},
          {{label: 'Comm loss', data: {json.dumps(comms)}, backgroundColor: 'rgba(239,140,30,.7)'}}
        ]}},
      options: {{ responsive: true, maintainAspectRatio: false,
        scales: {{ x: {{stacked: true}}, y: {{stacked: true, beginAtZero: true,
          title: {{display: true, text: 'Alarm count'}}, ticks: {{precision: 0}}}}}},
        plugins: {{legend: {{position: 'bottom'}}}}}}
    }});
    </script>"""


def render_critical_table(alerts):
    crit = [a for a in alerts if a["category"] in
            ("INVERTER_FAULT", "GRID", "TRACKER_FAULT", "METER")]
    crit.sort(key=lambda a: a["start"] or datetime.min, reverse=True)
    sites = sorted(set(a["site_name"] for a in crit))
    rows = []
    for a in crit:
        sev = SEVERITY_LABEL.get(a["severity"], a["severity"])
        status = "✔ resolved" if a["is_resolved"] else "<b class='red'>OPEN</b>"
        ts = a["start"].strftime("%Y-%m-%d %H:%M") if a["start"] else ""
        ts_display = a["start"].strftime("%m-%d %H:%M") if a["start"] else "—"
        rows.append(f"""<tr data-site="{esc(a['site_name'])}" data-ts="{ts}" data-cat="{esc(a['cat_label'])}">
          <td><span class="pill pill-{a['category'].lower()}">{esc(a['cat_label'])}</span></td>
          <td>{esc(a['site_name'])}</td>
          <td title="{esc(a['hardware_name'])}">{esc((a['hardware_name'] or '—')[:34])}</td>
          <td title="{esc(a['description'])}">{esc(a['event_type'])}<div class="desc">{esc(a['description'][:120])}</div></td>
          <td>{ts_display}</td>
          <td>{a['duration_h']:.1f} h</td>
          <td>{esc(sev)}</td>
          <td>{status}</td>
          <td class="muted">{esc(a['measurement'])} ({esc(a['unit'])})</td>
        </tr>""")
    if not rows:
        rows = ["<tr><td colspan='9' class='muted'>No critical alerts in window</td></tr>"]
    site_opts = "".join(f'<option value="{esc(s)}">{esc(s)}</option>' for s in sites)
    cats = sorted(set(a["cat_label"] for a in crit))
    cat_opts = "".join(f'<option value="{esc(c)}">{esc(c)}</option>' for c in cats)
    return f"""<div class="alert-filters">
      <select id="alertSiteFilter" onchange="filterAlerts()">
        <option value="">All sites</option>{site_opts}</select>
      <select id="alertCatFilter" onchange="filterAlerts()">
        <option value="">All categories</option>{cat_opts}</select>
      <label>From:</label><input type="date" id="alertDateFrom" onchange="filterAlerts()">
      <label>To:</label><input type="date" id="alertDateTo" onchange="filterAlerts()">
      <span class="muted" id="alertCount">{len(rows)} alerts</span>
    </div>
    <table class="big" id="alertTable">
      <tr><th>Category</th><th>Site</th><th>Device</th><th>Alert</th>
          <th>Start</th><th>Duration</th><th>Severity</th><th>Status</th><th>Check unit</th></tr>
      {''.join(rows)}</table>"""


PAGE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Operational Alerts — SolRiver Capital</title>
<script>/* Chart.js v4.4.1 inlined for offline use */
{chartjs}
</script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f4f6f9; color: #333; }}
  header {{ background: #1F4E79; color: #fff; padding: 14px 24px; display: flex; justify-content: space-between; align-items: center; }}
  header h1 {{ font-size: 20px; }}
  header .meta {{ font-size: 12px; opacity: .75; text-align: right; }}
  .nav {{ background: #2E75B6; padding: 0 24px; display: flex; align-items: stretch; gap: 0; }}
  .nav a {{ color: #cde; text-decoration: none; font-size: 13px; padding: 10px 16px; }}
  .nav a:hover {{ background: rgba(255,255,255,.1); }}
  .nav a.tab-link.active {{ color: #fff; background: rgba(255,255,255,.15); border-bottom: 2px solid #fff; font-weight: 600; }}
  .tab-panel {{ display: none; }}
  .tab-panel.active {{ display: block; }}
  .kpis {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 14px; padding: 18px 24px 0; }}
  .card {{ background: #fff; border-radius: 6px; box-shadow: 0 1px 4px rgba(0,0,0,.12); padding: 16px; }}
  .card-title {{ font-size: 12px; color: #777; text-transform: uppercase; letter-spacing: .5px; margin-bottom: 10px; }}
  .unit-tag {{ float: right; text-transform: none; letter-spacing: 0; color: #2E75B6; font-size: 11px; }}
  .big-num {{ font-size: 32px; font-weight: 700; color: #1F4E79; }}
  .big-num.red {{ color: #c62828; }}
  .big-num.orange {{ color: #ef8c1e; }}
  .sub {{ font-size: 12px; color: #888; margin-top: 2px; }}
  section {{ padding: 16px 24px 0; }}
  section:last-of-type {{ padding-bottom: 24px; }}
  h2 {{ font-size: 15px; color: #1F4E79; margin: 4px 0 10px; }}
  table.big {{ width: 100%; border-collapse: collapse; font-size: 12.5px; background:#fff; }}
  table.big th {{ background: #1F4E79; color: #fff; font-size: 11px; padding: 6px 8px; text-align: left; position: sticky; top: 0; }}
  table.big td {{ padding: 6px 8px; border-bottom: 1px solid #eef; vertical-align: top; }}
  table.big tr:nth-child(even) td {{ background: #f6f9fd; }}
  .desc {{ color: #888; font-size: 11px; margin-top: 2px; }}
  .red {{ color: #c62828; }}
  .muted {{ color: #999; font-size: 12px; }}
  .pill {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; white-space: nowrap; }}
  .pill-inverter_fault {{ background: #fde3e3; color: #b71c1c; }}
  .pill-grid {{ background: #ede3fd; color: #4a148c; }}
  .pill-tracker_fault {{ background: #fff0db; color: #a35e00; }}
  .pill-meter {{ background: #e0f0ff; color: #0d47a1; }}
  .grid2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
  .hm-card {{ overflow-x: auto; }}
  .hm {{ display: block; min-width: 520px; }}
  .hm-r {{ display: grid; grid-template-columns: 210px repeat(var(--days), 1fr); gap: 3px; margin-bottom: 3px; }}
  .hm-l {{ font-size: 11px; color: #555; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; line-height: 26px; }}
  .hm-h {{ font-size: 10px; color: #888; text-align: center; }}
  .hm-c {{ height: 26px; border-radius: 3px; font-size: 10px; color: #333; display: flex; align-items: center; justify-content: center; }}
  .legend {{ margin-top: 10px; font-size: 11px; color: #666; }}
  .legend span {{ margin-right: 16px; }}
  .legend i {{ display: inline-block; width: 12px; height: 12px; border-radius: 2px; vertical-align: -2px; margin-right: 4px; }}
  .chart-wrap {{ position: relative; height: 220px; margin-bottom: 10px; }}
  table.mini {{ width: 100%; border-collapse: collapse; font-size: 11.5px; }}
  table.mini th {{ background: #eef3fa; color: #1F4E79; padding: 4px 6px; text-align: left; font-size: 10.5px; }}
  table.mini td {{ padding: 4px 6px; border-bottom: 1px solid #f0f0f0; }}
  /* diagnosis / recommender */
  .diag-card {{ background: #fff; border-left: 4px solid #1F4E79; border-radius: 4px;
               box-shadow: 0 1px 3px rgba(0,0,0,.1); padding: 12px 14px; margin-bottom: 10px; }}
  .diag-head {{ font-size: 13.5px; color: #222; }}
  .diag-rank {{ display: inline-block; background: #1F4E79; color: #fff; border-radius: 50%;
               width: 22px; height: 22px; text-align: center; line-height: 22px;
               font-size: 11px; font-weight: 700; margin-right: 8px; }}
  .diag-badge {{ font-size: 10.5px; font-weight: 700; padding: 2px 8px; border-radius: 10px; margin-left: 8px; }}
  .diag-high {{ background: #fde3e3; color: #b71c1c; }}
  .diag-med  {{ background: #fff0db; color: #a35e00; }}
  .diag-low  {{ background: #eef3fa; color: #2E75B6; }}
  .diag-impact {{ float: right; font-size: 11px; color: #c62828; font-weight: 600; }}
  .diag-ev {{ margin: 8px 0 6px 18px; font-size: 11.5px; color: #555; }}
  .diag-ev li {{ margin-bottom: 2px; }}
  .diag-action {{ font-size: 12.5px; color: #155724; background: #f0f8f1;
                 border-radius: 4px; padding: 6px 10px; }}
  /* site accordion */
  .site-acc {{ background: #fff; border-radius: 6px; box-shadow: 0 1px 4px rgba(0,0,0,.12);
              margin-bottom: 8px; overflow-x: auto; }}
  .site-acc summary {{ cursor: pointer; padding: 12px 16px; font-size: 14px;
                      display: flex; align-items: center; gap: 12px; user-select: none; }}
  .site-acc summary:hover {{ background: #f6f9fd; }}
  .acc-name {{ font-weight: 600; color: #1F4E79; min-width: 240px; }}
  .acc-meta {{ font-size: 11px; color: #999; }}
  .acc-badge {{ font-size: 11px; font-weight: 700; padding: 2px 10px; border-radius: 10px; margin-left: auto; }}
  .acc-red {{ background: #fde3e3; color: #b71c1c; }}
  .acc-orange {{ background: #fff0db; color: #a35e00; }}
  .acc-green {{ background: #e3f4e4; color: #1b6e23; }}
  .acc-body {{ padding: 4px 16px 14px; }}
  /* strip heatmap (PowerTrack inverter-heatmap style) */
  .strip-wrap {{ display: flex; gap: 8px; margin: 8px 0 6px; }}
  .strip-labels {{ display: flex; flex-direction: column; }}
  .strip-labels div {{ height: 14px; line-height: 14px; font-size: 10px; color: #555;
                      white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
                      max-width: 200px; min-width: 140px; }}
  .strip-labels div.avg {{ font-weight: 700; color: #1F4E79; }}
  .strip-canvas {{ width: 100%; display: block; }}
  .scale-ruler {{ display: inline-flex; flex-direction: column; }}
  .scalebar {{ width: 200px; height: 12px; border-radius: 2px;
    background: linear-gradient(90deg,#00007f,#0000ff,#007fff,#00ffff,#7fff7f,#ffff00,#ff7f00,#ff0000,#7f0000); }}
  .scale-ticks {{ display: flex; justify-content: space-between; width: 200px; }}
  .scale-ticks span {{ font-size: 9px; color: #888; }}
  .hm-tip {{ position: fixed; background: rgba(20,30,40,.95); color: #fff; font-size: 11px;
            padding: 6px 9px; border-radius: 4px; pointer-events: none; z-index: 99;
            max-width: 340px; display: none; line-height: 1.5; }}
  .controls {{ display: flex; gap: 10px; align-items: center; margin-bottom: 10px; }}
  .controls input {{ padding: 6px 10px; border: 1px solid #ccd; border-radius: 4px;
                    font-size: 13px; width: 260px; }}
  .controls button {{ padding: 6px 12px; border: 1px solid #2E75B6; background: #fff;
                     color: #2E75B6; border-radius: 4px; font-size: 12px; cursor: pointer; }}
  .controls button:hover {{ background: #2E75B6; color: #fff; }}
  .ai-summary {{ background: #f0f6ff; border-left: 3px solid #2E75B6; border-radius: 4px;
               padding: 10px 14px; margin-bottom: 10px; font-size: 12px; color: #444; line-height: 1.5; }}
  .ai-label {{ font-size: 10px; font-weight: 700; color: #2E75B6; text-transform: uppercase;
              letter-spacing: .5px; margin-bottom: 4px; }}
  .hm-controls {{ display: flex; align-items: center; gap: 8px; margin-bottom: 6px; }}
  .hm-controls label {{ font-size: 11px; color: #666; font-weight: 600; }}
  .hm-controls select {{ font-size: 12px; padding: 3px 8px; border: 1px solid #bcd; border-radius: 4px;
                        background: #fff; color: #333; cursor: pointer; }}
  .email-section {{ margin-bottom: 10px; }}
  .email-header {{ font-size: 11px; font-weight: 700; color: #8e4500; margin-bottom: 4px; }}
  .email-cat {{ font-size: 10px; font-weight: 600; color: #8e4500; background: #fff3e0;
               padding: 1px 6px; border-radius: 8px; white-space: nowrap; }}
  .email-body {{ font-size: 11px; color: #666; line-height: 1.4; }}
  .email-ev {{ border-left: 2px solid #ef8c1e; padding-left: 6px; margin-top: 4px; }}
  .alert-filters {{ display: flex; align-items: center; gap: 10px; padding: 10px 12px; background: #f6f9fd;
                   border-bottom: 1px solid #e0e6ef; flex-wrap: wrap; }}
  .alert-filters select, .alert-filters input[type="date"] {{ font-size: 12px; padding: 4px 8px; border: 1px solid #bcd;
                   border-radius: 4px; background: #fff; }}
  .alert-filters label {{ font-size: 11px; color: #666; }}
  .legend-title {{ font-size: 11px; color: #555; font-weight: 600; }}
  .night-swatch {{ display: inline-block; width: 14px; height: 12px; border-radius: 2px;
    background: repeating-linear-gradient(-45deg, #fff, #fff 2px, #d0d0d0 2px, #d0d0d0 3px); border: 1px solid #ccc; }}
  footer {{ text-align: center; font-size: 11px; color: #aaa; padding: 14px; }}
</style>
</head>
<body>
<header>
  <div>
    <h1>Operational Alerts</h1>
    <div style="font-size:13px;margin-top:2px;opacity:.8">SolRiver Capital</div>
  </div>
  <div class="meta">Generated: {generated_at}<br><a href="index.html" style="color:#cde">← Portfolio Overview</a></div>
</header>
<div class="nav">
  <a href="index.html">Portfolio</a>
  <a href="#" class="tab-link active" onclick="switchTab('diagnosis',this);return false">Diagnosis</a>
  <a href="#" class="tab-link" onclick="switchTab('heatmaps',this);return false">Inverter Heatmaps</a>
  <a href="#" class="tab-link" onclick="switchTab('trackers',this);return false">Tracker / TCU</a>
  <a href="#" class="tab-link" onclick="switchTab('alerts',this);return false">Critical Alerts</a>
  <a href="#" class="tab-link" onclick="switchTab('breakdown',this);return false">Category Breakdown</a>
</div>

<div class="tab-panel active" id="tab-diagnosis">
<section>
  <h2>Diagnosis &amp; Recommended Actions
    <span class="muted">(cross-referenced: alerts + rule tool + AI summaries + live power + emails)</span></h2>
  {recommender}
</section>
</div>

<div class="tab-panel" id="tab-heatmaps">
<section>
  <h2>Inverter Heatmaps — All Sites
    <span class="muted">— click to expand · 15-min capacity factor</span></h2>
  <div class="controls">
    <input id="siteSearch" type="text" placeholder="Search sites…" oninput="filterSites(this.value)">
    <button onclick="setAll(true)">Expand all</button>
    <button onclick="setAll(false)">Collapse all</button>
  </div>
  {heatmaps}
</section>
<div class="hm-tip" id="hmTip"></div>
</div>

<div class="tab-panel" id="tab-trackers">
<section>
  <h2>Tracker / TCU Issues</h2>
  <div class="grid2">{trackers}</div>
</section>
</div>

<div class="tab-panel" id="tab-alerts">
<section>
  <h2>Critical Alerts — faults, production stops, grid events</h2>
  <div class="card" style="padding:0; max-height: 700px; overflow-y:auto">{critical_table}</div>
</section>
</div>

<div class="tab-panel" id="tab-breakdown">
<section>
  <h2>Category Breakdown</h2>
  <div class="grid2">
    <div class="card"><div class="card-title">Alerts by category</div>
      <div class="chart-wrap"><canvas id="catChart"></canvas></div></div>
    <div class="card"><div class="card-title">Alerts per day (all categories)</div>
      <div class="chart-wrap"><canvas id="dayChart"></canvas></div></div>
  </div>
</section>
</div>

<script>
function switchTab(id, el) {{
  document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab-link').forEach(a => a.classList.remove('active'));
  document.getElementById('tab-' + id).classList.add('active');
  if (el) el.classList.add('active');
  /* trigger heatmap draw for newly visible canvases */
  if (id === 'heatmaps') {{
    document.querySelectorAll('details.site-acc[open] canvas.strip-canvas').forEach(cv => {{
      if (!cv.dataset.drawn) {{
        const sid = cv.id.slice(2);
        drawStrip(sid);
      }}
    }});
  }}
}}
function filterAlerts() {{
  const site = document.getElementById('alertSiteFilter').value;
  const cat = document.getElementById('alertCatFilter').value;
  const from = document.getElementById('alertDateFrom').value;
  const to = document.getElementById('alertDateTo').value;
  const rows = document.querySelectorAll('#alertTable tr[data-site]');
  let shown = 0;
  rows.forEach(tr => {{
    const rs = tr.dataset.site;
    const rc = tr.dataset.cat;
    const rt = tr.dataset.ts;
    let show = true;
    if (site && rs !== site) show = false;
    if (cat && rc !== cat) show = false;
    if (from && rt < from) show = false;
    if (to && rt > to + 'T23:59') show = false;
    tr.style.display = show ? '' : 'none';
    if (show) shown++;
  }});
  document.getElementById('alertCount').textContent = shown + ' alerts';
}}
new Chart(document.getElementById('catChart'), {{
  type: 'doughnut',
  data: {{ labels: {cat_labels}, datasets: [{{ data: {cat_counts},
    backgroundColor: ['#c62828','#7b1fa2','#ef8c1e','#0d47a1','#f4b942','#888','#5cb85c','#2E75B6','#aaa','#ccc'] }}] }},
  options: {{ responsive: true, maintainAspectRatio: false, plugins: {{legend: {{position: 'right'}}}} }}
}});
new Chart(document.getElementById('dayChart'), {{
  type: 'line',
  data: {{ labels: {day_labels}, datasets: [
    {{label: 'Critical', data: {day_crit}, borderColor: '#c62828', backgroundColor: 'rgba(198,40,40,.15)', fill: true, tension: .3}},
    {{label: 'All (filtered)', data: {day_all}, borderColor: '#2E75B6', backgroundColor: 'rgba(46,117,182,.1)', fill: true, tension: .3}}
  ]}},
  options: {{ responsive: true, maintainAspectRatio: false,
    scales: {{y: {{beginAtZero: true, title: {{display: true, text: 'Alert count'}}, ticks: {{precision: 0}}}}}},
    plugins: {{legend: {{position: 'bottom'}}}} }}
}});

/* ── AlsoEnergy PowerTrack heatmap (jet colormap, 15-min bins) ──── */
function jet(v) {{
  const r = Math.max(0, Math.min(255, Math.round(255*(1.5 - Math.abs(4*v - 3)))));
  const g = Math.max(0, Math.min(255, Math.round(255*(1.5 - Math.abs(4*v - 2)))));
  const b = Math.max(0, Math.min(255, Math.round(255*(1.5 - Math.abs(4*v - 1)))));
  return [r, g, b];
}}

const ROW_H = 14, GAP = 1, AVG_H = 16, IRR_H = 16, AXIS_H = 36;
const stripCache = {{}};

function drawStrip(sid, mode) {{
  const dEl = document.getElementById('d' + sid);
  const cv = document.getElementById('hm' + sid);
  if (!dEl || !cv) return;
  if (!stripCache[sid]) stripCache[sid] = JSON.parse(dEl.textContent);
  const data = stripCache[sid];
  const nInv = data.invs.length;
  if (!nInv) return;
  mode = mode || (document.getElementById('sel' + sid) || {{}}).value || 'cf';
  const bm = data.binMin || 15;
  const binsPerHour = 60 / bm;
  const hasIrrad = data.irrad && data.irrad.length > 0;
  const irradOffset = hasIrrad ? IRR_H : 0;
  const totalH = irradOffset + AVG_H + nInv * ROW_H + AXIS_H;
  const wCss = Math.max(cv.parentElement.clientWidth, 500);
  const dpr = window.devicePixelRatio || 1;
  cv.width = wCss * dpr; cv.height = totalH * dpr;
  cv.style.height = totalH + 'px';
  const ctx = cv.getContext('2d'); ctx.scale(dpr, dpr);
  const bw = wCss / data.bins;
  const t0 = new Date(data.start.replace(' ', 'T'));

  function isNight(binIdx) {{
    const hr = (t0.getHours() + binIdx / binsPerHour) % 24;
    return hr < 5.5 || hr >= 20.5;
  }}

  const nightPat = (() => {{
    const pc = document.createElement('canvas');
    pc.width = 4; pc.height = 4;
    const pctx = pc.getContext('2d');
    pctx.fillStyle = '#fff'; pctx.fillRect(0, 0, 4, 4);
    pctx.strokeStyle = '#d0d0d0'; pctx.lineWidth = 0.5;
    pctx.beginPath(); pctx.moveTo(0, 4); pctx.lineTo(4, 0); pctx.stroke();
    return ctx.createPattern(pc, 'repeat');
  }})();

  /* pick data source based on mode */
  const isCF = mode === 'cf';
  const srcRows = isCF ? data.vals : data.kw;

  /* compute max kW per inverter for kW-mode scaling */
  let maxKwPerInv = [];
  if (!isCF) {{
    maxKwPerInv = data.caps.map((c, i) => {{
      if (c > 0) return c;
      let mx = 0;
      if (data.kw[i]) for (const v of data.kw[i]) if (v !== null && v > mx) mx = v;
      return mx || 1;
    }});
  }}

  /* compute average row */
  const avg = [];
  for (let b = 0; b < data.bins; b++) {{
    let s = 0, n = 0;
    for (let i = 0; i < srcRows.length; i++) {{
      const v = srcRows[i][b];
      if (v !== null && v !== undefined && v >= 0) {{
        s += isCF ? v : (v / (maxKwPerInv[i] || 1) * 100);
        n++;
      }}
    }}
    avg.push(n ? s / n : null);
  }}

  function paintRow(rowVals, y, h, scale) {{
    for (let b = 0; b < data.bins; b++) {{
      const v = rowVals[b];
      const night = isNight(b);
      if (v === null || v === undefined) {{
        ctx.fillStyle = night ? nightPat : '#e8e8e8';
      }} else if (night && v < (isCF ? 1 : 0.5)) {{
        ctx.fillStyle = nightPat;
      }} else {{
        const norm = Math.min(v / scale, 1);
        const c = jet(norm);
        ctx.fillStyle = `rgb(${{c[0]}},${{c[1]}},${{c[2]}})`;
      }}
      ctx.fillRect(b * bw, y, Math.ceil(bw) + 0.5, h - GAP);
    }}
  }}

  let y = 0;

  /* irradiance heatmap row */
  if (hasIrrad) {{
    for (let b = 0; b < data.bins; b++) {{
      const idx = Math.min(b, data.irrad.length - 1);
      const v = data.irrad[idx];
      const night = isNight(b);
      if (v === null || v === undefined) {{
        ctx.fillStyle = night ? nightPat : '#e8e8e8';
      }} else if (night && v < 5) {{
        ctx.fillStyle = nightPat;
      }} else {{
        const norm = Math.min(v / 1200, 1);
        const c = jet(norm);
        ctx.fillStyle = `rgb(${{c[0]}},${{c[1]}},${{c[2]}})`;
      }}
      ctx.fillRect(b * bw, y, Math.ceil(bw) + 0.5, IRR_H - GAP);
    }}
    y += IRR_H;
  }}

  /* average row */
  paintRow(avg, y, AVG_H, 100);
  y += AVG_H;

  /* inverter rows */
  srcRows.forEach((row, i) => {{
    const scale = isCF ? 100 : (maxKwPerInv[i] || 1);
    paintRow(row, y + i * ROW_H, ROW_H, scale);
  }});

  /* x-axis */
  const gridTop = irradOffset + AVG_H + nInv * ROW_H;
  const tickHours = [5, 8, 12, 16, 20];
  const tickLabels = ['5a', '8a', '12p', '4p', '8p'];
  ctx.font = '11px Segoe UI, Arial';
  for (let b = 0; b < data.bins; b++) {{
    const dt = new Date(t0.getTime() + b * bm * 60000);
    const hr = dt.getHours(), mn = dt.getMinutes();
    if (hr === 0 && mn === 0) {{
      ctx.fillStyle = 'rgba(0,0,0,.2)';
      ctx.fillRect(b * bw, 0, 1, gridTop);
    }}
    const ti = tickHours.indexOf(hr);
    if (ti >= 0 && mn === 0) {{
      ctx.fillStyle = '#777'; ctx.textAlign = 'center';
      ctx.fillText(tickLabels[ti], b * bw, gridTop + 14);
    }}
  }}
  ctx.fillStyle = '#444'; ctx.font = 'bold 12px Segoe UI, Arial'; ctx.textAlign = 'center';
  for (let b = 0; b < data.bins; b += binsPerHour * 24) {{
    const d = new Date(t0.getTime() + b * bm * 60000);
    ctx.fillText((d.getMonth()+1) + '/' + d.getDate(), (b + binsPerHour * 12) * bw, gridTop + 30);
  }}

  /* labels */
  const lb = document.getElementById('lb' + sid);
  let labelsHtml = '';
  if (hasIrrad) labelsHtml += '<div style="height:' + IRR_H + 'px;line-height:' + IRR_H + 'px;color:#d48800;font-weight:600">Irradiance</div>';
  labelsHtml += '<div class="avg" style="height:' + AVG_H + 'px;line-height:' + AVG_H + 'px">Average</div>';
  labelsHtml += data.invs.map(n => `<div title="${{n}}">${{n}}</div>`).join('');
  lb.innerHTML = labelsHtml;
  cv.dataset.drawn = '1';
  cv.dataset.mode = mode;

  /* update legend */
  const lt = document.getElementById('lt' + sid);
  const st = document.getElementById('st' + sid);
  if (lt) lt.textContent = isCF ? 'Capacity Factor (%)' : 'AC Power (kW)';
  if (st) {{
    if (isCF) {{
      st.innerHTML = '<span>0%</span><span>20%</span><span>40%</span><span>60%</span><span>80%</span><span>100%</span>';
    }} else {{
      st.innerHTML = '<span>0</span><span>20%</span><span>40%</span><span>60%</span><span>80%</span><span>rated</span>';
    }}
  }}

  /* tooltip */
  const tip = document.getElementById('hmTip');
  cv.onmousemove = e => {{
    const rect = cv.getBoundingClientRect();
    const mx = e.clientX - rect.left, my = e.clientY - rect.top;
    const b = Math.floor(mx / bw);
    if (b < 0 || b >= data.bins) {{ tip.style.display = 'none'; return; }}

    const tStart = new Date(t0.getTime() + b * bm * 60000);
    const tEnd = new Date(tStart.getTime() + bm * 60000);
    const fmtT = d => {{
      let h = d.getHours(), m = d.getMinutes(), ap = 'am';
      if (h >= 12) {{ ap = 'pm'; if (h > 12) h -= 12; }}
      if (h === 0) h = 12;
      return h + ':' + (m < 10 ? '0' : '') + m + ap;
    }};
    const fmtD = d => (d.getMonth()+1) + '/' + d.getDate() + '/' + (d.getFullYear() % 100);
    const timeStr = `(${{bm}}m) ${{fmtT(tStart)}}–${{fmtT(tEnd)}} EDT, ${{fmtD(tStart)}}`;

    let html;
    /* irradiance row */
    if (hasIrrad && my < irradOffset) {{
      const idx = Math.min(b, data.irrad.length - 1);
      const v = data.irrad[idx];
      html = `<b>POA Irradiance</b><br>${{timeStr}}<br>` +
             (v === null || v === undefined ? 'No data' : `<b>${{v.toFixed(1)}} W/m²</b>`);
    }}
    /* average row */
    else if (my < irradOffset + AVG_H) {{
      const v = avg[b];
      html = `<b>Average of All Inverters</b><br>${{timeStr}}<br>`;
      if (v === null) html += 'No data';
      else if (isCF) html += `<b>Capacity Factor (CF): ${{v.toFixed(2)}}%</b>`;
      else html += `<b>Avg: ${{v.toFixed(2)}}% of rated</b>`;
    }}
    /* inverter row */
    else {{
      const rowIdx = Math.floor((my - irradOffset - AVG_H) / ROW_H);
      if (rowIdx < 0 || rowIdx >= nInv) {{ tip.style.display = 'none'; return; }}
      const cf = data.vals[rowIdx] ? data.vals[rowIdx][b] : null;
      const kw = data.kw && data.kw[rowIdx] ? data.kw[rowIdx][b] : null;
      const cap = data.caps ? data.caps[rowIdx] : 0;
      const name = data.invs[rowIdx];
      html = `<b>${{name}}</b><br>${{timeStr}}<br>`;
      if (cf === null && kw === null) {{
        html += '<span style="color:#bbb">No data</span>';
      }} else {{
        if (cf !== null) html += `<b>Capacity Factor (CF): ${{cf.toFixed(2)}}%</b><br>`;
        if (kw !== null) html += `${{kw.toFixed(2)}} kW actual`;
        if (cap > 0) html += ` / ${{Math.round(cap)}} kW capacity`;
      }}
      if (data.alerts[rowIdx] && data.alerts[rowIdx].length)
        html += '<br><span style="color:#ff9999;font-size:10px">' +
                data.alerts[rowIdx].slice(0, 3).join('<br>') + '</span>';
    }}
    tip.innerHTML = html; tip.style.display = 'block';
    tip.style.left = Math.min(e.clientX + 14, window.innerWidth - 360) + 'px';
    tip.style.top = (e.clientY + 14) + 'px';
  }};
  cv.onmouseleave = () => tip.style.display = 'none';
}}

function switchMode(sid) {{
  const sel = document.getElementById('sel' + sid);
  if (sel) drawStrip(sid, sel.value);
}}

document.querySelectorAll('details.site-acc').forEach(d => {{
  const cvEl = d.querySelector('canvas.strip-canvas');
  if (!cvEl) return;
  const sid = cvEl.id.slice(2);
  if (d.open) requestAnimationFrame(() => drawStrip(sid));
  d.addEventListener('toggle', () => {{ if (d.open) drawStrip(sid); }});
}});

function filterSites(q) {{
  q = q.toLowerCase();
  document.querySelectorAll('details.site-acc').forEach(d =>
    d.style.display = d.dataset.site.includes(q) ? '' : 'none');
}}
function setAll(open) {{
  document.querySelectorAll('details.site-acc').forEach(d => {{
    if (d.style.display === 'none') return;
    d.open = open;
  }});
}}
</script>
<footer>ae_alert_dashboard.py · data: AlsoEnergy PowerTrack alerthistory API · units per PowerTrack Chart Builder</footer>
</body>
</html>
"""


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=7)
    ap.add_argument("--offline", action="store_true",
                    help="use cached ae_alerts.xlsx instead of the live API")
    ap.add_argument("--out", default=str(HERE / "dashboards" / "alerts.html"))
    args = ap.parse_args()

    d_to = date.today()
    d_from = d_to - timedelta(days=args.days - 1)

    if args.offline:
        print(f"[offline] reading {ALERTS_XLSX.name} ...")
        alerts = fetch_alerts_offline(d_from, d_to)
        if not alerts:
            all_alerts = fetch_alerts_offline(date(2000, 1, 1), d_to)
            if all_alerts:
                maxd = max(a["start"].date() for a in all_alerts if a["start"])
                d_to, d_from = maxd, maxd - timedelta(days=args.days - 1)
                alerts = [a for a in all_alerts
                          if a["start"] and d_from <= a["start"].date() <= d_to]
                print(f"[offline] window shifted to cached data: {d_from} .. {d_to}")
    else:
        session = get_session()
        print(f"Fetching alerts {d_from} .. {d_to} ...")
        alerts = fetch_alerts_live(session, d_from, d_to)

    print(f"{len(alerts)} alerts pulled")
    alerts = enrich(alerts, d_to)

    filtered = [a for a in alerts if not (
        a["category"] in ("COMMS_LOW", "OTHER") or
        (a["is_resolved"] and a["duration_h"] < 0.084 and
         a["category"] in ("INVERTER_COMM", "TRACKER_COMM")))]

    days = day_range(d_from, d_to)
    hw = load_hardware()
    heatmap_data = build_inverter_heatmaps(filtered, hw, days)
    tracker_data = build_tracker_data(filtered, days)

    crit_cats = ("INVERTER_FAULT", "GRID", "TRACKER_FAULT", "METER")
    n_critical = sum(1 for a in filtered if a["category"] in
                     ("INVERTER_FAULT", "GRID", "METER"))
    n_tracker = sum(1 for a in filtered if a["category"].startswith("TRACKER"))
    n_open = sum(1 for a in filtered if not a["is_resolved"])
    sites_hit = {a["site_name"] for a in filtered}

    # ── site emails ───────────────────────────────────────────────────
    site_emails = load_site_emails()

    # ── diagnosis + recommender ───────────────────────────────────────
    try:
        from ae_diagnosis import diagnose_portfolio
        diags = diagnose_portfolio(filtered, hw)
    except Exception as e:
        print(f"[warn] diagnosis engine failed: {e}")
        diags = []
    by_site_diag = defaultdict(list)
    for d in diags:
        if d["site"] != "PORTFOLIO":
            by_site_diag[d["site"]].append(d)
    recommender_html = "".join(
        render_diag_card(d, i + 1, emails=site_emails.get(d["site"], []))
        for i, d in enumerate(diags[:10])) or \
        "<div class='card muted'>No diagnoses generated for this window.</div>"

    # ── AI summaries ───────────────────────────────────────────────────
    ai_sums = load_ai_summaries()
    site_keys = load_site_keys()
    # back-fill site keys from hw data
    for sn, hwd in hw.items():
        if sn not in site_keys and hwd.get("site_key"):
            site_keys[sn] = hwd["site_key"]

    # ── inverter production + irradiance fetch ──────────────────────
    prod_by_site = {}
    prod_bins_by_site = {}
    irrad_by_site = {}
    if not args.offline:
        print("Fetching inverter production (15-min CF) per site ...")
        for sn, hwd in hw.items():
            sk = site_keys.get(sn)
            inv_keys = hwd.get("inv_keys", {})
            if inv_keys and sk:
                prod, n_bins = fetch_inverter_production(
                    session, sk, inv_keys, d_from, d_to)
                if prod:
                    prod_by_site[sn] = prod
                    prod_bins_by_site[sn] = n_bins
                    print(f"  {sn}: {len(prod)} inverters, {n_bins} bins")
                time.sleep(SLEEP)
            wk = hwd.get("weather_keys", [])
            if wk and sk:
                vals = fetch_irradiance(session, sk, wk, d_from, d_to)
                if vals:
                    irrad_by_site[sn] = vals
                time.sleep(SLEEP)
        print(f"  production: {len(prod_by_site)} sites, irradiance: {len(irrad_by_site)} sites")
        if prod_by_site:
            save_prod_cache(prod_by_site, prod_bins_by_site, irrad_by_site, d_from, d_to)

    # Fall back to cached production data if live fetch got nothing
    if not prod_by_site:
        cached_prod, cached_bins, cached_irrad, c_from, c_to = load_prod_cache()
        if cached_prod:
            prod_by_site = cached_prod
            prod_bins_by_site = cached_bins
            if not irrad_by_site:
                irrad_by_site = cached_irrad

    # ── per-site strip heatmaps for EVERY site ────────────────────────
    alerts_by_site = defaultdict(list)
    for a in filtered:
        alerts_by_site[a["site_name"]].append(a)
    all_sites = sorted(set(hw.keys()) | set(heatmap_data.keys()))
    print(f"  all_sites: {len(all_sites)}, hw: {len(hw)}, heatmap_data: {len(heatmap_data)}")

    def hours(site, kind):
        return sum(c[kind] for invs in heatmap_data.get(site, {}).values()
                   for c in invs.values() if isinstance(c, dict))

    site_order = sorted(all_sites,
                        key=lambda s: (-hours(s, "fault"), -hours(s, "comm"), s))
    heatmaps_html = ""
    skipped = []
    for sid, site in enumerate(site_order):
        strip = build_strip_data(site, hw, d_from, d_to,
                                 production=prod_by_site.get(site),
                                 n_bins_api=prod_bins_by_site.get(site, 0))
        if not strip["invs"]:
            skipped.append(site)
            continue
        strip = apply_alerts_to_strips(strip, alerts_by_site.get(site, []), d_from)
        heatmaps_html += render_site_accordion(
            sid, site, strip, by_site_diag.get(site, []),
            hours(site, "fault"), hours(site, "comm"),
            ai_summary=ai_sums.get(site, ""),
            irrad_data=irrad_by_site.get(site),
            site_emails=site_emails.get(site))

    if skipped:
        print(f"  SKIPPED (no invs): {skipped}")

    trackers_html = "".join(
        render_tracker_html(s, tracker_data[s], days, f"trk{i}")
        for i, s in enumerate(sorted(
            tracker_data.keys(),
            key=lambda s: -sum(v['fault'] + v['comm']
                               for v in tracker_data[s]['per_day'].values())))) or \
        "<div class='card muted'>No tracker/TCU alerts in this window.</div>"

    from collections import Counter
    cat_counter = Counter(a["cat_label"] for a in filtered)
    day_all = [sum(1 for a in filtered if a["start"] and a["start"].date() == d)
               for d in days]
    day_crit = [sum(1 for a in filtered if a["start"] and a["start"].date() == d
                    and a["category"] in crit_cats) for d in days]

    chartjs = CHARTJS_FILE.read_text(encoding="utf-8") if CHARTJS_FILE.exists() else ""
    if not chartjs:
        print("[warn] assets/chart.umd.min.js missing — charts will not render offline")

    html = PAGE.format(
        chartjs=chartjs,
        days_n=args.days,
        window_label=f"{d_from.strftime('%b %d')} – {d_to.strftime('%b %d, %Y')}",
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        n_total=len(alerts), n_filtered=len(filtered),
        n_critical=n_critical, n_tracker=n_tracker, n_open=n_open,
        n_sites=len(sites_hit), n_sites_total=35,
        critical_table=render_critical_table(filtered),
        recommender=recommender_html,
        heatmaps=heatmaps_html,
        trackers=trackers_html,
        cat_labels=json.dumps([k for k, _ in cat_counter.most_common()]),
        cat_counts=json.dumps([v for _, v in cat_counter.most_common()]),
        day_labels=json.dumps([d.strftime("%m/%d") for d in days]),
        day_all=json.dumps(day_all), day_crit=json.dumps(day_crit),
    )

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html, encoding="utf-8")
    print(f"Wrote {out}  ({len(html)/1024:.0f} KB)")


if __name__ == "__main__":
    main()

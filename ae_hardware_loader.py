#!/usr/bin/env python3
"""
Pulls hardware inventory for every site in the portfolio using:
  GET /api/scriptsite/{siteKey}?lastChanged=1900-01-01T00:00:00.000Z

Writes ae_hardware.xlsx with:
  - Hardware Inventory  : one row per device across all sites
  - Archive Columns     : which telemetry fields each device logs
  - Function Code Map   : meaning of functionCode integers
  - Site Summary        : device count per site

USAGE
-----
    python ae_hardware_loader.py
    python ae_hardware_loader.py --output ae_hardware.xlsx
"""

import argparse
import shlex
import sys
import time

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------
# CONFIG
# -----------------------------------------------------------------------

CURL_FILE  = "alsoenergy_curl.txt"
API_BASE   = "https://apps.alsoenergy.com/api"
PORTFOLIO  = "C12941"
SLEEP      = 0.5

FUNCTION_CODE_MAP = {
    0:  "Site",
    1:  "Inverter",
    2:  "Production Meter",
    5:  "Weather / Pyranometer",
    6:  "DC Zone / String Monitor",
    10: "Datalogger",
    11: "Generic Device / Module",
    14: "Camera",
    19: "Consumption Meter",
    24: "Tracker Controller",
    28: "String Inverter Group",
    31: "Cellular Modem",
}

# -----------------------------------------------------------------------
# SESSION
# -----------------------------------------------------------------------

def build_session():
    from ae_auth import get_session
    return get_session()

# -----------------------------------------------------------------------
# FETCH
# -----------------------------------------------------------------------

def get_portfolio_sites(session):
    r = session.get(f"{API_BASE}/view/portfolio/{PORTFOLIO}", timeout=20)
    if r.status_code in (401, 403):
        sys.exit("Auth expired — re-copy a fresh cURL into alsoenergy_curl.txt")
    r.raise_for_status()
    return r.json().get("sites", [])


def get_site_hardware(session, site_key):
    url = f"{API_BASE}/scriptsite/{site_key}?lastChanged=1900-01-01T00:00:00.000Z"
    r = session.get(url, timeout=20)
    if not r.ok:
        print(f"  WARN {site_key}: HTTP {r.status_code}", file=sys.stderr)
        return None
    return r.json()

# -----------------------------------------------------------------------
# EXCEL
# -----------------------------------------------------------------------

DARK  = "1F4E79"
MID   = "2E75B6"
WHITE = "FFFFFF"
ALT   = "EBF3FB"


def hdr(cell, text, bg=DARK):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def val(cell, text, row_idx=0):
    cell.value = text
    if row_idx % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=ALT)
    cell.alignment = Alignment(vertical="top", wrap_text=False)


def autosize(ws, max_w=50):
    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, max_w)


def write_xlsx(path, all_hw, site_data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Hardware Inventory ───────────────────────────────────
    ws = wb.create_sheet("Hardware Inventory")
    cols = ["Site Key", "Site Name", "HW Key", "HW Name", "Function Code",
            "Device Type", "Hardware Status", "Sort Order", "Archive Cols Count"]
    for ci, c in enumerate(cols, 1):
        hdr(ws.cell(1, ci), c, MID)
    ws.freeze_panes = "A2"

    row = 2
    for rec in all_hw:
        alt = row % 2
        val(ws.cell(row, 1), rec["site_key"], alt)
        val(ws.cell(row, 2), rec["site_name"], alt)
        val(ws.cell(row, 3), rec["hw_key"], alt)
        val(ws.cell(row, 4), rec["hw_name"], alt)
        val(ws.cell(row, 5), rec["function_code"], alt)
        val(ws.cell(row, 6), rec["device_type"], alt)
        val(ws.cell(row, 7), rec["hardware_status"], alt)
        val(ws.cell(row, 8), rec["sort"], alt)
        val(ws.cell(row, 9), rec["archive_col_count"], alt)
        row += 1

    autosize(ws)

    # ── Sheet 2: Archive Columns (one row per device×column) ─────────
    ws2 = wb.create_sheet("Archive Columns")
    cols2 = ["Site Key", "Site Name", "HW Key", "HW Name", "Device Type", "Column Name"]
    for ci, c in enumerate(cols2, 1):
        hdr(ws2.cell(1, ci), c, MID)
    ws2.freeze_panes = "A2"

    row = 2
    for rec in all_hw:
        for col_name in rec["archive_columns"]:
            alt = row % 2
            val(ws2.cell(row, 1), rec["site_key"], alt)
            val(ws2.cell(row, 2), rec["site_name"], alt)
            val(ws2.cell(row, 3), rec["hw_key"], alt)
            val(ws2.cell(row, 4), rec["hw_name"], alt)
            val(ws2.cell(row, 5), rec["device_type"], alt)
            val(ws2.cell(row, 6), col_name, alt)
            row += 1

    autosize(ws2)

    # ── Sheet 3: Site Summary ─────────────────────────────────────────
    ws3 = wb.create_sheet("Site Summary")
    s_cols = ["Site Key", "Site Name", "Lat", "Lon", "Timezone",
              "Total Devices", "Inverters", "Meters", "Weather Stations",
              "Dataloggers", "String Groups", "Other"]
    for ci, c in enumerate(s_cols, 1):
        hdr(ws3.cell(1, ci), c, MID)
    ws3.freeze_panes = "A2"

    for ri, (sk, sd) in enumerate(site_data.items(), start=2):
        alt = ri % 2
        hw = sd.get("hardware", [])
        fc_counts = {}
        for h in hw:
            fc = h.get("functionCode", -1)
            fc_counts[fc] = fc_counts.get(fc, 0) + 1

        val(ws3.cell(ri, 1), sk, alt)
        val(ws3.cell(ri, 2), sd.get("name", ""), alt)
        val(ws3.cell(ri, 3), sd.get("latitude"), alt)
        val(ws3.cell(ri, 4), sd.get("longitude"), alt)
        val(ws3.cell(ri, 5), sd.get("timeZone", ""), alt)
        val(ws3.cell(ri, 6), len(hw), alt)
        val(ws3.cell(ri, 7), fc_counts.get(1, 0) + fc_counts.get(28, 0), alt)  # inverters
        val(ws3.cell(ri, 8), fc_counts.get(2, 0) + fc_counts.get(19, 0), alt)  # meters
        val(ws3.cell(ri, 9), fc_counts.get(5, 0), alt)
        val(ws3.cell(ri, 10), fc_counts.get(10, 0), alt)
        val(ws3.cell(ri, 11), fc_counts.get(28, 0), alt)
        other = sum(v for k, v in fc_counts.items() if k not in (1, 2, 5, 10, 19, 28))
        val(ws3.cell(ri, 12), other, alt)

    autosize(ws3)

    # ── Sheet 4: Function Code Reference ─────────────────────────────
    ws4 = wb.create_sheet("Function Code Map")
    hdr(ws4.cell(1, 1), "Function Code", MID)
    hdr(ws4.cell(1, 2), "Device Type", MID)
    for ri, (fc, desc) in enumerate(sorted(FUNCTION_CODE_MAP.items()), start=2):
        alt = ri % 2
        val(ws4.cell(ri, 1), fc, alt)
        val(ws4.cell(ri, 2), desc, alt)
    autosize(ws4, max_w=40)

    wb.save(path)
    print(f"Saved -> {path}")
    print(f"  {len(all_hw)} hardware rows across {len(site_data)} sites")

# -----------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", default="ae_hardware.xlsx")
    args = ap.parse_args()

    session = build_session()

    print("Fetching site list from portfolio...")
    sites = get_portfolio_sites(session)
    print(f"  {len(sites)} sites found")

    all_hw = []
    site_data = {}

    for i, site in enumerate(sites):
        sk = site["key"]
        sn = site.get("name", sk)
        print(f"  [{i+1:2}/{len(sites)}] {sk} {sn}")
        data = get_site_hardware(session, sk)
        if not data:
            continue
        site_data[sk] = data
        for hw in data.get("hardware", []):
            fc = hw.get("functionCode", -1)
            all_hw.append({
                "site_key":        sk,
                "site_name":       sn,
                "hw_key":          hw.get("key", ""),
                "hw_name":         hw.get("name", ""),
                "function_code":   fc,
                "device_type":     FUNCTION_CODE_MAP.get(fc, f"Unknown ({fc})"),
                "hardware_status": hw.get("hardwareStatus", ""),
                "sort":            hw.get("sort", ""),
                "archive_columns": hw.get("archiveColumns", []),
                "archive_col_count": len(hw.get("archiveColumns", [])),
            })
        time.sleep(SLEEP)

    write_xlsx(args.output, all_hw, site_data)

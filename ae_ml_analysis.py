#!/usr/bin/env python3
"""
NLP + correlation analysis on ae_alerts.xlsx.

Outputs ae_analysis.xlsx with:
  - Cluster Summary    : TF-IDF KMeans clusters of alert descriptions
  - Top Terms          : highest-TF-IDF words per cluster
  - Alerts by Site     : frequency, resolution rate, avg duration
  - Alerts by Type     : same, ranked by volume
  - Hourly Pattern     : alert counts by hour-of-day (UTC)
  - Numeric Extraction : values parsed from description text
  - Correlation Matrix : pearson r between numeric columns

USAGE
-----
    pip install pandas scikit-learn openpyxl matplotlib seaborn
    python ae_ml_analysis.py                          # default input/output
    python ae_ml_analysis.py --input ae_alerts.xlsx --output ae_analysis.xlsx
"""

import argparse
import re
import sys
import warnings
warnings.filterwarnings("ignore")

import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import io

# -----------------------------------------------------------------------
# CONFIG
# -----------------------------------------------------------------------
N_CLUSTERS   = 7     # KMeans k — tune this
MAX_FEATURES = 300   # TF-IDF vocabulary size
DARK_BLUE    = "1F4E79"
MID_BLUE     = "2E75B6"
LIGHT        = "D6E4F0"
WHITE_FG     = "FFFFFF"

# -----------------------------------------------------------------------
# LOAD
# -----------------------------------------------------------------------

def load_alerts(path):
    df = pd.read_excel(path)
    # normalise column names
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    # parse timestamps
    for col in ("alert_start", "alert_end", "resolved_time", "trigger_time"):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    # numeric coerce
    for col in ("severity", "impact", "capacity"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    # booleans
    for col in ("is_resolved", "is_acknowledged"):
        if col in df.columns:
            df[col] = df[col].map({"True": True, "False": False, True: True, False: False})
    # duration
    if "alert_start" in df.columns and "alert_end" in df.columns:
        df["duration_hr"] = (df["alert_end"] - df["alert_start"]).dt.total_seconds() / 3600
    return df

# -----------------------------------------------------------------------
# NLP
# -----------------------------------------------------------------------

STOPWORDS = {
    "a", "an", "the", "of", "in", "at", "to", "and", "or", "is", "are",
    "was", "were", "for", "with", "this", "that", "from", "by", "on", "be",
    "not", "no", "all", "has", "have", "had", "it", "its", "as", "so",
}

def clean_text(s):
    if pd.isna(s):
        return ""
    s = str(s).lower()
    # normalise whitespace and tabs
    s = re.sub(r"[\t]+", " ", s)
    # remove numbers-only tokens but keep "kwh", "kw", "amps" etc.
    s = re.sub(r"\b\d+(\.\d+)?\b", "NUM", s)
    # remove punctuation except letters/digits
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    # collapse spaces
    s = re.sub(r"\s+", " ", s).strip()
    return s


def run_nlp(df):
    texts = df["description"].fillna("").apply(clean_text).tolist()

    vec = TfidfVectorizer(
        max_features=MAX_FEATURES,
        ngram_range=(1, 2),
        stop_words=list(STOPWORDS),
        min_df=2,
    )
    X = vec.fit_transform(texts)

    km = KMeans(n_clusters=N_CLUSTERS, random_state=42, n_init=10)
    labels = km.fit_predict(X)
    df = df.copy()
    df["cluster"] = labels

    # top terms per cluster
    feature_names = vec.get_feature_names_out()
    cluster_terms = {}
    for cid in range(N_CLUSTERS):
        center = km.cluster_centers_[cid]
        top_idx = center.argsort()[::-1][:15]
        cluster_terms[cid] = [(feature_names[i], round(float(center[i]), 4)) for i in top_idx]

    return df, cluster_terms


def cluster_label(df, cluster_terms):
    """Auto-label each cluster from its top term."""
    labels = {}
    for cid, terms in cluster_terms.items():
        top = [t for t, _ in terms[:3]]
        labels[cid] = " / ".join(top)
    df["cluster_label"] = df["cluster"].map(labels)
    return df, labels

# -----------------------------------------------------------------------
# NUMERIC EXTRACTION from description text
# -----------------------------------------------------------------------

PATTERNS = {
    "phase_A_kw":    r"A\s*=\s*([-\d.]+)\s*kW",
    "phase_B_kw":    r"B\s*=\s*([-\d.]+)\s*kW",
    "phase_C_kw":    r"C\s*=\s*([-\d.]+)\s*kW",
    "current_A_amp": r"A\s*=\s*([\d.]+)\s*Amps?",
    "current_B_amp": r"B\s*=\s*([\d.]+)\s*Amps?",
    "current_C_amp": r"C\s*=\s*([\d.]+)\s*Amps?",
    "perf_index":    r"Performance index[:\s]+([\d.]+)",
    "module_temp_c": r"module\s+\d+:\s*([\d.]+)\s*°C",
    "num_errors":    r"Performance\s*\((\d+)\)",
    "failed_suites": r"(\d+)\s+test\s+suites?\s+failed",
}


def extract_numerics(df):
    out = df[["alert_id", "site_name", "event_type_name", "description"]].copy()
    for col, pat in PATTERNS.items():
        out[col] = df["description"].str.extract(pat, expand=False).astype(float)
    return out

# -----------------------------------------------------------------------
# CORRELATION
# -----------------------------------------------------------------------

def correlation_analysis(df):
    num_cols = [c for c in ["severity", "impact", "capacity", "duration_hr"] if c in df.columns]
    if not num_cols:
        return None, None
    sub = df[num_cols].dropna(how="all")
    corr = sub.corr()

    fig, ax = plt.subplots(figsize=(7, 5))
    sns.heatmap(corr, annot=True, fmt=".2f", cmap="RdYlGn", center=0,
                linewidths=0.5, ax=ax, vmin=-1, vmax=1)
    ax.set_title("Pearson Correlation — numeric alert fields", fontsize=11, pad=10)
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return corr, buf

# -----------------------------------------------------------------------
# AGGREGATIONS
# -----------------------------------------------------------------------

def agg_by(df, groupby_col, label):
    g = df.groupby(groupby_col).agg(
        alert_count=("alert_id", "count"),
        resolved_pct=("is_resolved", lambda x: round(100 * x.sum() / len(x), 1) if len(x) else 0),
        avg_severity=("severity", "mean"),
        avg_impact_kw=("impact", "mean"),
        avg_duration_hr=("duration_hr", "mean") if "duration_hr" in df.columns else ("alert_id", "count"),
    ).reset_index().sort_values("alert_count", ascending=False)
    g.columns = [label, "Alert Count", "Resolved %", "Avg Severity", "Avg Impact kW", "Avg Duration hr"]
    return g


def hourly_pattern(df):
    if "alert_start" not in df.columns:
        return pd.DataFrame()
    df2 = df.copy()
    df2["hour_utc"] = df2["alert_start"].dt.hour
    return df2.groupby("hour_utc").size().reset_index(name="alert_count")


def cluster_summary(df, labels):
    rows = []
    for cid, lbl in labels.items():
        sub = df[df["cluster"] == cid]
        rows.append({
            "Cluster": cid,
            "Label (top terms)": lbl,
            "Count": len(sub),
            "Top Event Types": ", ".join(sub["event_type_name"].dropna().value_counts().head(3).index.tolist()),
            "Top Sites": ", ".join(sub["site_name"].dropna().value_counts().head(3).index.tolist()),
            "Resolved %": round(100 * sub["is_resolved"].sum() / len(sub), 1) if "is_resolved" in sub else "",
            "Avg Severity": round(sub["severity"].mean(), 2) if "severity" in sub else "",
            "Sample Description": sub["description"].dropna().iloc[0][:120] if len(sub) else "",
        })
    return pd.DataFrame(rows)

# -----------------------------------------------------------------------
# EXCEL WRITER
# -----------------------------------------------------------------------

def hdr(cell, text, bg=DARK_BLUE, size=10):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE_FG, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_df(ws, df, start_row=2, alt_color="EBF3FB"):
    for ci, col in enumerate(df.columns, start=1):
        hdr(ws.cell(start_row - 1, ci), col, MID_BLUE)
    for ri, row in enumerate(df.itertuples(index=False), start=start_row):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(ri, ci)
            if pd.isna(val) if not isinstance(val, str) else False:
                c.value = ""
            else:
                c.value = val if not isinstance(val, float) else round(val, 3)
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor=alt_color)
            c.alignment = Alignment(vertical="top", wrap_text=True)
    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 55)
    ws.freeze_panes = ws.cell(start_row, 1)


def write_report(path, df, cluster_terms, labels, corr, corr_img):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 1. Cluster Summary ──────────────────────────────────────────────
    ws = wb.create_sheet("Cluster Summary")
    cs = cluster_summary(df, labels)
    write_df(ws, cs)

    # ── 2. Top Terms per Cluster ────────────────────────────────────────
    ws = wb.create_sheet("Top NLP Terms")
    rows = []
    for cid, terms in cluster_terms.items():
        for rank, (term, score) in enumerate(terms, 1):
            rows.append({"Cluster": cid, "Label": labels[cid], "Rank": rank, "Term": term, "TF-IDF Score": score})
    write_df(ws, pd.DataFrame(rows))

    # ── 3. Alerts by Site ────────────────────────────────────────────────
    ws = wb.create_sheet("By Site")
    write_df(ws, agg_by(df, "site_name", "Site"))

    # ── 4. Alerts by Event Type ─────────────────────────────────────────
    ws = wb.create_sheet("By Event Type")
    write_df(ws, agg_by(df, "event_type_name", "Event Type"))

    # ── 5. Hourly Pattern ───────────────────────────────────────────────
    ws = wb.create_sheet("Hourly Pattern")
    hp = hourly_pattern(df)
    write_df(ws, hp)
    if not hp.empty:
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.bar(hp["hour_utc"], hp["alert_count"], color="#2E75B6", edgecolor="white")
        ax.set_xlabel("Hour of day (UTC)")
        ax.set_ylabel("Alert count")
        ax.set_title("Alert frequency by hour of day (UTC)")
        ax.set_xticks(range(0, 24))
        plt.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        img = XLImage(buf)
        img.anchor = f"D2"
        ws.add_image(img)

    # ── 6. Numeric Extraction ────────────────────────────────────────────
    ws = wb.create_sheet("Numeric Extraction")
    ne = extract_numerics(df)
    write_df(ws, ne)

    # ── 7. Correlation Matrix ────────────────────────────────────────────
    ws = wb.create_sheet("Correlation Matrix")
    if corr is not None:
        write_df(ws, corr.reset_index().rename(columns={"index": "Field"}))
    if corr_img:
        img = XLImage(corr_img)
        img.anchor = f"H2"
        ws.add_image(img)

    # ── 8. Raw Data ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Raw + Clusters")
    cols = [c for c in df.columns if c not in ("description",)] + ["description"]
    write_df(ws, df[cols].head(1000))

    wb.save(path)
    print(f"Saved -> {path}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")


# -----------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default="ae_alerts.xlsx")
    ap.add_argument("--output", default="ae_analysis.xlsx")
    ap.add_argument("--clusters", type=int, default=N_CLUSTERS)
    args = ap.parse_args()
    N_CLUSTERS = args.clusters

    print(f"Loading {args.input} ...")
    df = load_alerts(args.input)
    print(f"  {len(df)} rows, columns: {list(df.columns)}")

    print(f"Running TF-IDF + KMeans (k={N_CLUSTERS}) on description text ...")
    df, cluster_terms = run_nlp(df)
    df, labels = cluster_label(df, cluster_terms)

    print("Computing correlations ...")
    corr, corr_img = correlation_analysis(df)

    print(f"Writing {args.output} ...")
    write_report(args.output, df, cluster_terms, labels, corr, corr_img)

    # Quick console summary
    print("\n--- Cluster sizes ---")
    for cid, lbl in sorted(labels.items()):
        n = (df["cluster"] == cid).sum()
        print(f"  [{cid}] {n:4d}  {lbl}")

    if "site_name" in df.columns:
        print("\n--- Top 5 sites by alert count ---")
        print(df["site_name"].value_counts().head(5).to_string())

    if "event_type_name" in df.columns:
        print("\n--- Top 5 event types ---")
        print(df["event_type_name"].value_counts().head(5).to_string())
